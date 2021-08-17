# -*- coding: utf-8 -*-
#---------Various Utilities function for Python--------------------------------------
from __future__ import division
from __future__ import print_function
from future import standard_library
standard_library.install_aliases()
from builtins import next
from builtins import map
from builtins import zip
from builtins import str
from builtins import range
from past.builtins import basestring
from past.utils import old_div
from builtins import object
import os, sys
#if sys.platform.find('win') > -1 :
#  from guidata import qthelpers  #Otherwise Erro with Spyder Save

import datetime, time, arrow,  shutil,  IPython, gc
import matplotlib.pyplot as plt
import numexpr as ne, numpy as np, pandas as pd, scipy as sci
import urllib3
from bs4 import BeautifulSoup
from numba import jit, float32

#####################################################################################
'''
PLOTLY_APIKEY=    'aBq2k5nrzLK0chV5T32x'
PLOTLY_USERNAME=  'kevinno'

import plotly.tools as tls
tls.set_credentials_file(username=PLOTLY_USERNAME, api_key=PLOTLY_APIKEY)
'''

PCLOC=  'win,asus1'  if os.path.expanduser('~').find('asus1') >-1  and sys.platform.find('win')>-1 else  'win,unerry' if sys.platform.find('win')> -1 else  'lin,gcloud' if  os.path.expanduser('~').find('noel') >-1 and  sys.platform.find('linux')> -1   else 'lin, virtualbox'
DIRCWD= 'D:/_devs/Python01/project27/' if PCLOC=='win,asus' else   'G:/_devs/project27/' if PCLOC=='win,unerry'  else  '/home/noel/project27/' if PCLOC=='lin,gcloud' else '/media/sf_projec27'

# DIRCWD=  'D:/_devs/project27/'  if  os.environ['COMPUTERNAME']=='ASUS1-PC' and sys.platform.find('win')> -1 else  'G:/_devs/project27/' if sys.platform.find('win')> -1   and  os.environ['COMPUTERNAME']=='KEVIN'   else  '/home/ubuntu/notebook/' if os.environ['HOME'].find('ubuntu')>-1 else '/media/sf_project27/'    
__path__= DIRCWD +'/aapackage/'

EC2CWD='/home/ubuntu/notebook/'




####################################################################################################

'''
https://console.developers.google.com/iam-admin/projects
Google Drive API
https://googledrive.github.io/PyDrive/docs/_build/html/quickstart.html#authentication

https://console.developers.google.com/iam-admin/quotas?project=refreshing-code-142020

Create an empty file called __init__.py in all you folders. Then you can import using . as a folder separator. Documentation here.


#--------------------Utilities---------------------------
# http://win32com.goermezer.de/content/blogsection/7/284/

WhenPackage is destroyed resintall and copy paste the data.
from saved package

ps://pythonconquerstheuniverse.wordpress.com/2009/10/15/python-packages/

import spyderlib
74	os.environ["SPYDER_PARENT_DIR"] = os.path.abspath(os.path.join(spyderlib.__file__, "../.."))
…	
199	from spyderlib.utils.programs import (run_python_script, is_module_installed,
200	                                      start_file, run_python_script_in_terminal)
201	from spyderlib.utils.iofuncs import load_session, save_session, reset_session


https://service.wi2.ne.jp/wi2net/SbjLogin/2/

klepto is built on top of dill, and can cache function inputs and outputs to memory (so that calculations don't need to be run twice), and it can seamlessly persist objects in the cache to disk or to a database.
The versions on github are the ones you want. https://github.com/uqfoundation/klepto or https://github.com/joblib/joblib. Klepto is newer, but has a much broader set of caching and archiving solutions than joblib. Joblib has been in production use longer, so it's better tested -- especially for parallel computing.
Here's an example of typical klepto workflow: https://github.com/uqfoundation/klepto/blob/master/tests/test_workflow.py
Here's another that has some numpy in it: https://github.com/uqfoundation/klepto/blob/master/tests/test_cache.py

'''

'''
Unpacking dictionaries

** unpacks dictionaries.

func(a=1, b=2, c=3)
is the same as

args = {'a': 1, 'b': 2, 'c':3}
func(**args)
It's useful if you have to construct parameters:

args = {'name': person.name}
if hasattr(person, "address"):
    args["address"] = person.address
func(**args)  # either expanded to func(name=person.name) or
              #                    func(name=person.name, address=person.address)
Packing parameters of a function

def setstyle(**styles):
    for key, value in styles.iteritems():      # styles is a regular dictionary
        setattr(someobject, key, value)
This lets you use the function like this:

setstyle(color="red", bold=False)


'''



'''
Consider the Bunch alternative:

class Bunch(object):
  def __init__(self, adict):
    self.__dict__.update(adict)
so if you have a dictionary d and want to access (read) its values with the syntax x.foo instead of the clumsier d['foo'], just do
'''


############## #Serialize Python Session ################################################
# https://github.com/uqfoundation/dill

class testclass(object):
   def __init__(self,x) :
       pass

   def z_autotest(self) :
     import sys, io
     # create file-like string to capture output
     codeOut = io.StringIO()
     codeErr = io.StringIO()

     code = """
         def f(x):
         x = x + 1
         return x
         print 'This is my output.'
     """

     # capture output and errors
     sys.stdout = codeOut
     sys.stderr = codeErr

     exec(code)

     # restore stdout and stderr
     sys.stdout = sys.__stdout__
     sys.stderr = sys.__stderr__

     print(f(4))
     s = codeErr.getvalue(); print("error:\n%s\n" % s)
     s = codeOut.getvalue(); print("output:\n%s" % s)
     codeOut.close()
     codeErr.close()



def session_load_function(name='test_20160815'):
 import dill
 n1= DIRCWD+ '/aaserialize/session/'+ name + '.pkl'




 dill.load_session(n1)
 print(n1)



def session_save_function(name='test'):
 import dill, picle
 t1= date_now()
 n1= DIRCWD+ '/aaserialize/session/dill_session_'+ name + '_'+t1+'.pkl'

 dill.dump_session(n1)
 print(n1)


def py_save_obj_dill(obj1, keyname) :
   import dill, pickle, numpy, pandas
   dir0, keyname= z_key_splitinto_dir_name(keyname)
   os_folder_create(DIRCWD+'/aaserialize/' + dir0)
   dir1= DIRCWD+'/aaserialize/' + dir0 + '/'+ keyname + '.pkl' if otherfolder==0 else keyname

   type_list= [numpy, pandas.core.series, dill.source.getmodule(int), dill.source.getmodule(str), dill.source.getmodule(float)]
   name_type= []

   type1= dill.source.getmodule(type(obj1))
   name1= ''
   if not name1 in type_list and not type1 in name_LIST :
     with open( dir1, 'wb') as f:
       dill.dumps(object1, protocol=pickle.HIGHEST_PROTOCOL)
   else :
       print('Primitive type not dill saved')
   return dir1


def session_spyder_showall():
  ls= os_file_listall(DIRCWD+ '/aaserialize/session/', pattern="session*.pkl", dirlevel=2)
  lsname= ls[:,0]
  for s in lsname:
    print(s)

def session_guispyder_save(filename):
 from spyderlib.utils.iofuncs import save_session
 save_session(filename+'.session.tar')

def session_guispyder_load(filename):
 from spyderlib.spyder import load_session
 load_session(filename+'.session.tar')
 '''
  spyderlib/spyder.lib
    #---- Sessions
    def load_session(self, filename=None):
        """Load session"""
        if filename is None:
            self.redirect_internalshell_stdio(False)
            filename, _selfilter = getopenfilename(self, _("Open session"),
                        getcwd(), _("Spyder sessions")+" (*.session.tar)")
            self.redirect_internalshell_stdio(True)
            if not filename:
                return
        if self.close():
            self.next_session_name = filename


 '''


def session_load(filename, dict1=None, towhere='main'):
  ''' .spydata file,  dict1: already provided Dict,  towhere= main, function, dict '''
  from spyderlib.utils.iofuncs import load_dictionary
  print('Loading : ', end=' ')
  if dict1 is None :
    if filename.find('.') == -1 : #Use default folder
        dir0, keyname= z_key_splitinto_dir_name(filename)
        os_folder_create(DIRCWD+ '/aaserialize/session/' + dir0)
        filename=DIRCWD + '/aaserialize/session/' + dir0 + '/' + keyname + '.spydata'
    print(filename)
    data = load_dictionary(filename)

  if towhere=='main' :  #Load in interpreter module
    module = sys.modules['__main__']
    for k, v in list(data[0].items()):
       setattr(module, k, v)
       print(k, end=' ')

  elif towhere=='function' :
      from inspect import getmembers, stack
      globals1 = dict(getmembers(stack()[1][0]))["f_globals"]

  elif towhere=='dict' :
    return data


def session_save(filename="/folder1/name1", globals1=None):
    '''Need to pass globals() Cannot Get Save data to .spydata file

   BIG issue with Import, Impor FULL MODULE ----> BIG ISSUE
     BIG ISSUE with DICT, USE LIST INSTEAD
        If you try to put this code in a module and import the function then you will have to pass globals() to the function explicitly as the globals() in the function is not the IPython global namespace. However, you can put the above code inside your ~/.ipython/profile_PROFILE/startup/startup.ipy file and it will work as expected.
       PROFILE is the name of the profile that you plan to start IPython with.
    '''
    #from spyderlib.utils.iofuncs import save_dictionary
    #from spyderlib.widgets.dicteditorutils import globalsfilter
    #from spyderlib.plugins.variableexplorer import VariableExplorer
    #from spyderlib.baseconfig import get_supported_types

    # Local imports

    import spyderlib.utils.iofuncs as iofunc
    import spyderlib.widgets.dicteditorutils as spyutil
    import spyderlib.plugins.variableexplorer as spyvar
    import spyderlib.baseconfig as spybas

    if globals1 is None :  #Extract Global from the caller, Be careful can have issue
      print('Error, Please 2nd Arguments    (...,globals()')
      return None
      #from inspect import getmembers, stack
      #globals1 = dict(getmembers(stack()[1][0]))["f_globals"]
    
    settings = spyvar.VariableExplorer.get_settings()
    data = spyutil.globalsfilter(globals1, check_all=True, filters=tuple(spybas.get_supported_types()['picklable']),
                         exclude_private=settings['exclude_private'],
                         exclude_uppercase=settings['exclude_uppercase'],
                         exclude_capitalized=settings['exclude_capitalized'],
                         exclude_unsupported=settings['exclude_unsupported'],
                         excluded_names=settings['excluded_names']+['settings','In'])

    if filename.find('.') == -1 : #Use default folder
        dir0, keyname= z_key_splitinto_dir_name(filename)
        os_folder_create(DIRCWD+ '/aaserialize/session/' + dir0)
        filename= DIRCWD+'/aaserialize/session/' + dir0 + '/'+ keyname + '.spydata'

    iofunc.save_dictionary(data,filename)
    os.chdir(DIRCWD)
    print('Saved: '+filename)


'''
#Used for Spyder Load /Session
def spyder_save_gui_session(name1) :
 path= DIRCWD+'/'+name1
 os.makedirs(path)

 savepath = path+'/'+name1+'.spydata'
 session_save(savepath)
 spyder_savesession( path+'/'+name1+'.session.tar')

 shutil.make_archive(name1, 'zip',  root_dir=path, base_dir=None)
 shutil.rmtree(path)

def spyder_load_gui_session(name1) :
 import zipfile
 path= DIRCWD
 path2= DIRCWD+'/'+name1
 #shutil.unpack_archive(path+'/'+name1+'.zip')
 
 with zipfile.ZipFile(path+'/'+name1+'.zip', 'r') as myzip:
  myzip.extractall(  path2)
 
 spyder_loadsession(path2 + '/'+name1+'.session.tar')
 session_load(path2 + '/' + name1 + '.spydata')
 shutil.rmtree(path2)
'''



#####################################################################################
 #-------- Python General---------------------------------------------------------
#  Import File
# runfile('D:/_devs/Python01/project27/stockMarket/google_intraday.py', wdir='D:/_devs/Python01/project27/stockMarket')

def aa_unicode_ascii_utf8_issue():
   '''Take All csv in a folder and provide Table, Column Schema, type

 METHOD FOR Unicode / ASCII issue
1. Decode early
Decode to <type 'unicode'> ASAP
df['PREF_NAME']=       df['PREF_NAME'].apply(to_unicode)

2. Unicode everywhere


3. Encode late
>>> f = open('/tmp/ivan_out.txt','w')
>>> f.write(ivan_uni.encode('utf-8'))

Important methods
s.decode(encoding)  <type 'str'> to <type 'unicode'>
u.encode(encoding)  <type 'unicode'> to <type 'str'>

http://farmdev.com/talks/unicode/

   '''


def isexist(a) :
 try:
   a; return True
 except NameError: 
  return False

def isfloat(x):
  try:    
   v=   False if x == np.inf else True
   float(x)
   return v
  except :    return False
      
def isint(x): return isinstance(x, ( int, np.int, np.int64, np.int32 ) )

def a_isanaconda():
 import sys; 
 txt= sys.version
 if txt.find('Continuum') > 0 : return True
 else: return False

if a_isanaconda() :  #Import the Packages
   pass
  

def a_run_ipython(cmd1):
 ''' Execute Ipython Command in python code
     run -i :  run including current interprete variable
 '''
 IPython.get_ipython().magic(cmd1)

def a_autoreload() :
 a_run_ipython("load_ext autoreload"); a_run_ipython("autoreload 2")

def a_get_platform():
   return "linux"


def a_start_log(id1='', folder='aaserialize/log/')  :
 a_run_ipython('logstart -o -r -t ' + folder + 'log' + str(id1) + '_' + a_get_platform() + '_' + date_now() + ' rotate')

def a_cleanmemory():
  import gc;  gc.collect() 

def a_module_codesample(module_str='pandas') :
  dir1= 'D:/_devs/Python01/project27/docs/'
  file1= dir1+'/' + module_str + '/codesample.py'
  txt= os_file_read(file1)
  os_gui_popup_show(txt)

def a_module_doc(module_str='pandas') :
  dir1= 'D:/_devs/Python01/project27/docs/'
  file1= dir1+'/' + module_str + '/doc.py'
  txt= os_file_read(file1)
  os_gui_popup_show(txt)

def a_module_generatedoc(module_str="pandas", fileout=''):
  ''' #  getmodule_doc("jedi", r"D:\_devs\Python01\aapackage\doc.txt")'''
  from automaton import codeanalysis as ca
  #pathout= DIRCWD+'/docs/'+ module1.__name__
  pathout= DIRCWD+'/docs/'+ module_str
  if not os.path.exists(pathout) : os.makedirs(pathout)
  ca.getmodule_doc(module_str, file2=pathout+'/signature.py')

def a_info_conda_jupyter():
 s= '''
 CONDA COMMAND :
  'roll back' an installation, downgrading
     conda list  --revisions
     conda install --revision [revision number]         #Roll back to this number

  conda update PackageName

  conda uninstall PackageName


 IPYPARALLEL :
    Enable tab in Jupyter:   ipcluster nbextension enable
    Disable :                ipcluster nbextension enable
    Start 3 clsuter          ipcluster start -n 3
        <Stop them           ipclsuter stop

 JUPYTER Notebook :
      jupyter notebook      :  Enable Jupyter Server
      Close : Ctrl + C
      %connect_info      : To get Connection Details of Ipytho notebook


 '''
 print(s)

def a_run_cmd(cmd1):os_process_run(cmd1, shell=True)
# cmd("ipconfig")

def a_help():
 str= """
  PYCHARM shortcut :
    Highlight the Text  +  Alt +G   : Google Search
    Atl+C :  Doc   Alt+X : Doc     Alt+W : Doc Internet
    Ctrl+W : Refactor Name
    Shift + Shift  : Search Everywhere
    TouchRigh of Altgr + TouchRigh of Altgr: Definition of method

  #Save Session       session_save('/kaggle/supermarket_02', globals())

  #Load Session       session_load('/kaggle/kaggle_supermarket_01')


 1) %load_ext autoreload     #Reload the packages
    %autoreload 2

 2) Install PIP:
     If Permission access,Exit Spyder,  Use CMD  
       D:\WinPython-64-2710\scripts>pip install scikit-learn --upgrade -
 
     !pip install NAME --upgrade --no-deps  #NO dependencies installed.     
     
     !pip install git+git://github.com/fchollet/keras.git --upgrade --no-deps
     !pip install https://github.com//pymc3/archive/master.zip  
     !pip install  https://github.com/jpype-py3 /zipball/master  or  /tarball/master   

     !pip install c:\this_is_here\numpy-1.10.4+mkl-cp27-cp27m-win_amd64.whl
     !pip freeze   to get the list of package isntall

     !pip install -h    Help on pip  

 3) Windows CMD: 
    a_run_cmd("ipconfig")

    A: [enter] Changes the default drive from C to A.
   cd \furniture\chairs.  Moves you to the directory  'FURNITURE'
   cd .. Moves you up one level in the path.
   cd \  Takes you back to the root directory (c: in this case).


 4) import sys  
    reload(sys)  
    sys.setdefaultencoding('utf8')
    sys.getdefaultencoding()

 5) 
    !pip install packagename
    The ! prefix is a short-hand for the %sc command to run a shell command.

    You can also use the !! prefix which is a short-hand for the %sx command to execute a shell command 
    and capture its output (saved into the _ variable by default).

  6) Compile a modified package
    1) Put into a folder
    2) Add the folder to the Python Path (Spyder Python path)
 
 
    3) Compile in Spyder using, (full directory)
       !!python D:\\_devs\\Python01\\project27\\scikit_learn\\sklearn\\setup.py install

    4) Project is built here:
      D:\_devs\Python01\project27\build\lib.win-amd64-2.7
      http://scikit-learn.org/dev/developers/contributing.html#git-repo
python setup.py develop

Convert Python 2 to Python 3
import lib2to3

!2to3 D:\_devs\Python01\project\aapackage\codeanalysis.py

D:\_devs\Python01\project\zjavajar

# os.system('cd D:\_app\visualstudio13\Common7\IDE') #Execute any command
# os.path.abspath('memo.txt') #get absolute path
# os.path.exists('memo.txt')
# os.path.exists('memo.txt')
# os.getenv('PATH')    #Get environ variable

# sitepackage= r'D:\_devs\Python01\WinPython-64-2710\python-2.7.10.amd64\Lib\site-packages/'




  """
 os_gui_popup_show(str)
 #print( str)

def a_info_system():
 import platform; print(platform.platform() + '\n')
 # conda install -c anaconda psutil=5.0.0
 # https://github.com/nicolargo/glances
 import sys; print(("Python", sys.version))
 a_info_packagelist()

def a_info_packagelist() :
 import pip
 installed_packages = pip.get_installed_distributions()
 installed_packages_list = sorted(["%s==%s" % (i.key, i.version) for i in installed_packages])
 for p in  installed_packages_list : print(p)
 # def aa_getlistofpackage() : np.__config__.show(); a_info_packagelist()

def a_get_pythonversion():    return sys.version_info[0]


#################### Printers ###################################################################
def print_object(vv, txt='')  :
 ''' #Print Object Table  '''
 print ("\n\n" + txt +"\n")
 sh= np.shape(vv)
 kkmax, iimax= sh[0], sh[1]
 for k in range(0, kkmax):
   aux= ""
   for i in range(0,iimax) :
     if vv[k,0] != None :
      aux+=  str(vv[k,i])    + "," 
      
   if vv[k,0] != None :   
       print (aux )

def print_object_tofile(vv, txt, file1='d:/regression_output.py')  :
 ''' #Print to file Object   Table   '''
 with open(file1, mode='a')  as file1 :
  file1.write("\n\n" + txt +"\n" )  
  sh= np.shape(vv)
  kkmax, iimax= sh[0], sh[1]
  for k in range(0, kkmax):
    aux= ""
    for i in range(0,iimax) :
     if vv[k,0] != None :
      aux+=  str(vv[k,i])    + "," 
      
    if vv[k,0] != None :   
       #print (aux )
       file1.write(aux + 	"\n")

def print_ProgressBar (iteration, total, prefix = '', suffix = '', decimals = 1, barLength = 100):
    """# Print iterations progress
     Call in a loop to create terminal progress bar
    @params:
        iteration   - Required  : current iteration (Int)
        total       - Required  : total iterations (Int)
        prefix      - Optional  : prefix string (Str)
        suffix      - Optional  : suffix string (Str)
        decimals    - Optional  : positive number of decimals in percent complete (Int)
        barLength   - Optional  : character length of bar (Int)
    """
    formatStr = "{0:." + str(decimals) + "f}"
    percent = formatStr.format(100 * (old_div(iteration, float(total))))
    filledLength = int(round(barLength * iteration / float(total)))
    bar = '█' * filledLength + '-' * (barLength - filledLength)
    sys.stdout.write('\r%s |%s| %s%s %s' % (prefix, bar, percent, '%', suffix)),
    if iteration == total:
        sys.stdout.write('\n')
    sys.stdout.flush()
    '''
    for item in items:
     # Do stuff...
     # Update Progress Bar
     i += 1
     printProgress(i, l, prefix = 'Progress:', suffix = 'Complete', barLength = 50)
    '''


###################### OS- ######################################################################
def os_zip_checkintegrity(filezip1):
  import zipfile
  zip_file = zipfile.ZipFile(filezip1)
  try :
    ret= zip_file.testzip()
    if ret is not None:
        print("First bad file in zip: %s" % ret)
        return  False
    else: return True
  except Exception as e : return False

def os_zipfile(folderin, folderzipname, iscompress=True):
  import os, zipfile
  compress = zipfile.ZIP_DEFLATED if iscompress else  zipfile.ZIP_STORED
  zf= zipfile.ZipFile(folderzipname, "w", compress, allowZip64=True)

  for dirname, subdirs, files in os.walk(folderin):
        zf.write(dirname)
        for filename in files:  zf.write(os.path.join(dirname, filename))
  zf.close()

  r= os_zip_checkintegrity(folderzipname)
  if r : return  folderin, folderzipname
  else :
     print('Corrupt File')
     return  folderin, False

def os_zipfolder(dir_tozip='/zdisks3/output', zipname='/zdisk3/output.zip', dir_prefix=None, iscompress=True):
 '''
 shutil.make_archive('/zdisks3/results/output', 'zip',
                     root_dir=/zdisks3/results/',
                     base_dir='output')

 os_zipfolder('zdisk/test/aapackage', 'zdisk/test/aapackage.zip', 'zdisk/test')'''

 dir_tozip= dir_tozip if dir_tozip[-1] != '/' else dir_tozip[:-1]
 # dir_prefix= dir_prefix if dir_prefix[-1] != '/' else dir_prefix[:-1]

 if dir_prefix is None :
   dir_tozip, dir_prefix= '/'.join(dir_tozip.split('/')[:-1]), dir_tozip.split('/')[-1]

 import shutil
 shutil.make_archive(zipname.replace('.zip',''), 'zip', dir_tozip, base_dir=dir_prefix)


 '''
 shutil.make_archive(base_name, format[, root_dir[, base_dir[, verbose[, dry_run[, owner[, group[, logger]]]]]]])
Create an archive file (eg. zip or tar) and returns its name.

base_name is the name of the file to create, including the path, minus any format-specific extension. format is the archive format: one of “zip” (if the zlib module or external zip executable is available), “tar”, “gztar” (if the zlib module is available), or “bztar” (if the bz2 module is available).

root_dir is a directory that will be the root directory of the archive; ie. we typically chdir into root_dir before creating the archive.

base_dir is the directory where we start archiving from; ie. base_dir will be the common prefix of all files and directories in the archive.

root_dir and base_dir both default to the current directory.

owner and group are used when creating a tar archive. By default, uses the current owner and group.

logger must be an object compatible with PEP 282, usually an instance of logging.Logger.
  '''

 '''
 import zipfile
 compress = zipfile.ZIP_DEFLATED if iscompress else  zipfile.ZIP_STORED
 zf= zipfile.ZipFile(zipname, "w", compress, allowZip64=True)
 for dirname, subdirs, files in os.walk(dir_tozip):
    zf.write(dirname)
    for filename in files:  zf.write(os.path.join(dirname, filename))
 zf.close()
 '''
 r= os_zip_checkintegrity(zipname)
 if r : return  zipname
 else :
    print('Corrupt File'); return False

def os_zipextractall(filezip_or_dir="folder1/*.zip", tofolderextract='zdisk/test', isprint=1):
   '''os_zipextractall( 'aapackage.zip','zdisk/test/'      )  '''
   import zipfile

   if filezip_or_dir.find("*") > - 1 :    #  Many Zip
      ziplist1= os_file_listall(filezip_or_dir[:filezip_or_dir.find("*")],'*.zip')
      fileziplist_full=   ziplist1[2]

   else :                                 # Only 1
    filename=    os_file_getname(filezip_or_dir)
    fileziplist_full= [filezip_or_dir]


   # if os.path.exists( foldernew2  ) :      #Either File or Folder exists
   #   print('Renaming Folder ' + foldernew + ' with _')
   #   # lastfolder, beforelast= tofolderextract.split('/')[]
   #   os_folder_copy(foldernew2,  foldernew2+'_' )


   for filezip in fileziplist_full :
     filezip_name=    os_file_getname(filezip)
     zip_ref = zipfile.ZipFile(filezip, 'r')
     zip_ref.extractall(tofolderextract)   #Will create the path
     zip_ref.close()
     isok= os.path.exists( tofolderextract  )
     if not isok : print('Error: ' + filezip_name)

   if isok : return  tofolderextract
   else :    return -1

def os_folder_copy(src, dst, symlinks=False, pattern1="*.py", fun_file_toignore=None):
   '''
       callable(src, names) -> ignored_names
       'src' parameter, which is the directory being visited by copytree(), and
       'names' which is the list of `src` contents, as returned by os.listdir():

    Since copytree() is called recursively, the callable will be called once for each directory that is copied.
    It returns a  list of names relative to the `src` directory that should not be copied.
   :param fun_ignore:
   '''
   import shutil, errno, fnmatch
   def fun_file_toignore(src, names) :
       pattern= '!' + pattern1
       file_toignore= fnmatch.filter(names, pattern)
       return file_toignore

   try:
        shutil.copytree(src, dst, symlinks=False, ignore=fun_file_toignore)
   except OSError as exc: # python >2.5
        if exc.errno == errno.ENOTDIR:
            shutil.copy(src, dst,symlinks=False, ignore=fun_file_toignore)
        else: raise

def os_folder_create(directory) :
   DIR0= os.getcwd()
   if not os.path.exists(directory): os.makedirs(directory)
   os.chdir(DIR0)

def os_folder_robocopy(from_folder='', to_folder='', my_log='H:/robocopy_log.txt'):
    """
    Copy files to working directory
    robocopy <Source> <Destination> [<File>[ ...]] [<Options>]
    We want to copy the files to a fast SSD drive
    """
    if os.path.isdir(from_folder) &  os.path.isdir(to_folder):
        subprocess.call(["robocopy", from_folder, to_folder, "/LOG:%s" % my_log])
    else:
        print("Paths not entered correctly")

def os_file_replacestring1(findStr, repStr, filePath):
    "replaces all findStr by repStr in file filePath"
    import fileinput  
    file1= fileinput.FileInput(filePath, inplace=True, backup='.bak')
    for line in file1:
       line= line.replace(findStr,  repStr)
       sys.stdout.write(line)
    file1.close()

    print("OK: "+format(filePath))

def os_file_replacestring2(findstr, replacestr, some_dir, pattern="*.*", dirlevel=1  ):
  ''' #fil_replacestring_files("logo.png", "logonew.png", r"D:/__Alpaca__details/aiportfolio",    pattern="*.html", dirlevel=5  )
  '''
  list_file= listallfile(some_dir, pattern=pattern, dirlevel=dirlevel)
  list_file= list_file[2]
  for file1 in list_file : fil_replacestring_onefile(findstr, replacestr, file1)

def os_file_getname(path):
    import ntpath
    head, tail = ntpath.split(path)
    return tail or ntpath.basename(head)

def os_file_getpath(path):
    import ntpath
    head, tail = ntpath.split(path)
    return head

def os_file_gettext(file1):
 with open(file1, 'r',encoding='UTF-8',) as f:      
  return f.read()
 #def os_file_listall(some_dir, pattern="*.*", dirlevel=1):
 #  return listallfile(some_dir, pattern=pattern, dirlevel=dirlevel)

def os_file_listall(dir1, pattern="*.*", dirlevel=1, onlyfolder=0):
  '''
   # DIRCWD=r"D:\_devs\Python01\project"
   # aa= listallfile(DIRCWD, "*.*", 2)
   # aa[0][30];   aa[2][30]

   :param dir1:
   :param pattern:
   :param dirlevel:
   :param onlyfolder:
   :return:
  '''
  import fnmatch; import os; import numpy as np;  matches = []
  dir1 = dir1.rstrip(os.path.sep)
  num_sep = dir1.count(os.path.sep)

  if onlyfolder :
   for root, dirs, files in os.walk(dir1):
    num_sep_this = root.count(os.path.sep)
    if num_sep + dirlevel <= num_sep_this: del dirs[:]
    matches.append([]); matches.append([]); matches.append([]);   # Filename, DirName
    for dirs in fnmatch.filter(dirs, pattern):
      matches[0].append(os.path.splitext(dirs)[0])
      matches[1].append(os.path.splitext(root)[0])
      matches[2].append(os.path.join(root, dirs))
   return np.array(matches)

  for root, dirs, files in os.walk(dir1):
    num_sep_this = root.count(os.path.sep)
    if num_sep + dirlevel <= num_sep_this: del dirs[:]
    matches.append([]); matches.append([]); matches.append([]);   # Filename, DirName
    for files in fnmatch.filter(files, pattern):
      matches[0].append(os.path.splitext(files)[0])   
      matches[1].append(os.path.splitext(files)[1])  
      matches[2].append(os.path.join(root, files))   
  return np.array(matches)

def os_file_rename(some_dir, pattern="*.*", pattern2="", dirlevel=1):
  import fnmatch; import os; import numpy as np; import re;  matches = []
  some_dir = some_dir.rstrip(os.path.sep)
  assert os.path.exists(some_dir)
  num_sep = some_dir.count(os.path.sep)
  for root, dirs, files in os.walk(some_dir):
    num_sep_this = root.count(os.path.sep)
    if num_sep + dirlevel <= num_sep_this: del dirs[:]
    matches.append([]); matches.append([]); matches.append([]);   # Filename, DirName
    for files in fnmatch.filter(files, pattern):
        
     #replace pattern by pattern2        
      nfile= re.sub(pattern, pattern2, files) 
      os.path.abspath(root)
      os.rename(files, nfile)
 
      matches[0].append(os.path.splitext(nfile)[0])   
      matches[1].append(os.path.splitext(nfile)[1])  
      matches[2].append(os.path.join(root, nfile))   
  return np.array(matches).T

def os_gui_popup_show(txt):
  def os_gui_popup_show2(txt):
    from tkinter import Tk, Scrollbar,  Text, mainloop, RIGHT, END, LEFT, Y
    root = Tk()
    S = Scrollbar(root)
    T = Text(root, height=50, width=90)
    S.pack(side=RIGHT, fill=Y)
    T.pack(side=LEFT, fill=Y)
    S.config(command=T.yview)
    T.config(yscrollcommand=S.set)
    T.insert(END, txt)
    root.attributes('-topmost', True) # note - before topmost
    mainloop(  )
  os_gui_popup_show2(txt)
  # import _thread
  # _thread.start_new_thread(os_gui_popup_show2, (txt,))  #issues

def os_print_tofile(vv, file1,  mode1='a'):  # print into a file='a
    '''
    Here is a list of the different modes of opening a file:
r
Opens a file for reading only. The file pointer is placed at the beginning of the file. This is the default mode.

rb

Opens a file for reading only in binary format. The file pointer is placed at the beginning of the file. This is the default mode.
r+

Opens a file for both reading and writing. The file pointer will be at the beginning of the file.
rb+

Opens a file for both reading and writing in binary format. The file pointer will be at the beginning of the file.
w

Opens a file for writing only. Overwrites the file if the file exists. If the file does not exist, creates a new file for writing.
wb

Opens a file for writing only in binary format. Overwrites the file if the file exists. If the file does not exist, creates a new file for writing.
w+

Opens a file for both writing and reading. Overwrites the existing file if the file exists. If the file does not exist, creates a new file for reading and writing.
wb+

Opens a file for both writing and reading in binary format. Overwrites the existing file if the file exists. If the file does not exist, creates a new file for reading and writing.
a

Opens a file for appending. The file pointer is at the end of the file if the file exists. That is, the file is in the append mode. If the file does not exist, it creates a new file for writing.
ab

Opens a file for appending in binary format. The file pointer is at the end of the file if the file exists. That is, the file is in the append mode. If the file does not exist, it creates a new file for writing.
a+

Opens a file for both appending and reading. The file pointer is at the end of the file if the file exists. The file opens in the append mode. If the file does not exist, it creates a new file for reading and writing.
ab+

Opens a file for both appending and reading in binary format. The file pointer is at the end of the file if the file exists. The file opens in the append mode. If the file does not exist, it creates a new file for reading and writing.
To open a text file, use:
fh = open("hello.txt", "r")

To read a text file, use:
print fh.read()

To read one line at a time, use:
print fh.readline()

To read a list of lines use:
print fh.readlines()

To write to a file, use:
fh = open("hello.txt", "w")
lines_of_text = ["a line of text", "another line of text", "a third line"]
fh.writelines(lines_of_text)
fh.close()

To append to file, use:
fh = open("Hello.txt", "a")
fh.close()

    '''
    with open(file1, mode1) as text_file:  text_file.write(str(vv))

def os_path_norm(pth): #Normalize path for Python directory
    ''' #r"D:\_devs\Python01\project\03-Connect_Java_CPP_Excel\PyBindGen\examples" '''
    if a_get_pythonversion()==2:
     ind = pth.find(":")
     if ind > -1 :
       a, b = pth[:ind], pth[ind + 1:].encode("string-escape").replace("\\x", "/")
       return "{}://{}".format(a, b.lstrip("\\//").replace("\\\\", "/"))        
     else: return pth
    else: 
      pth = pth.encode("unicode-escape").replace(b"\\x", b"/")
      return pth.replace(b"\\\\", b"/").decode("utf-8")

def os_path_change(path1): path1= normpath(path1); os.chdir(path1)    #Change Working directory path

def os_path_current(): return DIRCWD

def os_file_exist(file1): return os.path.exists(file1)

def os_file_size(file1): return os.path.getsize(file1)

def os_file_read(file1):
 fh = open(file1,"r")
 return fh.read()

def os_file_mergeall(nfile, dir1, pattern1, deepness=2):
 ll= listallfile(dir1,pattern1,deepness)
 with open(nfile, mode='a', encoding='UTF-8') as nfile1:
  txt=''; ii=0
  for l in ll[2]:
   txt= '\n\n\n\n' + gettext_fromfile(l)
   nfile1.write(txt) 
 nfile1.close()

def os_extracttext_allfile(nfile, dir1, pattern1="*.html", htmltag='p', deepness=2):
 ''' Extract text from html '''
 ll= listallfile(dir1, pattern1,5)
 with open(nfile, mode='a', encoding='UTF-8') as nfile1:
  txt=''; ii=0
  for l in ll[2]:
   page= gettext_fromfile(l)
   soup = BeautifulSoup(page, "lxml")
   txt2 = ' \n\n'.join([p.text for p in soup.find_all(htmltag)])
   
   txt= '\n\n\n\n' + txt2.strip()
   newfile1.write(txt) 
 nfile1.close()

def os_path_append(p1, p2=None, p3=None, p4=None):
 sys.path.append(p1)
 if p2 is not None:  sys.path.append(p2)
 if p3 is not None:  sys.path.append(p3)
 if p4 is not None:  sys.path.append(p4)



 #-----Get Documentation of the module

def os_split_dir_file(dirfile) :
    lkey= dirfile.split('/')
    if len(lkey)== 1 :dir1=""
    else :            dir1= '/'.join(lkey[:-1]); dirfile= lkey[-1]
    return dir1, dirfile

def os_process_run(cmd_list=['program', 'arg1', 'arg2'], capture_output=False):
    import subprocess
    #cmd_list= os_path_norm(cmd_list)
    PIPE=subprocess.PIPE
    STDOUT=subprocess.STDOUT
    proc = subprocess.Popen(cmd_list, stdout=PIPE, stderr=STDOUT, shell=False)  #Always put to False

    if capture_output :
     stdout, stderr = proc.communicate()
     err_code = proc.returncode
     print("Console Msg: \n")
     print(str(stdout)) #,"utf-8"))
     print("\nConsole Error: \n"+ str(stderr) )
     #    return stdout, stderr, int(err_code)


def os_process_2():
   pass
   '''
Ex: Dialog (2-way) with a Popen()
p = subprocess.Popen('Your Command Here',
                 stdout=subprocess.PIPE,
                 stderr=subprocess.STDOUT,
                 stdin=PIPE,
                 shell=True,
                 bufsize=0)
p.stdin.write('START\n')
out = p.stdout.readline()
while out:
  line = out
  line = line.rstrip("\n")

  if "WHATEVER1" in line:
      pr = 1
      p.stdin.write('DO 1\n')
      out = p.stdout.readline()
      continue

  if "WHATEVER2" in line:
      pr = 2
      p.stdin.write('DO 2\n')
      out = p.stdout.readline()
      continue

out = p.stdout.readline()
p.wait()
'''





#######################  Python Interpreter ###################################################
def py_importfromfile(modulename, dir1):
 #Import module from file:  (Although this has been deprecated in Python 3.4.)
 vv= a_get_pythonversion()
 if vv==3:
  from importlib.machinery import SourceFileLoader
  foo = SourceFileLoader("module.name", "/path/to/file.py").load_module()
  foo.MyClass()
 elif vv==2 :
  import imp
  foo = imp.load_source('module.name', '/path/to/file.py')
  foo.MyClass()

def py_memorysize(o, ids, hint=" deep_getsizeof(df_pd, set()) "):
    """ deep_getsizeof(df_pd, set())
    Find the memory footprint of a Python object
    The sys.getsizeof function does a shallow size of only. It counts each
    object inside a container as pointer only regardless of how big it
    """
    from collections import Mapping, Container
    from sys import getsizeof

    d = py_memorysize
    if id(o) in ids:
        return 0

    r = getsizeof(o)
    ids.add(id(o))

    if isinstance(o, str) or isinstance(0, str):
        r= r

    if isinstance(o, Mapping):
        r= r + sum(d(k, ids) + d(v, ids) for k, v in list(o.items()))

    if isinstance(o, Container):
        r=  r + sum(d(x, ids) for x in o)

    return r * 0.0000001


def save(obj, folder='/folder1/keyname', isabsolutpath=0 ) :
 return py_save_obj(obj, folder=folder, isabsolutpath=isabsolutpath)

def load(folder='/folder1/keyname', isabsolutpath=0 ) :
 return py_load_obj(folder=folder, isabsolutpath=isabsolutpath)

def save_test(folder='/folder1/keyname', isabsolutpath=0  ) :
 ztest=  py_load_obj(folder=folder, isabsolutpath=isabsolutpath)
 print('Load object Type:' + str(type(ztest)) )
 del ztest; gc.collect()


def py_save_obj(obj, folder='/folder1/keyname', isabsolutpath=0):
    import pickle
    if isabsolutpath==0  and folder.find('.pkl') == -1 : #Local Path
       dir0, keyname= z_key_splitinto_dir_name(folder)
       os_folder_create(DIRCWD+'/aaserialize/' + dir0)
       dir1= DIRCWD+'/aaserialize/' + dir0 + '/'+ keyname + '.pkl'
    else :
       dir1= folder

    with open( dir1, 'wb') as f:
        pickle.dump(obj, f, pickle.HIGHEST_PROTOCOL)
    return dir1
    
def py_load_obj(folder='/folder1/keyname', isabsolutpath=0, encoding1='utf-8'):
    '''def load_obj(name, encoding1='utf-8' ):
         with open('D:/_devs/Python01/aaserialize/' + name + '.pkl', 'rb') as f:
            return pickle.load(f, encoding=encoding1)
    '''
    import pickle
    if isabsolutpath==0 and folder.find('.pkl') == -1 :
       dir0, keyname= z_key_splitinto_dir_name(folder)
       os_folder_create(DIRCWD+'/aaserialize/' + dir0)
       dir1= DIRCWD+'/aaserialize/' + dir0 + '/'+ keyname + '.pkl'
    else :
       dir1= folder

    with open(dir1, 'rb') as f:
        return pickle.load(f)

def z_key_splitinto_dir_name(keyname) :
    lkey= keyname.split('/')
    if len(lkey)== 1 :dir1=""
    else :            dir1= '/'.join(lkey[:-1]); keyname= lkey[-1]
    return dir1, keyname




############# Object ########################################################################
def sql_getdate():
  pass

'''
solver1= load_obj('batch/solver_ELVIS_USD_9assets_3_regime_perf_20160906_059')

solver1.x
solver1.convergence
._calculate_population_energies
solver1.next()  #One generation evolution
solver1.solve()  #Solve the pronlem

solver1.population           # Get the list of population sol
solver1.population_energies  # Population energy

aux= solver1.population


solver1.next()  #One generation evolution


def np_runsolver(name1, niter):

pfi= copy.deepcopy( get_class_that_defined_method(solver1.func) )

obj_getclass_property(pfi) 

solver1.x



id : int
date:
time:
name: str
solver
xbest
fbest
population
population_energies
params1
params2
params3
params4
params5
details

storage_optim= np.empty((1000, 15), dtype= np.object)

 
  id= ['id','date', 'time','name','solver','xbest', 'fbest', 'population','pop_energies',
  'params1', 'params2', 'params3', 'params4', 'params5', 'details']      )                                 
                                  
def tostore('store/'+dbname, dbid, vv) :                                  
 storage_optim=  load_obj(dbname)
 
 storage_optim[0,0]= 0
 storage_optim[0,1]= date_now()
 storage_optim[0,2]= copy.deepcopy(solver1)

 copyfile(dbname)
 save_obj(storage_optim, 'store/'+dbname)


aux= (solver1.population ,  solver1.population_energies  )
save_obj(aux, 'batch/elvis_usd_9assets_3_regime_perf_best_population_01')


'''


############### Object Class Introspection ##################################################
def obj_getclass_of_method(meth):
    import inspect
    for cls in inspect.getmro(meth.__self__.__class__):
        if meth.__name__ in cls.__dict__: 
            return cls
    return None


def obj_getclass_property(pfi) :
 for property, value in list(vars(pfi).items()):
    print(property, ": ", value)



################# XML / HTML processing  ####################################################
'''
https://pypi.python.org/pypi/RapidXml/
http://pugixml.org/benchmark.html

'''



#####################################################################################
#-------- PDF processing ------------------------------------------------------------
def print_topdf() :
 import datetime
 import numpy as np
 from matplotlib.backends.backend_pdf import PdfPages
 import matplotlib.pyplot as plt

 # Create the PdfPages object to which we will save the pages:
 # The with statement makes sure that the PdfPages object is closed properly at
 # the end of the block, even if an Exception occurs.
 with PdfPages('multipage_pdf.pdf') as pdf:
    plt.figure(figsize=(3, 3))
    plt.plot(list(range(7)), [3, 1, 4, 1, 5, 9, 2], 'r-o')
    plt.title('Page One')
    pdf.savefig()  # saves the current figure into a pdf page
    plt.close()

    plt.rc('text', usetex=True)
    plt.figure(figsize=(8, 6))
    x = np.arange(0, 5, 0.1)
    plt.plot(x, np.sin(x), 'b-')
    plt.title('Page Two')
    pdf.savefig()
    plt.close()

    plt.rc('text', usetex=False)
    fig = plt.figure(figsize=(4, 5))
    plt.plot(x, x*x, 'ko')
    plt.title('Page Three')
    pdf.savefig(fig)  # or you can pass a Figure object to pdf.savefig
    plt.close()

    # We can also set the file's metadata via the PdfPages object:
    d = pdf.infodict()
    d['Title'] = 'Multipage PDF Example'
    d['Author'] = u'Jouni K. Sepp\xe4nen'
    d['Subject'] = 'How to create a multipage pdf file and set its metadata'
    d['Keywords'] = 'PdfPages multipage keywords author title subject'
    d['CreationDate'] = datetime.datetime(2009, 11, 13)
    d['ModDate'] = datetime.datetime.today()




#####################################################################################
#--------CSV processing -------------------------------------------------------------
#Put Excel and CSV into Database / Extract CSV from database

def os_config_setfile(dict_params, outfile, mode1='w+') :
  with open(outfile, mode=mode1) as f1 :
     for key, item in list(dict_params.items()) :
       if isinstance(item, str)  or isinstance(item, str) :
          f1.write( str(key) +'= ' + "'"+ str(item) + "'" + ' \n')
       else :
          f1.write( str(key) +'= ' + str(item) + ' \n')

  f1.close()
  print(outfile)


def os_config_getfile(file1) :
   with open(file1, mode='r')  as f1 :
      ll= f1.readlines()

   for x in ll :
      print(x, end=' ')


def os_csv_process(file1):
 print(  '''
import csv
with open(file2, 'r',encoding='UTF-8',) as f:      
  reader = csv.reader(f,  delimiter='/' )
  kanjidict=  dict(reader) 
  
 return kanjidict  

import csv
with csv.reader(open(file2, 'r',encoding='UTF-8'),  delimiter=',' ) as reader :
  for row in reader:

# pip install fastcsv

with fastcsv.Reader(io.open(CSV_FILE)) as reader:
    for row in reader:
        pass


with fastcsv.Writer(io.open(CSV_FILE, 'w', encoding='cp932')) as writer:
    writer.writerow(row)

import csv
FUNDING = 'data/funding.csv'

def read_funding_data(path):
    with open(path, 'rU') as data:
        reader = csv.DictReader(data)
        for row in reader:
            yield row

if __name__ == "__main__":
    for idx, row in enumerate(read_funding_data(FUNDING)):
        if idx > 10: break
        print "%(company)s (%(numEmps)s employees) raised %(raisedAmt)s on %(fundedDate)s" % row

A couple of key points with the code above:
Always wrap the CSV reader in a function that returns a generator (via the yield statement).
Open the file in universal newline mode with 'rU' for backwards compatibility.
Use context managers with [callable] as [name] to ensure that the handle to the file is closed automatically.
Use the csv.DictReader class only when headers are present, otherwise just use csv.reader. (You can pass a list of fieldnames, but you'll see its better just to use a namedtuple as we discuss below).
This code allows you to treat the data source as just another iterator or list in your code. In fact if you do the following:

data = read_funding_data(FUNDING)
print repr(data)

Encoding reading :
These encoders translate the native file encoding to UTF-8, which the csv module can read because it's 8-bit safe.
try:
    import unicodecsv as csv
except ImportError:
    import warnings
    warnings.warn("can't import `unicodecsv` encoding errors may occur")
    import csv

pip install unicodecsv

from collections import namedtuple

fields = ("permalink","company","numEmps", "category","city","state","fundedDate", "raisedAmt","raisedCurrency","round")
FundingRecord = namedtuple('FundingRecord', fields)

def read_funding_data(path):
    with open(path, 'rU') as data:
        data.readline()            # Skip the header
        reader = csv.reader(data)  # Create a regular tuple reader
        for row in map(FundingRecord._make, reader):
            yield row

if __name__ == "__main__":
    for row in read_funding_data(FUNDING):
        print row
        break

https://districtdatalabs.silvrback.com/simple-csv-data-wrangling-with-python

from datetime import datetime

fields = ("permalink","company","numEmps", "category","city","state","fundedDate", "raisedAmt","raisedCurrency","round")

class FundingRecord(namedtuple('FundingRecord_', fields)):

    @classmethod
    def parse(klass, row):
        row = list(row)                                # Make row mutable
        row[2] = int(row[2]) if row[2] else None       # Convert "numEmps" to an integer
        row[6] = datetime.strptime(row[6], "%d-%b-%y") # Parse the "fundedDate"
        row[7] = int(row[7])                           # Convert "raisedAmt" to an integer
        return klass(*row)

    def __str__(self):
        date = self.fundedDate.strftime("%d %b, %Y")
        return "%s raised %i in round %s on %s" % (self.company, self.raisedAmt, self.round, date)

def read_funding_data(path):
    with open(path, 'rU') as data:
        data.readline()            # Skip the header
        reader = csv.reader(data)  # Create a regular tuple reader
        for row in map(FundingRecord.parse, reader):
            yield row

if __name__ == "__main__":
    for row in read_funding_data(FUNDING):
        print row
        break
''')


def pd_toexcel(df, outfile='file.xlsx', sheet_name='sheet1', append=1, returnfile=1):
   '''
# Create a Pandas Excel writer using XlsxWriter as the engine.
writer = pd.ExcelWriter('test.xlsx', engine='xlsxwriter')
df.to_excel(writer, sheet_name='Sheet1')
writer.save()

# Get the xlsxwriter objects from the dataframe writer object.
workbook  = writer.book
worksheet = writer.sheets['Sheet1']

# Add some cell formats.
format1 = workbook.add_format({'num_format': '#,##0.00'})
format2 = workbook.add_format({'num_format': '0%'})
format3 = workbook.add_format({'num_format': 'h:mm:ss AM/PM'})

# Set the column width and format.
worksheet.set_column('B:B', 18, format1)

# Set the format but not the column width.
worksheet.set_column('C:C', None, format2)

worksheet.set_column('D:D', 16, format3)

# Close the Pandas Excel writer and output the Excel file.
writer.save()

from openpyxl import load_workbook
wb = load_workbook(outfile)
ws = wb.active
ws.title = 'Table 1'

tableshape = np.shape(table)
alph = list(string.ascii_uppercase)

for i in range(tableshape[0]):
    for j in range(tableshape[1]):
        ws[alph[i]+str(j+1)] = table[i, j]

for cell in ws['A'] + ws[1]:
    cell.style = 'Pandas'

wb.save('Scores.xlsx')

   '''

   if append and os_file_exist(outfile) :
     from openpyxl import load_workbook
     book = load_workbook(outfile)
     writer = pd.ExcelWriter(outfile, engine='openpyxl')
     writer.book = book
     writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
     df.to_excel(writer, sheet_name=sheet_name)
     writer.save()
     if returnfile : return outfile
   else :
     writer = pd.ExcelWriter(outfile, engine='xlsxwriter')
     df.to_excel(writer, sheet_name=sheet_name, )
     writer.save()
     if returnfile : return outfile

def pd_toexcel_many(outfile='file1.xlsx', df1=None, df2=None, df3=None, df4=None, df5=None, df6=None) :
  pd_toexcel(df1, outfile,  sheet_name='df1')
  if df2 is not None :   pd_toexcel(df2, outfile,  sheet_name='df2')
  if df3 is not None :   pd_toexcel(df3, outfile,  sheet_name='df3')
  if df4 is not None :   pd_toexcel(df4, outfile,  sheet_name='df4')
  if df5 is not None :   pd_toexcel(df5, outfile,  sheet_name='df5')
  if df6 is not None :   pd_toexcel(df6, outfile,  sheet_name='df6')


  
def pd_df_tojson(df,  tname='table1', format='column'):
   ''' {'table1': {"index": [ 0, 1 ,3 ],
                 "col1" : [ "250", "1" ,"3" ] }
       }
   '''
   import json
   json1= {} 
   json1[tname]= df.to_dict(orient='list')
   return json.dumps(json1)


  

############# STRING- ########################################################################
def find_fuzzy(xstring, list_string):
 ''' if xstring matches partially, add to the list   '''
 return [ xi for xi in list_string if xi.find(xstring) > -1 ]


def str_match_fuzzy(xstring, list_string):
 ''' if any of list_strinf elt matches partially xstring '''
 for  xi in list_string:
   if xstring.find(xi) > -1:  return True
 return False


def str_parse_stringcalendar(cal):
    '''----------Parse Calendar  --------'''
    cal2 = cal.split('\n')
    cal3 = [x for x in cal2 if x != '']
    import dateutil
    cal4 = []
    for x in cal3:
        i0 = x.find(':')
        cal4.append([datetime_toint(dateutil.parser.parse(x[:i0])), x[i0 + 1:]])
    cal4 = np.array(cal4)
    cal5 = np.array(sortcol(cal4, 0), dtype=str)
    for x in cal5: print(x[0], ":", x[1].strip())
    return cal5

def str_make_unicode(input, errors='replace'):
    ttype= type(input) 
    if ttype != str  : 
        input =  input.decode('utf-8', errors=errors);  return input
    else:  return input

def str_empty_string_array(x,y=1):
 if y==1:  return ["" for x in range(x)]
 else: return [["" for row in range(0,x)] for col in range(0,y)]

def str_empty_string_array_numpy(nx, ny=1):
 arr = np.empty((nx, ny), dtype=object)
 arr[:, :] = ''
 return arr

def str_isfloat(value):
  try:    float(value);    return True
  except :    return False

def str_is_azchar(x):
  try:    float(x);    return True
  except :    return False  
    
def str_is_az09char(x):
  try:    float(x);    return True
  except :    return False  

def str_reindent(s, numSpaces): #change indentation of multine string
    '''
   if args:
       aux= name1+'.'+obj.__name__ +'('+ str(args) +')  \n' + str(inspect.getdoc(obj))
       aux= aux.replace('\n', '\n       ')
       aux= aux.rstrip()
       aux= aux + ' \n'
       wi( aux)
    '''
    s = string.split(s, '\n')
    s = [(numSpaces * ' ') + string.lstrip(line) for line in s]
    s = string.join(s, '\n')
    return s

def str_split2(delimiters, string, maxsplit=0):  #Split into Sub-Sentence
    import re
    regexPattern = '|'.join(map(re.escape, delimiters))
    return re.split(regexPattern, string, maxsplit)

def str_split_pattern(sep2, ll, maxsplit=0):  #Find Sentence Pattern
 import re
 regexPat = '|'.join(sep2) 
 regex3 = re.compile( '(' + regexPat + r')|(?:(?!'+ regexPat +').)*', re.S)
 #re.compile(r'(word1|word2|word3)|(?:(?!word1|word2|word3).)*', re.S)
 ll= regex3.sub(lambda m: m.group(1) if m.group(1) else "P", ll)
 return ll

def pd_str_isascii(x):
  try :
    x.decode('ascii'); return True
  except: return false


def str_to_utf8(x):
  """ Do it before saving/output to external printer """
  return x.encode('utf-8')


def str_to_unicode(x, encoding='utf-8'):
  ''' Do it First after Loading some text '''
  if isinstance(x, basestring):
    if not isinstance(x, str): return str(x, encoding)
  else : return x


'''
http://www.ibm.com/developerworks/library/l-pyint/

spam.upper().lower()

spam.strip()
spam.lstrip()
spam.rstrip()  Performs both lstrip() and rstrip() on string

count(str, beg= 0,end=len(string))  Counts how many times str occurs in string or in a substring of
find(str, beg=0 end=len(string))  str occurs in string or in a substring of strindex if found  -1 otherwise.
replace(old, new [, max])   Replaces all occurrences of old in string with new or at most max occurrences if max given.

isdecimal()   Returns true if a unicode string contains only decimal characters and false otherwise.
isalnum()  Returns true if string has at least 1 character and all characters are alphanumeric and false otherwise.
salpha()   Returns true if string has at least 1 character and all characters are alphabetic and false otherwise.
isdigit()  Returns true if string contains only digits and false otherwise.
islower()  Returns true if string has at least 1 cased character and all cased characters are in lowercase and false otherwise.
isnumeric() Returns true if a unicode string contains only numeric characters and false otherwise.
isspace()  Returns true if string contains only whitespace characters and false otherwise.
istitle()  Returns true if string is properly "titlecased" and false otherwise.

lower()   Converts all uppercase letters in string to lowercase.
capitalize()   Capitalizes first letter of string
center(width, fillchar)  Returns space-padded string with string centered  w
title()  Returns "titlecased" version of string, that is, all words begin with uppercase and the rest are lowercase.
upper()   Converts lowercase letters in string to uppercase.

join(seq)  Merges (concatenates) the string representations of elements in sequence seq into a string, with separator string.
	
split(str="", num=string.count(str))   Splits string according to delimiter str  and returns list of substrings;
splitlines( num=string.count('\n'))  Splits string at all (or num) NEWLINEs and returns a list of each line with NEWLINEs removed.

ljust(width[, fillchar])   Returns a space-padded string with the original string left-justified to a total of width columns.

maketrans()  Returns a translation table to be used in translate function.

index(str, beg=0, end=len(string))   Same as find(), but raises an exception if str not found.
rfind(str, beg=0,end=len(string))   Same as find(), but search backwards in string.

rindex( str, beg=0, end=len(string))   Same as index(), but search backwards in string.

rjust(width,[, fillchar])  Returns a space-padded string with string right-jus to a total of width columns.

startswith(str, beg=0,end=len(string))
Determines if string or a substring of string (if starting index beg and ending index end are given) starts with substring str; returns true if so and false otherwise.

decode(encoding='UTF-8',errors='strict')
Decodes the string using the codec registered for encoding. encoding defaults to the default string encoding.
	
encode(encoding='UTF-8',errors='strict')
Returns encoded string version of string; on error, default is to raise a ValueError

endswith(suffix, beg=0, end=len(string))
Determines if string or a substring of string  ends with suffix; returns true if so and false otherwise.

expandtabs(tabsize=8)
Expands tabs in string to multiple spaces; defaults to 8 spaces per tab if tabsize not provided.
	
zfill (width)
Returns original string leftpadded with zeros to a total of width characters; intended for numbers, zfill() retains any sign given (less one zero).

'''



##############Internet data connect- #################################################################
'''
https://moz.com/devblog/benchmarking-python-content-extraction-algorithms-dragnet-readability-goose-and-eatiht/
pip install numpy
pip install --upgrade cython
!pip install lxml
!pip install libxml2
!pip install dragnet
https://pypi.python.org/pypi/dragnet

Only Python 2.7
!pip install goose-extractor

https://github.com/grangier/python-goose

from goose import Goose
url = 'http://edition.cnn.com/2012/02/22/world/europe/uk-occupy-london/index.html?hpt=ieu_c2'
g = Goose()
article = g.extract(url=url)
 
article.title

article.meta_description

article.cleaned_text[:150]

'''

def web_restapi_toresp(apiurl1):
 import requests
 resp = requests.get(apiurl1)
 if resp.status_code != 200:     # This means something went wrong.
    raise ApiError('GET /tasks_folder/ {}'.format(resp.status_code))
 return resp
 
def web_getrawhtml(url1) :
 import requests
 resp = requests.get(url1)
 if resp.status_code != 200:  # This means something went wrong.
    raise ApiError('GET /tasks_folder/ {}'.format(resp.status_code))
 else:
    return resp.text

def web_importio_todataframe(apiurl1, isurl=1):
 import requests
 if isurl : 
   resp = requests.get(apiurl1)
   if resp.status_code != 200:     # This means something went wrong.
    raise ApiError('GET /tasks_folder/ {}'.format(resp.status_code))
 au= resp.json()
 txt= au['extractorData']['data'][0]['group']
 colname=[]; i=-1
 for row in txt :
   i+=1;
   if i==1: break;
   for key, value in list(row.items()):
     if i==0:  colname.append( str(key))
 colname= np.array(colname); colmax=len(colname)
 
 dictlist= np.empty((5000, colmax), dtype=np.object); i=-1
 for row in txt :
   j=0; i+=1
   for key, value in list(row.items()):
     dictlist[i,j]= str(value[0]['text'])
     j+=1

 dictlist= dictlist[0:i+1,:]
 df= pd_createdf(dictlist, col1=colname, idx1= np.arange(0, len(dictlist)))
 return df

def web_getjson_fromurl(url):
 import json
 http = urllib3.connection_from_url(url) 
 jsonurl = http.urlopen('GET',url)
 
 # soup = BeautifulSoup(page)
 print(jsonurl)
 data = json.loads(jsonurl.read())

 return data



 # return the title and the text of the article at the specified url

def web_gettext_fromurl(url, htmltag='p'):
 http = urllib3.connection_from_url(url) 
 page = http.urlopen('GET',url).data.decode('utf8')
 
 soup = BeautifulSoup(page, "lxml")
 text = ' \n\n'.join([p.text for p in soup.find_all('p')])
 return soup.title.text + "\n\n" + text

def web_gettext_fromhtml(file1, htmltag='p'):
 with open(file1, 'r',encoding='UTF-8',) as f:      
   page=f.read()
 
 soup = BeautifulSoup(page, "lxml")
 text = ' \n\n'.join([p.text for p in soup.find_all(htmltag)])
 return soup.title.text + "\n\n" + text




'''
I know its been said already, 
but I'd highly recommend the Requests python package
: http://docs.python-requests.org/en/latest/index.html

If you've used languages other than python, 
you're probably thinking urllib and urllib2 are easy to use, 
not much code, and highly capable, that's how I used to think. 
But the Requests package is so unbelievably useful and 
short that everyone should be using it.

First, it supports a fully restful API, and is as easy as:

import requests
...

resp = requests.get('http://www.mywebsite.com/user')
resp = requests.post('http://www.mywebsite.com/user')
resp = requests.put('http://www.mywebsite.com/user/put')
resp = requests.delete('http://www.mywebsite.com/user/delete')
Regardless of whether GET/POST you never have to encode parameters again, it simply takes a dictionary as an argument and is good to go.

userdata = {"firstname": "John", "lastname": "Doe", "password": "jdoe123"}
resp = requests.post('http://www.mywebsite.com/user', params=userdata)
Plus it even has a built in json decoder (again, i know json.loads() isn't a lot more to write, but this sure is convenient):

resp.json()
Or if your response data is just text, use:

resp.text
This is just the tip of the iceberg. This is the list of features from the requests site:

International Domains and URLs
Keep-Alive & Connection Pooling
Sessions with Cookie Persistence
Browser-style SSL Verification
Basic/Digest Authentication
Elegant Key/Value Cookies
Automatic Decompression
Unicode Response Bodies
Multipart File Uploads
Connection Timeouts
.netrc support
List item
Python 2.6—3.4
Thread-safe.
'''

def web_getlink_fromurl(url):
 http = urllib3.connection_from_url(url) 
 page = http.urlopen('GET',url).data.decode('utf8')
 soup = BeautifulSoup(page, "lxml")    
 soup.prettify()
 links=[]
 for anchor in soup.findAll('a', href=True):
    lnk= anchor['href']
    links.append(  anchor['href'])
    
 return set(links)

def web_send_email(FROM, recipient, subject, body, login1="mizenjapan@gmail.com", pss1="sophieelise237", server1="smtp.gmail.com", port1=465):
    '''  # send_email("Kevin", "brookm291@gmail.com", "JapaneseText:" , "txt") '''
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
#    TO = recipient if type(recipient) is list else [recipient]
    TO= recipient
    msg = MIMEMultipart("alternative");    msg.set_charset("utf-8")
    msg["Subject"] = subject
    msg["From"] = FROM
    msg["To"] = TO
    part2 = MIMEText(body, "plain", "utf-8")
    msg.attach(part2)
    
    try:   # SMTP_SSL Example
        server_ssl = smtplib.SMTP_SSL( server1, port1)
        server_ssl.ehlo() # optional, called by login()
        server_ssl.login(login1, pss1)
        server_ssl.sendmail(FROM, [TO], msg.as_string())
        server_ssl.close();        print ('successfully sent the mail'  )       
        return 1
    except:
        print( "failed to send mail")
        return -1

def web_send_email_tls(FROM, recipient, subject, body, login1="mizenjapan@gmail.com", pss1="sophieelise237",
                   server1="smtp.gmail.com", port1=465):
    # send_email("Kevin", "brookm291@gmail.com", "JapaneseText:" , "txt")
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    #    TO = recipient if type(recipient) is list else [recipient]
    TO = recipient
    msg = MIMEMultipart("alternative");
    msg.set_charset("utf-8")
    msg["Subject"] = subject
    msg["From"] = FROM
    msg["To"] = TO
    part2 = MIMEText(body, "plain", "utf-8")
    msg.attach(part2)

    try:  # SMTP_SSL Example
        mailserver = smtplib.SMTP(server1, port1)
        # identify ourselves to smtp gmail ec2
        mailserver.ehlo()
        # secure our email with tls encryption
        mailserver.starttls()
        # re-identify ourselves as an encrypted connection
        mailserver.ehlo()
        mailserver.login(login1, pss1)

        mailserver.sendmail(FROM, [TO], msg.as_string())
        mailserver.quit()

        print ('successfully sent the mail')
        return 1
    except:
        print("failed to send mail")
        return -1


def web_sendurl(url1):
 # Send Text by email
 mm= web_gettext_fromurl(url1)   
 send_email("Python", "brookm291@gmail.com", mm[0:30] , url1 + '\n\n'+ mm )    



################### LIST UTIL / Array  ############################################################
def np_minimize(fun_obj, x0=[0.0], argext=(0,0), bounds1=[(0.03, 0.20), (10, 150)], method='Powell' ) :
   def penalty(vv):      #Constraints Penalty
     penalty=0.0  
     for i,x in enumerate(vv) :
       penalty+= 5000000*( -min(x-bounds1[i][0], 0.0) + max(x-bounds1[i][1],0.0))   
     return   penalty

   def loss(vv, argext):
     return fun_obj(vv,argext) + penalty(vv)

   res= sci.optimize.minimize(loss, x0, args=argext, method=method,   bounds=bounds1, tol= 0.001)
   return res

def np_minimizeDE(fun_obj, bounds, name1, solver=None):
  solver= sci.optimize._differentialevolution.DifferentialEvolutionSolver(fun_obj, bounds=bounds, popsize=popsize) 
  imin=0      
      
  name1= '/batch/solver_'+ name1 
  fbest0=1500000.0
  for i in range(imin, imin+maxiter):
    xbest, fbest = next(solver)              
    print(0,i, fbest, xbest)
    res= (copy.deepcopy(solver), i, xbest, fbest)  
    try :
         save_obj(res, name1+date_now()+'_'+np_int_tostr(i))
    except :  pass      
    if np.mod(i+1, 11)==0 :
            if np.abs(fbest - fbest0) < 0.001 : break;
            fbest0= fbest  

def np_remove_NA_INF_2d(X):
 im, jm= np.shape(X)
 for i in range(0, im) :
   for j in range(0,jm) :
    if np.isnan(X[i,j]) or np.isinf(X[i,j]): X[i,j]= X[i-1,j]
 return X
  
def np_addcolumn(arr, nbcol):
  sh= np.shape(arr);  
  vv= np.zeros((sh[0], nbcol))
  arr2= np.column_stack((arr, vv))
  return arr2

def np_addrow(arr, nbrow):
  sh= np.shape(arr);
  if len(sh)  > 1 :  
   vv= np.zeros((nbrow, sh[1]))
   arr2= np.row_stack((arr, vv))
   return arr2
  else:
   return np.append(arr, np.zeros(nbrow))

def np_int_tostr(i):
  if i < 10 : return '00'+str(i)
  elif i < 100 : return '0'+str(i)
  else: return str(i)


def np_dictordered_create():
   from collections import OrderedDict
   return OrderedDict()

def np_list_unique(seq) :   
 from sets import Set
 set = Set(seq)
 return list(set)


def np_list_tofreqdict(l1, wweight=[]) :
   # Get list frequency
   dd= dict()
   if len(wweight) == 0 :
     for x in l1 :
       try :    dd[x]+= 1
       except : dd[x]=  1
     return dd
   else :
     for ii, x in enumerate(l1) :
        try :     dd[x]+= wweight[ii]
        except :  dd[x]=  wweight[ii]
     return dd



# used to flatten a list or tupel [1,2[3,4],[5,[6,7]]] -> [1,2,3,4,5,6,7]
def np_list_flatten(seq):
    l = []
    for elt in seq:
        t = type(elt)
        if t is tuple or t is list:
            for elt2 in flatten(elt):  l.append(elt2)
        else:   l.append(elt)
    return l



def np_dict_tolist(dd , withkey=0) :
   if withkey :
     return [ [key,val]  for key, val in list(dd.items()) ]
   else:
     return [ val  for _, val in list(dd.items()) ]

def np_dict_tostr_val(dd) :
    return ','.join([ str(val)  for _, val in list(dd.items()) ])

def np_dict_tostr_key(dd) :
    return ','.join([ str(key)  for key,_ in list(dd.items()) ])



def np_removelist(x0, xremove=[]) :
  xnew=[]
  for x in x0 :
    if np_findfirst(x, xremove) < 0 : xnew.append(x)
  return xnew

def np_transform2d_int_1d(m2d, onlyhalf=False) :
  imax, jmax= np.shape(m2d)
  v1d= np.zeros((imax*jmax, 4))  ; k=0
  for i  in range(0,imax):
    for j in range(i+1, jmax):
       v1d[k,0]= i; v1d[k,1]=j
       v1d[k,2]= m2d[i,j]
       v1d[k,3]= np.abs(m2d[i,j])
       k+=1
  v1d= v1d[v1d[:,2] !=0]
  return np_sortbycol(v1d,3, asc=False)

def np_mergelist(x0, x1) :
  xnew= list(x0)
  for x in x1 :
    xnew.append(x)
  return list(xnew)

def np_enumerate2(vec_1d):
 v2= np.empty((len(vec_1d), 2))
 for k,x in enumerate(vec_1d): v2[k,0]= k; v2[k,1]= x 
 return v2

#Pivot Table from List data    
def np_pivottable_count(mylist):
    mydict={}.fromkeys(mylist,0)
    for e in mylist:
        mydict[e]= mydict[e] + 1   # Map Reduce function 
    ll2= np_dict_tolist(mydict)
    ll2= sorted(ll2, key = lambda x: int(x[1]), reverse=True)
    return ll2

def np_nan_helper(y):
    """ Input:  - y, 1d numpy array with possible NaNs
        Output - nans, logical indices of NaNs - index, a function, with signature indices= index(logical_indices),
              to convert logical indices of NaNs to 'equivalent' indices
    """
    return np.isnan(y), lambda z: z.nonzero()[0]

def np_interpolate_nan(y):
 nans, x= np_nan_helper(y)
 y[nans]= np.interp(x(nans), x(~nans), y[~nans])
 return y

def np_and1(x,y, x3=None, x4=None, x5=None, x6=None, x7=None, x8=None):
  if not x8 is None :  return np.logical_and.reduce((x8,x7, x6, x5, x4,x3,x,y))
  if not x7 is None :  return np.logical_and.reduce((x7, x6, x5, x4,x3,x,y))
  if not x6 is None:  return np.logical_and.reduce((x6, x5, x4,x3,x,y))
  if not x5 is None:  return np.logical_and.reduce((x5, x4,x3,x,y))
  if not x4 is None :  return np.logical_and.reduce((x4,x3,x,y))
  if not x3 is None :  return np.logical_and.reduce((x3,x,y))

def np_sortcol(arr, colid, asc=1):
 """ df.sort(['A', 'B'], ascending=[1, 0])  """
 df = pd.DataFrame(arr)
 arr= df.sort_values(colid, ascending=asc)   
 return arr.values

def np_sort(arr, colid, asc=1):
 """ df.sort(['A', 'B'], ascending=[1, 0])  """
 df = pd.DataFrame(arr)
 arr= df.sort_values(colid, ascending=asc)   
 return arr.values

def np_ma(vv, n):
  '''Moving average '''
  return np.convolve(vv, old_div(np.ones((n,)),n))[(n-1):]


@jit(float32[:,:](float32[:,:]))
def np_cleanmatrix(m):
  m= np.nan_to_num(m)
  imax, jmax= np.shape(m)
  for i in range(0,imax):
    for j in range(0,jmax):
       if abs(m[i,j]) > 300000.0 : m[i,j]= 0.0
  return m

def np_torecarray(arr, colname) :
  return  np.core.records.fromarrays(arr.T, names=','.join(colname), formats = ','.join(['f8'] * np.shape(arr)[1]))

def np_sortbycolumn(arr, colid, asc=True): 
 df = pd.DataFrame(arr)
 arr= df.sort_values(colid, ascending=asc)   
 return arr.values

def np_sortbycol(arr, colid, asc=True): 
 if len(np.shape(arr)) > 1 :
  df = pd.DataFrame(arr)
  arr= df.sort_values(colid, ascending=asc)   
  return arr.values
 else :
  return np.reshape(arr, (1,len(arr)))  
    
# if colid==0 : return arr[np.argsort(arr[:])]
# else : return arr[np.argsort(arr[:, colid])]

def min_kpos(arr, kth):
   ''' return kth mininimun '''
   return np.partition(arr, kth-1)[kth-1]

def max_kpos(arr, kth):
   ''' return kth mininimun '''
   n= len(arr)
   return np.partition(arr, n-kth+1)[n-kth+1-1]


@jit(nopython=True, nogil=True)    #Only numericalpostioon
def np_findfirst(item, vec):
    """return the index of the first occurence of item in vec"""
    for i in range(len(vec)):
        if item == vec[i]: return i
    return -1
    
@jit(nopython=True, nogil=True) 
def np_find(item, vec):
    """return the index of the first occurence of item in vec"""
    for i in range(len(vec)):
        if item == vec[i]: return i
    return -1


def find(item, vec):
    """return the index of the first occurence of item in vec"""
    for i in range(len(vec)):
        if item == vec[i]: return i
    return -1

def findnone(vec):
    """return the index of the first occurence of item in vec"""
    for i in range(len(vec)):
        if vec[i]==None : return i
    return -1

def findx(item, vec):
    """return the index of the first occurence of item in vec"""
    try :    
      if type(vec)== list : i2= vec.index(item)
      else :   
        i=np.where(vec==item)[0]
        i2= i[0] if len(i) > 0 else -1
    except: 
       i2= -1
    return i2

def finds(itemlist, vec):
  """return the index of the first occurence of item in vec"""
  idlist= []  
  for x in itemlist :
    ix=-1    
    for i in range(len(vec)):
        if x == vec[i]:  idlist.append(i); ix=i
    if ix==-1: idlist.append(-1)       
  if idlist== [] : return -1 
  else : return idlist

def findhigher(x, vec):
    """return the index of the first occurence of item in vec"""
    for i in range(len(vec)):
        if vec[i] > x : return i
    return -1

def findlower(x, vec):
    """return the index of the first occurence of item in vec"""
    for i in range(len(vec)):
        if vec[i] < x: return i
    return -1

  
import operator
def np_find_minpos(values):
 min_index, min_value = min(enumerate(values), key=operator.itemgetter(1))
 return min_index, min_value
 
def np_find_maxpos(values):
 max_index, max_value = max(enumerate(values), key=operator.itemgetter(1))
 return max_index, max_value

def np_find_maxpos_2nd(numbers):
    count = 0
    m1 = m2 = float('-inf')
    for i,x in enumerate(numbers):
        count += 1
        if x > m2:
            if x >= m1:
                m1, m2 = x, m1            
            else:
                m2 = x; i2= i
    return i2, m2 if count >= 2 else None

def np_findlocalmax2(v, trig):
 n=len(v)  
 v2= np.zeros((n,8))
 tmax,_= np_find_maxpos(v)
 if n < 3:
   max_index, max_value= np_find_maxpos(v)
   v2=  [[max_index, max_value]] 
   return v2
 else:
  for i,x in enumerate(v):
    if i < n-1 and i > 0 :
     if x > v[i-1] and x > v[i+1] : 
           v2[i,0]= i;  v2[i,1]=x
  v2 =np_sortbycolumn(v2,1,asc=False)
 
  #Specify the max :
 for k in range(0, len(v2)) :
  kmax= v2[k,0]
  kmaxl=  findhigher(v2[k,1], v[:kmax][::-1])  #Find same level of max
  kmaxr=  findhigher(v2[k,1], v[kmax+1:])

  kmaxl= 0 if kmaxl==-1 else kmax-kmaxl
  kmaxr= n if kmaxr==-1 else kmaxr+kmax
  
  v2[k,2]= np.abs(kmaxr-kmaxl)   #Range 
  v2[k,3]= np.abs(kmax-tmax)   #Range of the Max After
  v2[k,4]= 0  #Range of the Max Before
  v2[k,5]= kmax-kmaxl
  v2[k,6]= kmaxr-kmax

 v2= v2[np.logical_and( v2[:,5] > trig , v2[:,6] > trig ) ]
 v2 =np_sortbycolumn(v2,0,asc=1)
 return v2

def np_findlocalmin2(v, trig):
 n=len(v)  
 v2= np.zeros((n,8))
 tmin,_= np_find_minpos(v)
 if n < 3:
   max_index, max_value= np_find_minpos(v)
   v2= [[max_index, max_value]]
   return v2
 else:
  for i,x in enumerate(v):
    if i < n-1 and i > 0 :
     if x < v[i-1] and x < v[i+1] :      
           v2[i,0]= i;  v2[i,1]=x
  v2 =np_sortbycolumn(v2,1,asc=False)
 
 #Classification of the Local min
  for k in range(0, len(v2)) :
   if v2[k,1] != 0.0 :
    kmin= v2[k,0]
    kminl=  findlower(v2[k,1], v[:kmin][::-1])  #Find same level of min
    kminr=  findlower(v2[k,1], v[kmin+1:])

    kminl= 0 if kminl==-1 else kmin-kminl
    kminr= n if kminr==-1 else kminr+kmin

    v2[k,2]= np.abs(kminr-kminl)   #Range 
    v2[k,3]= np.abs(kmin-tmin)   #Range of the min After
    v2[k,4]= 0  #Range of the min After
    v2[k,5]= kmin-kminl
    v2[k,6]= kminr-kmin
 v2= v2[np.logical_and( v2[:,5] > trig , v2[:,6] > trig ) ]
 v2 =np_sortbycolumn(v2,0,asc=1)
 return v2

def np_findlocalmax(v):
 n=len(v)   
 v2= np.zeros((n,2))
 if n > 2:
  for i,x in enumerate(v):
    if i < n-1 and i > 0 :
     if x > v[i-1] and x > v[i+1] : 
           v2[i,0]= i;  v2[i,1]=x
  v2 =np_sortbycolumn(v2,1,asc=False)
  return v2
 else : 
   max_index, max_value= np_find_maxpos(v)
   return  [[max_index, max_value]]

def np_findlocalmin(v):
 n=len(v)   
 v2= np.zeros((n,2))
 if n > 2:
  for i,x in enumerate(v):
    if i < n-1 and i > 0 :
     if x < v[i-1] and x < v[i+1] :      
           v2[i,0]= i;  v2[i,1]=x
 
  v2 =np_sortbycolumn(v2[:i2],0,asc=True)
  return v2
 else : 
   
   max_index, max_value= np_find_minpos(v)
   return  [[max_index, max_value]]

def np_stack(v1, v2=None, v3=None, v4=None, v5=None):
   sh= np.shape(v1)
   if sh[0] < sh[1] :    
     v1= np.row_stack((v1,v2))
     if v3 != None : v1= np.row_stack((v1,v3))
     if v4 != None : v1= np.row_stack((v1,v4))
     if v5 != None : v1= np.row_stack((v1,v5))
   else :
     v1= np.column_stack((v1,v2))
     if v3 != None : v1= np.column_stack((v1,v3))
     if v4 != None : v1= np.column_stack((v1,v4))
     if v5 != None : v1= np.column_stack((v1,v5))       
       
   return v1

def np_uniquerows(a):
    a = np.ascontiguousarray(a)
    unique_a = np.unique(a.view([('', a.dtype)]*a.shape[1]))
    return unique_a.view(a.dtype).reshape((unique_a.shape[0], a.shape[1]))

def np_remove_zeros(vv, axis1=1):
   return vv[~np.all(vv == 0, axis=axis1)]

def np_sort(vv): 
 return vv[np.lexsort(np.transpose(vv)[::-1])]      #Sort the array by different column

def np_memory_array_adress(x):
    # This function returns the memory block address of an array.# b = a.copy(); id(b) == aid
    return x.__array_interface__['data'][0]

    


################   SQL UTIL  ###############################################################
'''
from sqlalchemy import create_engine
engine = create_engine("postgresql://u:p@host/database")
'''
import sqlalchemy as sql

def sql_create_dbengine(type1='',  dbname='', login='', password='', url='localhost', port=5432) :
   ''' Return SQL Alchemy Connector

sql_create_dbengine(type1='mysql',  dbname='', login='', password='', url='localhost', port=5432)

# psycopg2
engine = create_engine('postgresql+psycopg2://scott:tiger@localhost/mydatabase')


# MySQL-connector-python  Official one
engine = create_engine('mysql+mysqlconnector://scott:tiger@localhost/foo')
conda install -c anaconda mysql-connector-python=2.0.4
engine = create_engine('postgresql://%s:%s@localhost:5432/%s' %(myusername, mypassword, mydatabase))

engine = create_engine('sqlite:///  folder/foo.db')

   '''
   if type1=='postgres' :        # psycopg2
      engine = sql.create_engine('postgresql+psycopg2://' + login + ':' + password + '@' + url + ':'+ str(port) + '/'+ dbname)

   elif type1=='mysql' :        # MYSQL
      engine = sql.create_engine('mysql+mysqlconnector://' + login + ':' + password + '@' + url + ':'+ str(port) + '/'+ dbname)

   elif type1=='sqlite' :        # SQLlite
      engine = sql.create_engine('sqlite:/// ' + url+ '/'+ dbname)

   return engine




   
def sql_query(sqlr='SELECT ticker,shortratio,sector1_id, FROM stockfundamental',  dbengine=None, output='df', dburl='sqlite:///aaserialize/store/finviz.db'):
 '''
 :param sqlr:       'SELECT ticker,shortratio,sector1_id, FROM stockfundamental'
 :param output:     df   /   file1.csv
 :param dburl:      'sqlite:///aaserialize/store/finviz.db'
 :param dbengine:   dbengine = sql.create_engine('postgresql+psycopg2://postgres:postgres@localhost/coke')
 :return:
 '''
 if dbengine is None :   dbengine = sql.create_engine(dburl)
 df= pd.read_sql_query( sqlr, dbengine)

 if output=='df' : return df
 elif output.find('.csv') > -1 :  df.to_csv(output)

def sql_get_dbschema(dburl='sqlite:///aapackage/store/yahoo.db', dbengine=None, isprint=0):
    if dbengine is None :
        dbengine = sql.create_engine(dburl)
    inspector = sql.inspect(dbengine)
    l1= []
    for table_name in inspector.get_table_names():
        aux= ""
        for column in inspector.get_columns(table_name):
            l1.append([table_name, column['name']])
            aux= aux + ', ' + column['name']
        if isprint:   print(table_name + ", Col: "+aux +"\n")
    return np.array(l1)

def sql_delete_table(name, dbengine):
 pd.io.sql.execute('DROP TABLE IF EXISTS '+name, dbengine)
 pd.io.sql.execute('VACUUM', dbengine)




def sql_insert_excel(file1='.xls', dbengine=None, dbtype='') :
 ''' http://flask-excel.readthedocs.io/en/latest/
 https://pythonhosted.org/pyexcel/tutorial_data_conversion.html#import-excel-sheet-into-a-database-table
 from sqlalchemy import create_engine
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy import Column , Integer, String, Float, Date
from sqlalchemy.orm import sessionmaker
engine = create_engine("sqlite:///birth.db")
Base = declarative_base()
Session = sessionmaker(bind=engine)

class BirthRegister(Base):
...     __tablename__='birth'
...     id=Column(Integer, primary_key=True)
...     name=Column(String)
...     weight=Column(Float)
...     birth=Column(Date)

Base.metadata.create_all(engine)

https://www.digitalocean.com/community/tutorials/how-to-use-celery-with-rabbitmq-to-queue-tasks-on-an-ubuntu-vps
import os
import pyexcel
import datetime

from sqlalchemy import create_engine
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy import Column, Integer, String, Float, Date
from sqlalchemy.orm import sessionmaker


engine = create_engine("sqlite:///birth.db")
Base = declarative_base()
Session = sessionmaker(bind=engine)


# here is the destination table
class BirthRegister(Base):
    __tablename__ = 'birth'
    id = Column(Integer, primary_key=True)
    name = Column(String)
    weight = Column(Float)
    birth = Column(Date)


Base.metadata.create_all(engine)


# create fixture
data = [
    ["name", "weight", "birth"],
    ["Adam", 3.4, datetime.date(2015, 2, 3)],
    ["Smith", 4.2, datetime.date(2014, 11, 12)]
]
pyexcel.save_as(array=data,
                dest_file_name="birth.xls")

# import the xls file
session = Session()  # obtain a sql session
pyexcel.save_as(file_name="birth.xls",
                name_columns_by_row=0,
                dest_session=session,
                dest_table=BirthRegister)

# verify results
sheet = pyexcel.get_sheet(session=session, table=BirthRegister)
print(sheet)

session.close()
os.unlink('birth.db')
os.unlink("birth.xls")

 This code uses the openpyxl package for playing around with excel using Python code
 to convert complete excel workbook (all sheets) to an SQLite database
 The code assumes that the first row of every sheet is the column name
 Every sheet is stored in a separate table
 The sheet name is assigned as the table name for every sheet
 '''
 import sqlite3, re
 from openpyxl import load_workbook

 def slugify(text, lower=1):
    if lower == 1: text = text.strip().lower()
    text = re.sub(r'[^\w _-]+', '', text)
    text = re.sub(r'[- ]+', '_', text)
    return text

 # dbengine= sqlite3.connect(dbname)
 wb = load_workbook(filename=file1)  # r'abc.xlsx'
 sheets = wb.get_sheet_names()

 for sheet in sheets:   #Insert Sheet as new table
    ws = wb[sheet];     columns= []
    query = 'CREATE TABLE ' + str(slugify(sheet)) + '(ID INTEGER PRIMARY KEY AUTOINCREMENT'
    for row in ws.rows[0]:
        query += ', ' + slugify(row.value) + ' TEXT'
        columns.append(slugify(row.value))
    query += ');'
    dbengine.execute(query)


    tup = []
    for i, rows in enumerate(ws):
        tuprow = []
        if i == 0: continue
        for row in rows:
            tuprow.append(str(row.value).strip()) if str(row.value).strip() != 'None' else tuprow.append('')
        tup.append(tuple(tuprow))

    #Insert tuples
    insQuery1 = 'INSERT INTO ' + str(slugify(sheet)) + '('
    insQuery2 = ''
    for col in columns:
        insQuery1 += col + ', '
        insQuery2 += '?, '
    insQuery1 = insQuery1[:-2] + ') VALUES('
    insQuery2 = insQuery2[:-2] + ')'
    insQuery = insQuery1 + insQuery2

    dbengine.executemany(insQuery, tup)
    dbengine.commit()
 dbengine.close()

def sql_insert_df(df, dbtable, dbengine, col_drop=['id'], verbose=1) :
 for c in df.columns:    #Remove columns
   if c in col_drop:  df = df.drop(c, axis=1)

 list1 = df.to_dict(orient='records')  # The orient='records' is the key of this, it allows to align with the format mentioned in the doc to insert in bulks.

 '''
 #list1= list1[0]
 list2=[]
 for i in xrange(0, len(list1)) :
    listx= list1[i]     
    for col in col_drop :
       del listx[col] #remove the ID, so SQL can create it, as well timeStamp
    list2.append(listx)
 '''

 if verbose : print(list1)

 metadata = sql.schema.MetaData(bind=dbengine,reflect=True)
 table = sql.Table(dbtable, metadata, autoload=True)

 Session = sql.orm.session.sessionmaker(bind=dbengine)
 session = Session()

 # Insert the dataframe into the database in one bulk
 res= dbengine.execute(table.insert(), list1)

 session.commit(); session.close()  
 return res

def sql_insert_csv(csvfile, dbtable, dbengine, col_drop=[]) :
 start = datetime.now()
 chunksize = 20000
 j = 0
 index_start = 1

 for df in pd.read_csv(csvfile, chunksize=chunksize, iterator=True, encoding='utf-8'):

   df.index += index_start
   df = df.rename(columns={c: c.replace(' ', '') for c in df.columns}) # Remove spaces from columns
   #df['CreatedDate'] = pd.to_datetime(df['CreatedDate']) # Convert to datetimes
   #df['ClosedDate'] =  pd.to_datetime(df['ClosedDate'])
   for c in df.columns:    #Remove columns
      if c in col_drop:  df = df.drop(c, axis=1)


   ################  df to SQL Alchemy ######################################
   listToWrite = df.to_dict(orient='records')  # The orient='records' is the key of this, it allows to align with the format mentioned in the doc to insert in bulks.

   metadata = sql.schema.MetaData(bind=dbengine,reflect=True)
   table =    sql.Table(dbtable, metadata, autoload=True)
   Session = sql.orm.session.sessionmaker(bind=dbengine)
   session = Session()

   # Inser the dataframe into the database in one bulk
   dbengine.execute(table.insert(), listToWrite)
   session.commit();   session.close()

   #     df.to_sql('data', disk_engine, if_exists='append')
   j+=1
   index_start = df.index[-1] + 1
   print('{} seconds: completed {} rows'.format((datetime.now() - start).seconds, j*chunksize))




 ''' Batch Mode :
 #  CSV to SQL Lite ----------------------------------------------------------------------
disk_engine = create_engine('sqlite:///311_8M.db') # Initializes database with filename 311_8M.db in current directory

start = dt.datetime.now()
chunksize = 20000
j = 0
index_start = 1

for df in pd.read_csv('311_100M.csv', chunksize=chunksize, iterator=True, encoding='utf-8'):

    df = df.rename(columns={c: c.replace(' ', '') for c in df.columns}) # Remove spaces from columns
    df['CreatedDate'] = pd.to_datetime(df['CreatedDate']) # Convert to datetimes
    df['ClosedDate'] = pd.to_datetime(df['ClosedDate'])

    df.index += index_start

    # Remove the un-interesting columns
    columns = ['Agency', 'CreatedDate', 'ClosedDate', 'ComplaintType', 'Descriptor',
               'CreatedDate', 'ClosedDate', 'TimeToCompletion','City']

    for c in df.columns:
        if c not in columns:
            df = df.drop(c, axis=1)

    j+=1
    print '{} seconds: completed {} rows'.format((dt.datetime.now() - start).seconds, j*chunksize)

    df.to_sql('data', disk_engine, if_exists='append')
    index_start = df.index[-1] + 1




 '''

def sql_insert_csv2(csvfile='', dbtable='', columns=[], dbengine=None, nrows= 10000):
    """
    Upload data to a temporary table first using PANDAs to identify optimal data-types for columns
    PANDAS is not speed-efficient as it uses INSERT commands rather than COPY e.g. it took COPY 16mins average
    to get a 15GB CSV into the database (door-to-door) whereas pandas.to_sql took 50mins
    """

    dbtable +='_temp'
    counter = 0
    for i in os.listdir(csvfile):
        # Cycle through all CSVs and upload a small chunk to make sure everything is OK
        if counter < 1:
            if i.endswith(".csv") & i.startswith("..."):
                print("Reading CSV: %s into PANDAs data-frame" % i)

                # First 1,000,000 rows
                #df = pd.read_csv(os.path.join(csvfile, i), nrows=1000000, header=None, sep='~') #sep=None; automatically find by sniffing

                # Upload whole file
                df = pd.read_csv(os.path.join(csvfile, i), nrows= nrows, header=None, sep='~') #sep=None; automatically find by sniffing
                # My dates were in columns 2 and 3
                # The column names were not present in the original CSVs
                df.columns =  columns
                print("CSV read-in successfully")
                print(df.shape)
                print("Uploading %s to SQL Table: %s" % (i, dbtable))
                df.to_sql(dbtable, engine, if_exists='append', index=False)
                counter += 1
                print("Successfully uploaded: %d" % counter)




def sql_postgres_create_table(mytable='', database='', username='', password='' ):
    """ Create table copying the structure of the temp table created using pandas  Timer to benchmark """
    # Connect
    import psycopg2
    con = psycopg2.connect(database=database, user=username, password=password)
    cur = con.cursor()
    if con:
        print('Connected: %s' % con)
    else:
        print('Connection lost')
        sys.exit(1)

    try:
        # Check if table exists already
        cur.execute("""
                    SELECT relname FROM pg_class WHERE relname = '{0}';
                    """.format(my_table))
        table_test = cur.fetchone()[0]
    except Exception as e:
        print('Table %s does not exist' % mytable)
        table_test = None

    if table_test:
        print('%s already exists' % mytable)
    else:
        print('Creating table: %s' % mytable)
        try:
            # Copy structure and no data (1=2 is false)
            cur.execute("""
                        CREATE TABLE {0} AS SELECT * FROM {1} WHERE 1=2;
                        """.format(mytable, mytable+'_temp'))
            con.commit()
            print('Table created successfully')
        except psycopg2.DatabaseError as e:
            if con:
                con.rollback()
            print('Error %s' % e)
            sys.exit(1)
    con.close()

def sql_postgres_insert_csv(path_2_csv='', mytable=''):
    """  Use the PostgreSQL COPY command to bulk-copy the CSVs into the newly created table """
    # Connect
    con = psycopg2.connect(database=mydatabase, user=myusername, password=mypassword)
    cur = con.cursor()
    if con:
        print('Connected: %s' % con)
    else:
        print('Connection lost')
        sys.exit(1)

    copy_sql = """
               COPY %s FROM stdin DELIMITERS '~' CSV;
               """ % mytable
    counter = 0
    start_time = time.time()

    for i in os.listdir(path_2_csv):
        if i.endswith(".csv") and i.startswith("..."):
            print("Uploading %s to %s" % (i, mytable))
            with open(os.path.join(path_2_csv, i), 'r') as f:
                cur.copy_expert(sql=copy_sql, file=f)
                con.commit()
                counter += 1
                print("Successfully uploaded %d CSVs" % counter)
                current_speed = old_div((old_div((time.time()-start_time),60)),counter)
                print("Average speed is %.2f minutes per database" % current_speed)
    con.close()
    end_time = time.time()
    print("Total duration of COPY: %.2f minutes" % (old_div((end_time - start_time),60)))

def sql_postgres_query_to_csv(sqlr='SELECT ticker,shortratio,sector1_id, FROM stockfundamental', csv_out=''):
    """ Submit query to created PostgreSQL database and output results to a CSV  """
    import psycopg2
    con = psycopg2.connect(database=mydatabase, user=myusername, password=mypassword)
    cur = con.cursor()
    if con:   print('Connected: %s' % con)
    else:
       print('Connection lost'); return  -1

    output_query = "COPY ({0}) TO STDOUT WITH CSV HEADER".format(sqlr)
    with open(csv_out, 'w') as f:
        cur.copy_expert(output_query, f)
        print("Successfully submitted results to: %s" % csv_out)
    con.close()

def sql_postgres_pivot():
   '''
Enabling the Crosstab Function
As we previously mentioned, the crosstab function is part of a PostgreSQL extension called tablefunc. To call the crosstab function,
you must first enable the tablefunc extension by executing the following SQL command:
CREATE extension tablefunc;

SELECT *
FROM crosstab( 'select student, subject, evaluation_result from evaluations order by 1,2')
AS final_result(Student TEXT, Geography NUMERIC,History NUMERIC,Language NUMERIC,Maths NUMERIC,Music NUMERIC);


##### Correct Even iF there are missing values :
http://www.vertabelo.com/blog/technical-articles/creating-pivot-tables-in-postgresql-using-the-crosstab-function

SELECT *
FROM crosstab( 'select student, subject, evaluation_result from evaluations
                where extract (month from evaluation_day) = 7 order by 1,2',
                'select name from subject order by 1')
     AS final_result(Student TEXT, Geography NUMERIC,History NUMERIC,Language NUMERIC,Maths NUMERIC,Music NUMERIC);


   '''
   pass


def sql_mysql_insert_excel():
 import MySQLdb, xlrd

 list= xlrd.open_workbook("prod.xls")
 sheet= list.sheet_by_index(0)

 database = MySQLdb.connect (host="localhost" , user="root" , passwd="" ,db="table")
 cursor = database.cursor()

 query= """INSERT INTO produits (idProduit, idCategorie, LibelleProduit, PrixProduit) VALUES (%s, %s, %s, %s)"""
 for r in range(1,sheet.nrows):
    idProduit = sheet.cell(r,0).value
    categorie = 999
    libelle=sheet.cell(r,1).value
    prix=sheet.cell(r,3).value #>>HERE THE PROBLEM the Imported Value <<<<

    values = (idProduit,categorie,libelle,prix)

    cursor.execute(query,values)

 cursor.close();
 database.commit()

 database.close()

 print("")
 print("All done !")

 columns= str(sheet.ncols)
 rows=str(sheet.nrows)
 print("i just import "+columns+" columns and " +rows+ " rows to MySQL DB")


def sql_pivotable(dbcon, ss='select  '):
  '''

 1) get the category

 2) Build the Pivot From category
  SELECT *
  FROM crosstab( 'select student, subject, evaluation_result from evaluations
                where extract (month from evaluation_day) = 7 order by 1,2',
                'select name from subject order by 1')
   AS final_result(Student TEXT, Geography NUMERIC,History NUMERIC,Language NUMERIC,Maths NUMERIC,Music NUMERIC);

  https://www.amazon.com/PostgreSQL-High-Performance-Gregory-Smith/dp/184951030X/ref=as_li_ss_tl?s=books&ie=UTF8&qid=1458352081&sr=1-6&keywords=postgres&linkCode=sl1&tag=postgres-bottom-20&linkId=c981783121cbd5542dc2b44a2297df57


http://blog.brakmic.com/data-science-for-losers-part-2/

Here we instruct Pandas to merge two tables by using certain primary keys from both when combining their rows into a new table. The parameter how instructs Pandas to use the inner-join which means it will only combine such rows which belong to both of the tables. Therefore we’ll not receive any NaN-rows. But in some cases this could be desirable. Then use the alternative options like left, right or outer.

Pivots with Tables from SQLAlchemy

And of course it’s possible to generate the same pivot tables with data that came from SQLAlchemy.
They’re nothing else but DataFrames all the way down. OK, not absolutely all the way down,
because there are also Series and NumPy arrays etc.,
but this is a little bit too much of knowledge for Losers like us. Maybe in some later articles.


 :return:
  '''

  ss= ''
  pass

def np_pivotable_create(table, left, top, value):
    """
    Creates a cross-tab or pivot table from a normalised input table. Use this
    function to 'denormalize' a table of normalized records.

    * The table argument can be a list of dictionaries or a Table object.
    (http://aspn.activestate.com/ASPN/Cookbook/Python/Recipe/334621)
    * The left argument is a tuple of headings which are displayed down the
    left side of the new table.
    * The top argument is a tuple of headings which are displayed across the
    top of the new table.
    Tuples are used so that multiple element headings and columns can be used.

    E.g. To transform the list (listOfDicts):

    Name,   Year,  Value
    -----------------------
    'Simon', 2004, 32
    'Simon', 2005, 128
    'Russel', 2004, 64
    'Eric', 2004, 52
    'Russel', 2005, 32

    into the new list:

    'Name',   2004, 2005
    ------------------------
    'Simon',  32,     128
    'Russel',  64,     32
    'Eric',   52,     NA

    you would call pivot with the arguments:

    newList = pivot(listOfDicts, ('Name',), ('Year',), 'Value')

    """
    rs = {}
    ysort = []
    xsort = []
    for row in table:
        yaxis = tuple([row[c] for c in left])       # e.g. yaxis = ('Simon',)
        if yaxis not in ysort: ysort.append(yaxis)
        xaxis = tuple([row[c] for c in top])        # e.g. xaxis = ('2004',)
        if xaxis not in xsort: xsort.append(xaxis)
        try:
            rs[yaxis]
        except KeyError:
            rs[yaxis] = {}
        if xaxis not in rs[yaxis]:
            rs[yaxis][xaxis] = 0
        rs[yaxis][xaxis] += row[value]


    # In the following loop we take care of missing data,
    # e.g 'Eric' has a value in 2004 but not in 2005
    for key in rs:
        if len(rs[key]) > len(xsort):
            for var in xsort:
                if var not in list(rs[key].keys()):
                    rs[key][var] = ''

    headings = list(left)
    xsort.sort()
    headings.extend(xsort)

    t = []

    # The lists 'sortedkeys' and 'sortedvalues' make sure that
    # even if the field 'top' is unordered, data will be transposed correctly.
    # E.g. in the example above the table rows are not ordered by the year
    for left in ysort:
        row = list(left)
        sortedkeys = list(rs[left].keys())
        sortedkeys.sort()
        sortedvalues = list(map(rs[left].get, sortedkeys))
        row.extend(sortedvalues)
        t.append(dict(list(zip(headings,row))))
    return t




############### PANDA UTIL  ################################################################


def pd_info(df, doreturn=1) :
 df.info()
 dtype0= df.dtypes.to_dict()
 for i in df.columns :
   print(i,  dtype0[i], type(df[i].values[0]),  df[i].values[0])
 if doreturn : return dtype0

def pd_info_memsize(df, memusage=0):
 df.info(memory_usage='deep')
 pd_info(df)
 if memusage==1 :return df.memory_usage().sum()



# selection
def pd_selectrow(df, **conditions):
    '''Select rows from a df according to conditions
    pdselect(data, a=2, b__lt=3) __gt __ge __lte  __in  __not_in
    will select all rows where 'a' is 2 and 'b' is less than 3
    '''
    if type(df) == pd.Series:   df = pd.DataFrame({'value': df})
    for c, value in list(conditions.items()):
        if c in df.columns:
            c = (df[c]==value)
        elif c.endswith('__neq') or c.endswith('__not_eq'):
            if c.endswith('__neq'):
                c =c[:-len('__neq')]
            elif c.endswith('__not_eq'):
                c =c[:-len('__not_eq')]
            c =df[c]!=value
        elif c.endswith('__gt'):
            c =c[:-len('__gt')]
            c =df[c] > value
        elif c.endswith('__ge'):
            c =c[:-len('__ge')]
            c =df[c] >= value
        elif c.endswith('__gte'):
            c =c[:-len('__gte')]
            c =df[c] >= value
        elif c.endswith('__lt'):
            c =c[:-len('__lt')]
            c =df[c] < value
        elif c.endswith('__le'):
            c =c[:-len('__le')]
            c =df[c] <= value
        elif c.endswith('__lte'):
            c =c[:-len('__lte')]
            c =df[c] <= value
        elif c.endswith('__in'):
            c =c[:-len('__in')]
            c = np.in1d(df[c], value)
        elif c.endswith('__not_in'):
            c =c[:-len('__not_in')]
            c = np.in1d(df[c], value, invert=True)
        else:
            raise ValueError("Cannot process condition '{}'".format(c))
        df = df[c]
    return df

def pd_csv_randomread(filename, nsample=10000, filemaxline=-1, dtype=None) :
 if filemaxline==-1 : n = sum(1 for line in open(filename)) - 1 #number of records in file (excludes header)
 else : n= filemaxline
 skip = np.sort(np.random.randint(1,n+1,n- nsample)) #the 0-indexed header will not be included in the skip list
 df= pd.read_csv(filename, skiprows=skip, dtype=dtype)
 return df



# Creation
def pd_array_todataframe(array, colname=None, index1=None, dotranspose=False):
   sh= np.shape(array)
   if len(sh) > 1 :
     if sh[0] < sh[1] and dotranspose :   #  masset x time , need Transpose
         return pd.DataFrame(data= array.T, index= index1, columns=colname)
     else :
         return pd.DataFrame(data= array, index= index1, columns=colname)
   else :   #1d vector
     return pd.DataFrame(data= np.array(array).reshape(-1, 1), index= index1, columns=colname)

def pd_dataframe_toarray(df):
  col1= df.index
  array1= df.reset_index().values[1:,:]
  column_name= df.columns
  return column_name, col1, array1

def pd_createdf(array1, col1=None, idx1=None) :
  return  pd.DataFrame(data=array1, index=idx1, columns=col1)

def pd_create_colmapdict_nametoint(df) :
  ''' 'close' ---> 5    '''
  col= df.columns.values; dict1={}
  for k,x in enumerate(col) :
    dict1[x]= k
  return dict1



# Extract
def pd_extract_col_idx_val(df):
 return df.columns.values,  df.index.values, df.values


def pd_extract_col_uniquevalue_tocsv(df, colname='', csvfile='') :
 ''' Write one column into a file   '''
 a= df[colname].unique()
 a= np.array(a)
 pd.DataFrame( a, columns=[colname]).to_csv( csvfile )
 print(csvfile)



def pd_split_col_idx_val(df):
 return df.columns.values,  df.index.values, df.values

def pd_splitdf_inlist(df, colid, type1="dict"):
    ''' Split df into dictionnary of dict/list '''
    UniqueNames = df.sym.unique()
    if type1=='dict' :
      dfDict = {elem: pd.DataFrame for elem in UniqueNames}
      for key in list(DataFrameDict.keys()):
        dfDict[key] = df[df[colid] == key]
      return dfDict

    if type1 == "list":
      l1 = []
      for key in UniqueNames:
        l1.append(df[df[colid] == key])
      return l1

def pd_find(df, regex_pattern='*', col_restrict=[], isnumeric=False, doreturnposition=False) :
 ''' Find string / numeric values inside df columns, return position where found
     col_restrict : restrict to these columns '''
 dtype0= df.dtypes.to_dict()
 col0=   df.columns if col_restrict==[] else col_restrict

 if not isnumeric :   #object string columns
   colx= [col for col in col0 if str(dtype0[col])== 'object'   ]
   print("Searching Cols: "+ str(colx))
   for i, coli in enumerate(colx) :
     dfm= df[coli].str.contains(regex_pattern, na=False, regex=True)
     if i==0 :  mask = dfm
     else    :  mask=  np.column_stack((mask, dfm))

 else :
   numval= regex_pattern
   colx= [col for col in col0 if str(dtype0[col]).find('float') > -1 or str(dtype0[col]).find('int') > -1 ]
   print("Searching Cols: "+ str(colx))
   for i, coli in enumerate(colx) :
     dfm= (df[coli]== numval).values        #df[col].loc[ df[col]== numval  ]
     if i==0 :  mask = dfm
     else    :  mask=  np.column_stack((mask, dfm))

 # print mask
 if len(mask.shape) < 2 :  mask= mask.values.reshape(-1,1)

 gc.collect()
 if doreturnposition :
   locate= np_dictordered_create()        #Position in Dict_Column --> Indice_j
   for j in range(0, mask.shape[1])  :
     pos_tempj= np.array([ i for i in range(0,mask.shape[0])   if mask[i,j] ], dtype= np.int32)
     locate[colx[j]] = pos_tempj

   # locate= np.array([ (colx[j], i) for i in xrange(0,mask.shape[0]) for j in xrange(0, mask.shape[1])  if mask[i,j] ])
   return df.loc[mask.any(axis=1)], locate
 else :
   return df.loc[mask.any(axis=1)]



#dtypes
'''
def pd_dtypes_tocategory(df, columns=[], targetype='category'):
   for col in columns : df[col]= df[col].astype(targetype)
   return df
'''
def pd_dtypes_totype2(df, columns=[], targetype='category'):
   for col in columns : df[col]= df[col].astype(targetype)
   return df


def pd_dtypes(df, returnasdict=0) :
 from collections import OrderedDict
 dtype0= OrderedDict(df.dtypes.apply(lambda x: x.name))
 ss='''{ '''
 for i, col in enumerate(df.columns.values) :
   ss+= "'"+col+ "':" + "'"+ dtype0[col]+"', "
   if i%3==0 and i> -1 : ss+= '\n'
 ss= ss[0:-2] + ' }'

 if returnasdict : return eval(ss)
 print(ss)
 print('''\n df.astype(typedict)  Pandas 'object' : 'category', 'unicode' , 'str'  'boolean',
    	float16, float32, int8, int16, int32,uint8, uint16, uint32 ''')


def pd_df_todict(df, colkey='table', excludekey=[''], onlyfirstelt= True) :
      df1= df.drop_duplicates(colkey).reset_index(level=0, drop=True)
      dict0 = {}
      for i in range(len(df)):
         id0 = df.iloc[i,0]
         val0= df.iloc[i,1]
         if id0 not in excludekey :
            dict0.setdefault(id0, [])
            if onlyfirstelt :  dict0[id0]= val0
            else:          dict0[id0].append(val0)
      return dict0


'''
def pd_dtypes_getdict(df=None, csvfile=None) :
   if df is not None :   return df.dtypes.to_dict()
   elif csvfile is not None :
      df= pd.read_csv(csvfile, nrows= 1000)
      return df.dtypes.to_dict()


def pd_dtypes_getblaze(df1) :
 from collections import OrderedDict
 x= str(OrderedDict(df1.dtypes.apply(lambda x: x.name)))
 x=x.replace("',", "':")
 x=x.replace("(", "")
 x=x.replace(")", "")
 x=x.replace("OrderedDict[", "{")
 x=x.replace("]", "}")
 print(" string 5char: |S5 object in Pandas,  object_bool, string_, unicode
	float16, float32, float64, int8, int16, int32, int64,	uint8, uint16, uint32, uint64")
 return x
'''


'''
 # return df1.dtypes.apply(lambda x: x.name).to_dict()
 from collections import OrderedDict
 ds=OrderedDict()
 for i,x in enumerate(df1.columns.values) :
   ds[x]=  df1.dtypes[i].name
 print ds
'''



# Apply / Transform df
def pd_applyfun_col(df, newcol, ff, use_colname="all/[colname]") :
   ''' use all Columns to compute values '''
   if use_colname=="all/[colname]" : df[newcol]= ff(df.values)
   else :  df[newcol]= ff(df[use_colname].values)
   return df

def pd_cleanquote(q):
 col= q.columns.values
 for kid in col:
   if kid not in ['date', 'day','month','year'] :
      q[kid]= pd.to_numeric(q[kid], errors='coerce').values  #Put NA on string
 
 q= q.fillna(method='pad')  
 return q

def pd_date_intersection(qlist) :
 date0= set(qlist[0]['date'].values)
 for qi in qlist :
   qs= set(qi['date'].values)
   date0 = set.intersection(date0, qs)
 date0= list(date0); date0= sorted(date0)
 return date0

def pd_is_categorical(z):
    if isinstance(z, pd.Categorical): return True
    try:
        return isinstance(z.values, pd.Categorical)
    except:
        return False

def pd_str_encoding_change(df, cols, fromenc='iso-8859-1', toenc='utf-8'):
    #  Western European: 'cp1252'
    for col in cols:
        df[col] = df[col].str.decode(fromenc).str.encode(toenc)
    return df

def pd_str_unicode_tostr(df, targetype=str) :
 '''
 https://www.azavea.com/blog/2014/03/24/solving-unicode-problems-in-python-2-7/
 Nearly every Unicode problem can be solved by the proper application of these tools;
 they will help you build an airlock to keep the inside of your code nice and clean:

encode(): Gets you from Unicode -> bytes
decode(): Gets you from bytes -> Unicode
codecs.open(encoding=”utf-8″): Read and write files directly to/from Unicode (you can use any encoding,
 not just utf-8, but utf-8 is most common).
u”: Makes your string literals into Unicode objects rather than byte sequences.
Warning: Don’t use encode() on bytes or decode() on Unicode objects

>>> uni_greeting % utf8_name
Traceback (most recent call last):
 File "<stdin>", line 1, in <module>
UnicodeDecodeError: 'ascii' codec can't decode byte 0xc3 in position 3: ordinal not in range(128)
# Solution:
>>> uni_greeting % utf8_name.decode('utf-8')
u'Hi, my name is Josxe9.'

 '''
 return pd_dtypes_type1_totype2(df, fromtype=str, targetype=str)

def pd_dtypes_type1_totype2(df, fromtype=str, targetype=str) :
 for ii in df.columns :
  if isinstance(df[ii].values[0], fromtype) :
     df[ii]= df[ii].astype(targetype)
 return df



#----Insert/update col row  -----------------------------------------
def pd_resetindex(df):
  df.index = list(np.arange(0,len(df.index)))
  return df

def pd_insertdatecol(df_insider, format1="%Y-%m-%d %H:%M:%S:%f") :
 df_insider= pd_addcol(df_insider, 'dateinsert')
 df_insider['dateinsert']=  df_insider['dateinsert'].astype(str)
 df_insider.dateinsert.values[:]  =  date_nowtime('str', format1= format1)
 return df_insider

def pd_replacevalues(df,  matrix):
 ''' Matrix replaces df.values  '''
 imax, jmax= np.shape(vec)
 colname= df.columns.values
 for j in jmax :
   df.loc[colname[j]] = matrix[:,j]

 return df

def pd_removerow(df, row_list_index=[23, 45]) :
 return df.drop(row_list_index)

def pd_removecol(df1, name1):
 return df1.drop(name1, axis=1)

def pd_addcol(df1, name1='new'):
 df1[name1]= 0.0
 '''
 tmax= len(df1.index)
 if type(name1)==str : 
    df1.loc[:, name1] = pd.Series(np.zeros(tmax), index=df1.index)
 else :
   for name0 in name1 :
     df1.loc[:, name0] = pd.Series(np.zeros(tmax), index=df1.index)
 '''
 return df1

def pd_insertcol(df, colname, vec):
 ''' Vec and Colname must be aligned '''
 df[colname]= vec
 '''
 ncol= len(df.columns.values)
 sh= np.shape(vec)
 if len(sh) > 1 :
    imax, jmax= sh[0], sh[1]
    for j in range(0,jmax) :
      df.insert(ncol, colname[j], vec[:,j])

 else :
    df.insert(ncol, colname, vec)
 '''
 return df

def pd_insertrow(df, rowval, index1=None, isreset=1):
 df2=  pd_array_todataframe(rowval, df.columns.values, index1 )
 df=   df.append(df2, ignore_index=True)
 #if isreset : df.reset_index(inplace=True)
 return df


#---- h5 In / Out  ------------------------------------------------
def pd_h5_cleanbeforesave(df):
   '''Clean Column type before Saving in HDFS: Unicode, Datetime  '''
   #Unicode String :
   df= pd_resetindex(df)   #Reset Index 0 to 100000
   df= pd_str_unicode_tostr(df, targetype=str)

   '''
   for col in  ['date'] :
   # df_pd[col] = df_pd[col].astype(str)
   df_pd[col] = df_pd[col].apply(lambda x: x.encode('utf-8').strip())

   df_pd.to_hdf('test.h5','df',format='table',mode='w',data_columns=True,encoding='latin1')

   '''
   return df

def pd_h5_addtable(df, tablename, dbfile='F:\temp_pandas.h5') :
  store = pd.HDFStore(dbfile)
  if find(tablename, list(store.keys())) > 0 :
     # tablename=tablename + '_1';
     print('Table Exist, change table name')
  else :
     store.append(tablename, df); store.close()


def pd_h5_tableinfo(filenameh5, table):
    store= pd.HDFStore(filenameh5)
    return store.get_storer(table).table


def pd_h5_dumpinfo(dbfile='E:\_data\stock\intraday_google.h5'):
  store = pd.HDFStore(dbfile)
  extract=[]; errsym=[]
  for symbol in list(store.keys()):
     try:
       df= pd.read_hdf(dbfile, symbol)
       extract.append([symbol[1:], df.shape[1],
                       df.shape[0], datetime_tostring(df.index.values[0]),
                        datetime_tostring(df.index.values[-1]) ])

     except: errsym.append(symbol)
  return np.array(extract), errsym

def pd_h5_save(df, filenameh5='E:/_data/_data_outlier.h5', key='data'):
 ''' File is release after saving it'''
 store = pd.HDFStore(filenameh5)  
 store.append(key, df); store.close()

def pd_h5_load(filenameh5='E:/_data/_data_outlier.h5', table_id='data', exportype="pandas", rowstart=-1, rowend=-1, cols=[]):
  if rowend==-1 :  df=  pd.read_hdf(filenameh5, table_id)
  else :           df=  pd.read_hdf(filenameh5, table_id,  start=rowstart, end=rowend)
  if exportype=='pandas':   return df
  elif exportype=='numpy' : return df.values

def pd_h5_fromcsv_tohdfs(dircsv='dir1/dir2/', filepattern='*.csv', tofilehdfs='file1.h5', tablename='df', col_category=[],
                         dtype0=None,  encoding='utf-8', chunksize= 2000000, mode='a',format='table', complib=None):

  csvlist=  os_file_listall(dircsv, filepattern)
  csvlist=  csvlist[2]   # 2: Full_path + filename

  store = pd.HDFStore(tofilehdfs)
  max_size= {}
  for ii, tablei_file in enumerate(csvlist) :

    #Inference of Type, String size from top 1000 records......
    if ii== 0 :
       if dtype0 is None :
         df_i= pd.read_csv(tablei_file, nrows=1000, sep=',') #Use Top 1000 to estimate size....
         dtype0= df_i.dtypes.to_dict()
         col_list=[]
         # for col, x  in dtype0.items():
         #   if x == np.dtype('O') :    #Object == date, string, need to convert to string....
         #      col_list.append(col)
             # df_i[col] =    df_i[col].map(lambda x:  str(str(x).encode(encoding)))
             # max_size[col]= max(15, df_i[col].str.len().max())

    #ISSUE: DO NOT USE encoding='utf-8' when reading the Japanese Character.....
    list01= pd.read_csv(tablei_file,  chunksize=chunksize, dtype= dtype0, sep=',')  #, encoding=encoding)
    for k, df_i in enumerate(list01) :
      print('.', end=' ')
      for col in col_category :
          df_i[col] =    df_i[col].astype('category')

      #for col in col_list :
         # df_i[col] = df_i[col].map(lambda x:  str(str(x).encode(encoding)))
      store.append(tablename, df_i,  mode=mode, format=format, complib=complib)  #,  min_itemsize=max_size)

  store.close(); del store; print('\n')
  return os_file_exist(tofilehdfs)


def pd_np_toh5file(numpyarr, fileout="file.h5", table1='data'):
 pd= pd.DataFrame(numpyarr);  st= pd.HDFStore(fileout);  st.append(table1, pd); del pd



##############Date Manipulation#######################################################################
from dateutil import parser

def date_allinfo():
   '''

   https://aboutsimon.com/blog/2016/08/04/datetime-vs-Arrow-vs-Pendulum-vs-Delorean-vs-udatetime.html


   '''


def date_convert(t1, fromtype, totype) :
   try :
      n= len(t1)
   except : t1= [t1]; n=1
   t0= t1[0]

   if isinstance(t0, str) :
      pass

   elif isinstance(t0, int) :
      pass

   elif isinstance(t0, datetime) :
      pass

   elif isinstance(t0, np.datetime64) :
     pass

   for t in t1 :
      t2= _dateconvert_from(t, type1)   #  to Datetime
      t3= _dateconvert_from(t, totype)   # Datetime  to target type
      tlist.append(t3)
   return tlist

def datetime_tostring(datelist1):
 if isinstance(datelist1, datetime.date) :      return datelist1.strftime("%Y%m%d")
 if isinstance(datelist1, np.datetime64 ) :
   t= pd.to_datetime(str(datelist1)) ;  return t.strftime("%Y%m%d") 
  
 date2 = []
 for t in datelist1:
     date2.append(t.strftime("%Y%m%d"))
 return date2

def date_remove_bdays(from_date, add_days):
    isint1= isint(from_date)  
    if isint1 :  from_date= dateint_todatetime(from_date)
    business_days_to_add = add_days
    current_date = from_date
    while business_days_to_add < 0:
        current_date += datetime.timedelta(days=-1)
        weekday = current_date.weekday()
        if weekday >= 5: # sunday = 6
            continue
        business_days_to_add += 1
    if isint1 : return datetime_toint(current_date) 
    else :      return   current_date

def date_add_bdays(from_date, add_days):
    isint1= isint(from_date)
    if isint1 :  from_date= dateint_todatetime(from_date)
    business_days_to_add = add_days
    current_date = from_date
    while business_days_to_add > 0:
        current_date += datetime.timedelta(days=1)
        weekday = current_date.weekday()
        if weekday >= 5: # sunday = 6
            continue
        business_days_to_add -= 1
    if isint1 : return datetime_toint(current_date) 
    else :      return   current_date

def datenumpy_todatetime(tt, islocaltime=True):
   #  http://stackoverflow.com/questions/29753060/how-to-convert-numpy-datetime64-into-datetime
   if type(tt)==np.datetime64:
      if islocaltime:
         return datetime.datetime.fromtimestamp(old_div(tt.astype('O'), 1e9))
      else:
         return datetime.datetime.utcfromtimestamp(old_div(tt.astype('O'), 1e9))
   elif type(tt[0])==np.datetime64:
      if islocaltime:
         v=[datetime.datetime.fromtimestamp(old_div(t.astype('O'), 1e9)) for t in tt]
      else:
         v=[datetime.datetime.utcfromtimestamp(old_div(t.astype('O'), 1e9)) for t in tt]
      return v
   else:
      return tt  # datetime case

def datetime_tonumpydate(t, islocaltime=True):
   #  http://stackoverflow.com/questions/29753060/how-to-convert-numpy-datetime64-into-datetime
   return np.datetime64(t)

def datestring_todatetime(datelist1, format1= "%Y%m%d") :
 if isinstance(datelist1, str)  :  return parser.parse(datelist1)
 date2 = []
 for s in datelist1:
     date2.append(parser.parse(s))
     #date2.append(datetime.datetime.strptime(s, format1))
 return date2    

def datetime_toint(datelist1):
 if isinstance(datelist1, datetime.date) :      return int(datelist1.strftime("%Y%m%d"))
 date2 = []
 for t in datelist1:
     date2.append(int(t.strftime("%Y%m%d")))
 return date2

def date_holiday():
   '''
   https://jakevdp.github.io/blog/2015/07/23/learning-seattles-work-habits-from-bicycle-counts/

from pandas.tseries.holiday import USFederalHolidayCalendar
cal = USFederalHolidayCalendar()
holidays = cal.holidays('2012', '2016', return_name=True)
holidays.head()

holidays_all = pd.concat([holidays, "Day Before " + holidays.shift(-1, 'D'),  "Day After " + holidays.shift(1, 'D')])
holidays_all = holidays_all.sort_index()
holidays_all.head()

from pandas.tseries.offsets import CustomBusinessDay
from pandas.tseries.holiday import USFederalHolidayCalendar
bday_us = CustomBusinessDay(calendar=USFederalHolidayCalendar())
dateref[-1] - bday_us- bday_us

   '''

def date_add_bday(dint, nbday, country='us') :
  d= dateint_todatetime(dint)

  if country=='us'  :
     from pandas.tseries.offsets import CustomBusinessDay
     from pandas.tseries.holiday import USFederalHolidayCalendar
     bday_us = CustomBusinessDay(calendar=USFederalHolidayCalendar())
     d= d + bday_us*nbday
  else :
     d= d+ pd.tseries.offsets.BDay(nbday)

  return datetime_toint(d)

def dateint_todatetime(datelist1) :
 if isinstance(datelist1, int)  :  return parser.parse(str(datelist1))
 date2 = []
 for s in datelist1:
     date2.append(parser.parse(str(s)))
     #date2.append(datetime.datetime.strptime(s, format1))
 return date2    

def date_diffinday(intdate1, intdate2):
  dt= dateint_todatetime(intdate2) - dateint_todatetime(intdate1)
  return dt.days

def date_diffinyear(startdate, enddate):
 return date_as_float(startdate) - date_as_float(enddate)

def date_diffinbday( intd2, intd1)   :
  d1= dateint_todatetime(intd1)
  d2= dateint_todatetime(intd2)
  d1= d1.date()
  d2= d2.date()

  return np.busday_count(d1, d2)

def date_gencalendar(start='2010-01-01', end='2010-01-15', country='us') :
  from pandas.tseries.holiday import USFederalHolidayCalendar
  from pandas.tseries.offsets import CustomBusinessDay
  us_bd = CustomBusinessDay(calendar=USFederalHolidayCalendar())
  return np.arrray(pd.DatetimeIndex(start=start,end=end, freq=us_bd))

def date_finddateid(date1, dateref) :
  i= np_findfirst(date1, dateref)
  if i==-1 : i= np_findfirst(date1+1, dateref) 
  if i==-1 : i= np_findfirst(date1-1, dateref)     
  if i==-1 : i= np_findfirst(date1+2, dateref)    
  if i==-1 : i= np_findfirst(date1-2, dateref) 
  if i==-1 : i= np_findfirst(date1+3, dateref)    
  if i==-1 : i= np_findfirst(date1-3, dateref) 
  if i==-1 : i= np_findfirst(date1+5, dateref)    
  if i==-1 : i= np_findfirst(date1-5, dateref)  
  if i==-1 : i= np_findfirst(date1+7, dateref)    
  if i==-1 : i= np_findfirst(date1-7, dateref)  
  return i 
   
def datestring_toint(datelist1) :
 if isinstance(datelist1, str) :      return int(datelist1)    
 date2 = []
 for s in datelist1:   date2.append(int(s))
 return date2   

def date_now(i=0):
 from datetime import datetime
 d= datetime.now()
 if i > 0 : d= date_add_bdays(d,i)
 else:      d= date_remove_bdays(d,i)
 return str(datetime_toint(d))

def date_nowtime(type1='str', format1= "%Y-%m-%d %H:%M:%S:%f"):
 ''' str / stamp /  '''
 from datetime import datetime
 #d= datetime.now()
 d=datetime.today()  #today = datetime.today().strftime('%Y%m%d_%H%M%S%f')
 if   type1== 'str' :     return d.strftime(format1)
 elif type1== 'stamp' : return d.strftime('%Y%m%d_%H%M%S%f')
 else : return d

def date_tofloat(dt):
    size_of_day = old_div(1., 366.)
    size_of_second = old_div(size_of_day, (24. * 60. * 60.))
    days_from_jan1 = dt - datetime.datetime(dt.year, 1, 1)
    if not isleap(dt.year) and days_from_jan1.days >= 31+28:
        days_from_jan1 += timedelta(1)
    return dt.year + days_from_jan1.days * size_of_day + days_from_jan1.seconds * size_of_second

def date_generatedatetime(start="20100101", nbday=10, end=""):
  from dateutil.rrule import DAILY, rrule, MO, TU, WE, TH, FR
  start= datestring_todatetime(start)
  if end=="" :  end = date_add_bdays(start,nbday-1) #  + datetime.timedelta(days=nbday) 
  date_list= list(rrule(DAILY, dtstart=start, until=end, byweekday=(MO,TU,WE,TH,FR)))
  
  return np.array(date_list)





############################# Utilities for Numerical Calc ######################################
def np_numexpr_vec_calc(filename, expr, i0=0, imax=1000, fileout='E:\_data\_QUASI_SOBOL_gaussian_xx3.h5'  ):
 ''' New= xx*xx  over very large series
 #numexpr_vect_calc(filename, 0, imax=16384*4096, "xx*xx", 'E:\_data\_QUASI_SOBOL_gaussian_xx3.h5'  ):
'''
 pdframe=  pd.read_hdf(filename,'data', start=i0, stop=imax)    #from file
 xx= pdframe.values;  del pdframe    #to numpy vector
 xx= ne.evaluate(expr)  
 pdf =pd.DataFrame(xx); del xx  
# filexx3=   'E:\_data\_QUASI_SOBOL_gaussian_xx3.h5'   
 store = pd.HDFStore(fileout) 
 store.append('data', pdf); del pdf



def np_numexpr_tohdfs(filename, expr,  i0=0, imax=1000, fileout='E:\_data\_QUASI_SOBOL_gaussian_xx3.h5'  ):
 pdframe=  pd.read_hdf(filename,'data', start=i0, stop=imax)    #from file
 xx= pdframe.values;  del pdframe   
 xx= ne.evaluate(expr)  
 pdf =pd.DataFrame(xx); del xx    # filexx3=   'E:\_data\_QUASI_SOBOL_gaussian_xx3.h5' 
 store = pd.HDFStore(fileout);  store.append('data', pdf); del pdf

#numexpr_vect_calc(filename, 0, imax=16384*4096, "xx*xx", 'E:\_data\_QUASI_SOBOL_gaussian_xx3.h5'  ):


# yy1= getrandom_tonumpy('E:\_data\_QUASI_SOBOL_gaussian_xx2.h5', 16384, 4096)
#----------------------------------------------------------------------------



##################### Statistics    #################################################################
def np_comoment(xx,yy,nsample, kx,ky) :
#   cx= ne.evaluate("sum(xx)") /  (nsample);   cy= ne.evaluate("sum( yy)")  /  (nsample)
#   cxy= ne.evaluate("sum((xx-cx)**kx * (yy-cy)**ky)") / (nsample)
   cxy= old_div(ne.evaluate("sum((xx)**kx * (yy)**ky)"), (nsample))
   return cxy 

def np_acf(data):
    #Autocorrelation
    n = len(data)
    mean = np.mean(data)
    c0 = old_div(np.sum((data - mean) ** 2), float(n))

    def r(h):
        acf_lag = ((data[:n - h] - mean) * (data[h:] - mean)).sum() / float(n) / c0
        return acf_lag
    x = np.arange(n) # Avoiding lag 0 calculation
    acf_coeffs = np.asarray(list(map(r, x)))
    return acf_coeffs
 


##################### Plot Utilities ################################################################
def plot_XY(xx, yy, zcolor=None, tsize=None, title1='', xlabel='', ylabel='', figsize=(8, 6), dpi=75, savefile='') :

  #Color change
  if zcolor is None :  c= [[0, 0, 0]]
  elif isinstance(zcolor, int) : zcolor= zcolor
  else :
     aux= np.array(zcolor, dtype= np.float64)
     c= np.abs(aux)
  cmhot = plt.get_cmap("Blues")

  #Marker size
  if tsize is None :  tsize=  50
  elif isinstance(tsize, int) : tsize= tsize
  else :
     aux= np.array(tsize, dtype= np.float64)
     tsize=  np.abs(aux)
     tsize=(tsize - np.min(tsize)) / (np.max(tsize) - np.min(tsize)) * 130 + 1

  #Plot
  fig, ax1 = plt.subplots(nrows=1, ncols=1)

  fig.set_size_inches(figsize[0], figsize[1] )
  fig.set_dpi(dpi)
  fig.tight_layout()

  ax1.scatter(xx, yy, c=c, cmap=cmhot, s=tsize, alpha=0.5)
  ax1.set_xlabel(xlabel, fontsize=11)
  ax1.set_ylabel(ylabel, fontsize=11)
  ax1.set_title(title1)
  ax1.grid(True)
  # fig.autoscale(enable=True, axis='both')
  #fig.colorbar(ax1)

  if savefile != '' :
     plt.savefig(savefile, dpi=dpi)
     plt.clf()
  else :
     plt.show()


def plot_heatmap( frame,   ax=None, cmap=None,   vmin=None,vmax=None, interpolation='nearest'  ):
    from matplotlib import pyplot as plt
    if ax is None:
        ax = plt.gca()
    ax.set_xticks(np.arange(frame.shape[1]))
    ax.set_xticklabels(frame.columns, rotation='vertical')

    ax.set_yticks(np.arange(frame.shape[0]))
    ax.set_yticklabels(frame.index)
    ax.grid(False)
    ax.set_aspect('auto')
    ax.imshow(frame.values, interpolation=interpolation, cmap=cmap, vmin=vmin, vmax=vmax, aspect='auto')
    return



##########
def gc_map_dict_to_bq_schema(source_dict, schema, dest_dict):
    '''
     new_dict = {}
     map_dict_to_bq_schema (my_dict, schema, new_dict)

     new_dict = {}
     map_dict_to_bq_schema (my_dict, schema, new_dict)

    :param source_dict:
    :param schema:
    :param dest_dict:
    :return:
    '''
    #iterate every field from current schema
    for field in schema['fields']:
        #only work in existant values
        if field['name'] in source_dict:
            #nested field
            if field['type'].lower()=='record' and 'fields' in field:
                #list
                if 'mode' in field and field['mode'].lower()=='repeated':
                    dest_dict[field['name']] = []
                    for item in source_dict[field['name']]:
                        new_item = {}
                        map_dict_to_bq_schema( item, field, new_item )
                        dest_dict[field['name']].append(new_item)
                #record
                else:
                    dest_dict[field['name']] = {}
                    map_dict_to_bq_schema( source_dict[field['name']], field, dest_dict[field['name']] )
            #list
            elif 'mode' in field and field['mode'].lower()=='repeated':
                dest_dict[field['name']] = []
                for item in source_dict[field['name']]:
                    dest_dict[field['name']].append(item)
            #plain field
            else:
                dest_dict[field['name']]=source_dict[field['name']]

                format_value_bq(source_dict[field['name']], field['type'])




###################### Amazon #####################################################################
def aws_accesskey_get(access='', key='') :
   if access!= '' and key!='' : return access, key   
   # access, key= config.get('IAM', 'access'), config.get('IAM', 'secret')
   access, key= AWS_ACCESS_KEY_ID, AWS_SECRET_ACCESS_KEY
   #access, key=(boto.config.get('Credentials', 'aws_access_key_id'), boto.config.get('Credentials', 'aws_secret_access_key'))
   # print  access, key
   return access, key

def aws_conn_do(action='', region="ap-northeast-2") :
   con= aws_conn_create(region=region)
   pass

def aws_conn_getallregions(conn=None):
   return conn.get_all_regions()

def aws_conn_create(region="ap-northeast-2", access='', key=''):
   from boto.ec2.connection import EC2Connection

   if access== '' and key=='' : access, key= aws_accesskey_get()
   conn=    EC2Connection(access,key )
   regions= aws_conn_getallregions(conn)
   for r in regions:
      if r.name== region:
         conn=EC2Connection(access, key, region= r)
         return conn
   print('Region not Find')
   return None

def aws_conn_getinfo(conn):
   print(conn.region.name)

'''
def aws_ec2_instance_spot_stop(conn, inst):
   try:
      print 'Terminating', str(inst.id), '...',
      conn.terminate_instances(instance_ids=[inst.id])
      print 'done.'
      inst.remove_tag('Name', config.get('EC2', 'tag'))
   except:
      print 'Failed to terminate:', sys.exc_info()[0]
'''

'''
def aws_s3_file_putons3(fromfile, tobucket_path='bucket/folder1/folder2', AWS_KEY='', AWS_SECRET='' ) :
  from boto.s3.connection import S3Connection
  access, key= aws_accesskey_get()
  conn = S3Connection(access, key )

  tobucket, path1= aws_s3_url_split(tobucket_path)
  filename=        os_file_getname(fromfile)
  bucket = conn.get_bucket(tobucket)
  dest = bucket.new_key(path1+'/'+filename)
  dest.set_contents_from_file(fromfile)
'''


'''
def aws_s3_file_getfroms3(s3file='bucket/folder/perl_poetry.pdf', tofilename='/home/larry/myfile.pdf', AWS_KEY='', AWS_SECRET='' ):
  from boto.s3.connection import S3Connection
  #if access== '' and key=='' : access, key= aws_accesskey_get()
  access, key= aws_accesskey_get()
  conn = S3Connection(access, key )
  bucket, path1= aws_s3_url_split(s3file)
  bucket = conn.get_bucket(bucket)
  key = bucket.get_key(path1)
  key.get_contents_to_filename(tofilename)
'''

def aws_s3_url_split(url):
  '''Split into Bucket, url '''
  url1= url.split("/")
  return url1[0], "/".join(url1[1:])

def aws_s3_getbucketconn(s3dir) :
    import boto.s3
    bucket_name, todir= aws_s3_url_split(s3dir)
    ACCESS, SECRET= aws_accesskey_get()
    conn = boto.connect_s3(ACCESS, SECRET)
    bucket = conn.get_bucket(bucket_name)  #, location=boto.s3.connection.Location.DEFAULT)
    return bucket

def aws_s3_puto_s3(fromdir_file='dir/file.zip', todir='bucket/folder1/folder2') :
 ''' Copy File or Folder to S3 '''
 import boto.s3
 bucket= aws_s3_getbucketconn(todir)
 bucket_name, todir= aws_s3_url_split(todir)

 MAX_SIZE = 20 * 1000 * 1000
 PART_SIZE = 6 * 1000 * 1000

 if fromdir_file.find('.') > -1 :   #Case of Single File
    filename= os_file_getname(fromdir_file)
    fromdir_file=os_file_getpath(fromdir_file) + '/'
    uploadFileNames = [filename]
 else :
    uploadFileNames = []
    for (fromdir_file, dirname, filename) in os.walk(fromdir_file):
      uploadFileNames.extend(filename)
      break

 def percent_cb(complete, total):
    sys.stdout.write('.'); sys.stdout.flush()

 for filename in uploadFileNames:
    sourcepath = os.path.join(fromdir_file + filename)
    destpath =   os.path.join(todir, filename)
    print('Uploading %s to Amazon S3 bucket %s' % (sourcepath, bucket_name))

    filesize = os.path.getsize(sourcepath)
    if filesize > MAX_SIZE:
        print("multipart upload")
        mp = bucket.initiate_multipart_upload(destpath)
        fp = open(sourcepath,'rb')
        fp_num = 0
        while (fp.tell() < filesize):
            fp_num += 1
            print("uploading part %i" %fp_num)
            mp.upload_part_from_file(fp, fp_num, cb=percent_cb, num_cb=10, size=PART_SIZE)
        mp.complete_upload()
    else:
        print("singlepart upload: " + fromdir_file + ' TO ' + todir)
        k = boto.s3.key.Key(bucket)
        k.key = destpath
        k.set_contents_from_filename(sourcepath, cb=percent_cb, num_cb=10)

def aws_s3_getfrom_s3(froms3dir='task01/', todir='', bucket_name='zdisk') :
 ''' Get from S3 file/folder  '''
 bucket_name, dirs3= aws_s3_url_split(froms3dir)
 bucket= aws_s3_getbucketconn(froms3dir)
 bucket_list= bucket.list(prefix=dirs3)  #  /DIRCWD/dir2/dir3

 for l in bucket_list:
   key1=  str(l.key)
   file1, path2= os_file_getname(key1), os_file_getpath(key1)
   path1=  os.path.relpath(path2, dirs3).replace('.', '')  #Remove prefix path of S3 to mach
   d = todir + '/' + path1
   # print d, path2
   # sys.exit(0)
   if not os.path.exists(d): os.makedirs(d)
   try:
     l.get_contents_to_filename(d+ '/'+ file1)
   except OSError:
     pass

def aws_s3_folder_printtall(bucket_name='zdisk'):
   ACCESS, SECRET= aws_accesskey_get()
   conn = boto.connect_s3(ACCESS, SECRET)
   bucket = conn.create_bucket(bucket_name, location=boto.s3.connection.Location.DEFAULT)
   folders = bucket.list("","/")
   for folder in folders:
      print(folder.name)
   return folders

def aws_s3_file_read(filepath, isbinary=1) :
  ''' s3_client = boto3.client('s3')
    #Download private key file from secure S3 bucket
  s3_client.download_file('s3-key-bucket','keys/keyname.pem', '/tmp/keyname.pem')
  '''
  from boto.s3.connection import S3Connection
  conn = S3Connection(aws_accesskey_get( AWS_SECRET,AWS_KEY) )
  response = conn.get_object(Bucket=bucket1,Key=filename1)
  file1 = response["Body"]
  return file1





class aws_ec2_ssh(object):
    '''
    ssh= aws_ec2_ssh(host)
    print ssh.command('ls ')
    ssh.put_all( DIRCWD +'linux/batch/task/elvis_prod_20161220', EC2CWD + '/linux/batch/task' )
    ssh.get_all(  EC2CWD + '/linux/batch/task',  DIRCWD +'/zdisks3/fromec2' )

    # Detects DSA or RSA from key_file, either as a string filename or a file object.  Password auth is possible, but I will judge you for
    # ssh=SSHSession('targetserver.com','root',key_file=open('mykey.pem','r'))
    # ssh=SSHSession('targetserver.com','root',key_file='/home/me/mykey.pem')
    # ssh=SSHSession('targetserver.com','root','mypassword')
    # ssh.put('filename','/remote/file/destination/path')
    # ssh.put_all('/path/to/local/source/dir','/path/to/remote/destination')
    # ssh.get_all('/path/to/remote/source/dir','/path/to/local/destination')
    # ssh.command('echo "Command to execute"')
    '''
    def __init__(self,hostname,username='ubuntu',key_file=None,password=None):
        import paramiko,  socket
        from stat import S_ISDIR
        #  Accepts a file-like object (anything with a readlines() function)
        #  in either dss_key or rsa_key with a private key.  Since I don't
        #  ever intend to leave a server open to a password auth.
        self.host= hostname
        self.sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        self.sock.connect((hostname,22))
        self.t = paramiko.Transport(self.sock)
        self.t.start_client()
        #keys = paramiko.load_host_keys(os.path.expanduser('~/.ssh/known_hosts'))
        #key = self.t.get_remote_server_key()
        # supposed to check for key in keys, but I don't much care right now to find the right notation
        key_file= AWS_KEY_PEM
        pkey = paramiko.RSAKey.from_private_key_file(key_file)
        '''
        if key_file is not None:
            if isinstance(key,str):
                key_file=open(key,'r')
            key_head=key_file.readline()
            key_file.seek(0)
            if 'DSA' in key_head:
                keytype=paramiko.DSSKey
            elif 'RSA' in key_head:
                keytype=paramiko.RSAKey
            else:
                raise Exception("Can't identify key type")
            pkey=keytype.from_private_key(key_file)
            self.t.auth_publickey(username, pkey)
        else:
            if password is not None:
                self.t.auth_password(username,password,fallback=False)
            else: raise Exception('Must supply either key_file or password')
        '''
        self.t.auth_publickey(username, pkey)
        self.sftp=paramiko.SFTPClient.from_transport(self.t)
        print(self.command('ls '))

    def command(self,cmd):
        #  Breaks the command by lines, sends and receives
        #  each line and its output separately
        #  Returns the server response text as a string

        chan = self.t.open_session()
        chan.get_pty()
        chan.invoke_shell()
        chan.settimeout(20.0)
        ret=''
        try:
            ret+=chan.recv(1024)
        except:
            chan.send('\n')
            ret+=chan.recv(1024)
        for line in cmd.split('\n'):
            chan.send(line.strip() + '\n')
            ret+=chan.recv(1024)
        return ret

    def put(self,localfile,remotefile):
        #  Copy localfile to remotefile, overwriting or creating as needed.
        self.sftp.put(localfile,remotefile)

    def put_all(self,localpath,remotepath):
        #  recursively upload a full directory
        localpath= localpath[:-1] if localpath[-1]=='/' else localpath
        remotepath= remotepath[:-1] if remotepath[-1]=='/' else remotepath

        os.chdir(os.path.split(localpath)[0])
        parent=os.path.split(localpath)[1]
        for walker in os.walk(parent):
            try:
                self.sftp.mkdir(os.path.join(remotepath,walker[0]).replace('\\',"/"))
            except:
                pass
            for file in walker[2]:
                print((os.path.join(walker[0],file).replace('\\','/').replace('\\','/'),  os.path.join(remotepath,walker[0],file).replace('\\','/')))
                self.put(os.path.join(walker[0],file).replace('\\','/'),      os.path.join(remotepath,walker[0],file).replace('\\','/'))

    def get(self,remotefile,localfile):
        #  Copy remotefile to localfile, overwriting or creating as needed.
        self.sftp.get(remotefile,localfile)

    def sftp_walk(self,remotepath):
        import paramiko,  socket
        from stat import S_ISDIR
        # Kindof a stripped down  version of os.walk, implemented for
        # sftp.  Tried running it flat without the yields, but it really
        # chokes on big directories.
        path=remotepath
        files=[]
        folders=[]
        for f in self.sftp.listdir_attr(remotepath):
            if S_ISDIR(f.st_mode):
                folders.append(f.filename)
            else:
                files.append(f.filename)
        print((path,folders,files))
        yield path,folders,files
        for folder in folders:
            new_path=os.path.join(remotepath,folder).replace('\\','/')
            for x in self.sftp_walk(new_path):
                yield x

    def get_all(self,remotepath,localpath):
        #  recursively download a full directory
        #  Harder than it sounded at first, since paramiko won't walk
        # For the record, something like this would gennerally be faster:
        # ssh user@host 'tar -cz /source/folder' | tar -xz
        localpath= localpath[:-1] if localpath[-1]=='/' else localpath
        remotepath= remotepath[:-1] if remotepath[-1]=='/' else remotepath

        self.sftp.chdir(os.path.split(remotepath)[0])
        parent=os.path.split(remotepath)[1]
        try:
            os.mkdir(localpath)
        except:
            pass
        for walker in self.sftp_walk(parent):
            try:
                os.mkdir(os.path.join(localpath,walker[0]).replace('\\','/') )
            except:
                pass
            for file in walker[2]:
                print((   os.path.join(walker[0],file).replace('\\','/')   ,  os.path.join(localpath,walker[0],file).replace('\\','/')))
                self.get(os.path.join(walker[0],file).replace('\\','/')   ,  os.path.join(localpath,walker[0],file).replace('\\','/'))


    def write_command(self,text,remotefile):
        #  Writes text to remotefile, and makes remotefile executable.
        #  This is perhaps a bit niche, but I was thinking I needed it.
        #  For the record, I was incorrect.
        self.sftp.open(remotefile,'w').write(text)
        self.sftp.chmod(remotefile,755)

    def python_script(self, script_path, args1):
      ipython_script='/home/ubuntu/anaconda2/bin/ipython '
      cmd1= ipython_script +' ' + script_path + ' ' + '"'+args1 +'"'
      self.cmd2(cmd1)
      # self.command(cmd1)

    def command_list(self, cmdlist) :
      for command in cmdlist:
        print("Executing {}".format(command))
        ret= self.command(cmd1)
        print(ret)
      print('End of SSH Command')

    def listdir(self, remotedir):
       return self.sftp.listdir(remotedir)

    def jupyter_kill(self):
       pid_jupyter= aws_ec2_cmd_ssh(cmdlist=  ['fuser 8888/tcp'], host=host ,doreturn=1)[0][0].strip()
       print(ssh.command('kill -9 '+pid_jupyter))

    def jupyter_start(self):
        pass

    def cmd2(self, cmd1):
        return aws_ec2_cmd_ssh(cmdlist=  [cmd1], host= self.host,doreturn=1)



    def _help_ssh(self):
       s='''
         fuser 8888/tcp     Check if Jupyter is running
           ps -ef | grep python     :List of  PID Python process
          kill -9 PID_number     (i.e. the pid returned)
        top     : CPU usage
       '''
       print(s)


def aws_ec2_cmd_ssh(cmdlist=  ["ls " ],  host='ip', doreturn=0, ssh=None, username='ubuntu', keyfilepath='') :
    ''' SSH Linux terminal Command
     https://www.siteground.com/tutorials/ssh/ssh_deleting.htm

     rm -rf foldername/


    fuser 8888/tcp     Check if Jupyter is running
    ps -ef | grep python     :List of  PID Python process
    kill -9 PID_number     (i.e. the pid returned)
    top     : CPU usage

      Run nohup python bgservice.py & to get the script to ignore the hangup signal and keep running.
      Output will be put in nohup.out.
        "aws s3 cp s3://s3-bucket/scripts/HelloWorld.sh /home/ec2-user/HelloWorld.sh",
        "chmod 700 /home/ec2-user/HelloWorld.sh",
        "/home/ec2-user/HelloWorld.sh"

    https://aws.amazon.com/blogs/compute/scheduling-ssh-jobs-using-aws-lambda/
   '''
    if ssh is None and len(host) > 5 :
      ssh= aws_ec2_create_con(contype='ssh', host=host, port=22, username=username, keyfilepath='')
      print('EC2 connected')

    c= cmdlist
    if isinstance(c, str) :  #Only Command to be launched
       if   c== 'python'   : cmdlist= [ 'ps -ef | grep python ' ]
       elif c== 'jupyter'  : cmdlist= [ 'fuser 8888/tcp  ' ]

    readall=[]
    for command in cmdlist:
        print("Executing {}".format(command))
        stdin , stdout, stderr = ssh.exec_command(command)
        outread, erread= stdout.read(), stderr.read()
        readall.append((outread, erread))
        print(outread); print(erread)
    print('End of SSH Command')
    ssh.close()
    if doreturn: return readall

def aws_ec2_python_script(script_path, args1, host):
   ipython_script='/home/ubuntu/anaconda2/bin/ipython'  #!!! No space after ipython
   cmd1= ipython_script +' ' + script_path + ' ' + '"'+args1 +'"'
   aws_ec2_cmd_ssh(cmdlist=  [cmd1], ssh=None, host=host, username='ubuntu')

def aws_ec2_create_con(contype='sftp/ssh', host='ip', port=22, username='ubuntu',  keyfilepath='', password='',keyfiletype='RSA', isprint=1):
    """ Transfert File  host = '52.79.79.1'
        keyfilepath = 'D:/_devs/aws/keypairs/ec2_instanc'

# List files in the default directory on the remote computer.
dirlist = sftp.listdir('.')
sftp.get('remote_file.txt', 'downloaded_file.txt')
sftp.put('testfile.txt', 'remote_testfile.txt')

http://docs.paramiko.org/en/2.1/api/sftp.html
    """
    import paramiko
    sftp,ssh,  transport= None, None,  None
    try:
        if keyfilepath=='': keyfilepath= AWS_KEY_PEM
        if keyfiletype == 'DSA':  key = paramiko.DSSKey.from_private_key_file(keyfilepath)
        else:                     key = paramiko.RSAKey.from_private_key_file(keyfilepath)

        if contype== 'sftp' :
          # Create Transport object using supplied method of authentication.
          transport = paramiko.Transport((host, port))
          transport.add_server_key(key)
          transport.connect(None, username,  pkey=key)
          sftp = paramiko.SFTPClient.from_transport(transport)
          if isprint : print(('Root Directory :\n ', sftp.listdir()))
          return sftp

        if contype== 'ssh' :
          ssh = paramiko.SSHClient()
          ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
          ssh.connect( hostname = host, username = username, pkey = key )

          # Test
          if isprint :
            stdin, stdout, stderr = ssh.exec_command("uptime;ls -l")
            stdin.flush()   #Execute
            data = stdout.read().splitlines()   #Get data
            print('Test Print Directory ls :')
            for line in data: print(line)
          return ssh

    except Exception as e:
        print('An error occurred creating client: %s: %s' % (e.__class__, e))
        if sftp is not None:      sftp.close()
        if transport is not None: transport.close()
        if ssh is not None: ssh.close()



def aws_ec2_allocate_elastic_ip(instance_id, region="ap-northeast-2") :
 con=  aws_conn_create(region=region)
 eip =con.allocate_address()
 con.associate_address(instance_id=instance_id, public_ip=eip.public_ip)
 print('Elastic assigned Public IP: '+eip.public_ip,  ',Instance_ID:', instance_id)
 return eip.public_ip



'''
def aws_ec2_jupyter(host):
   Launch Jupyter server with in Backgroun Process
   At the terminal execute which jupyter. Observe the path. In both local terminal and connecting via ssh, execute echo $PATH.
Ensure that the path to jupyter is included in the $PATH in your environment. – user4556274 10 hours ago

The simple stuff
PATH=$PATH:~/opt/bin
PATH=~/opt/bin:$PATH
depending on whether you want to add ~/opt/bin at the end (to be searched after all other directories, in case
there is a program by the same name in multiple directories) or at the beginning (to be searched before all other directories).

   aws_ec2_cmd_ssh(['nohup /home/ubuntu/anaconda2/bin/jupyter notebook'], host=host, username='ubuntu')


def aws_ec2_getfrom_ec2( fromfolder, tofolder, host) :
   sftp= aws_ec2_create_sftp(contype='sftp', host=host)

   if fromfolder.find('.') > -1 :   # file
      folder1, file1= z_key_splitinto_dir_name(fromfolder[:-1] if fromfolder[-1]=='/' else  fromfolder)
      tofolder2=      tofolder if tofolder.find(".") > -1 else  tofolder + '/'+file1
      sftp.get(fromfolder, tofolder2)

   else :   #Pass the Folder in Loop
     pass

def aws_ec2_putfolder(fromfolder='D:/_devs/Python01/project27//linux/batch/task/elvis_prod_20161220/', tofolder='/linux/batch', host='') :
  # fromfolder= DIRCWD +'/linux/batch/task/elvis_prod_20161220/'
  # tofolder=   '/linux/batch/'

  # If you don't care whether the file already exists and you always want to overwrite the files as they are extracted without prompting use the -o switch as follows:
  # https://www.lifewire.com/examples-linux-unzip-command-2201157
  # unzip -o filename.zip

  # sftp.mkdir(remotedirectory)
  # sftp.chdir(remotedirectory)
  # sftp.put(localfile, remotefile)
  folder1, file1= z_key_splitinto_dir_name(fromfolder[:-1] if fromfolder[-1]=='/' else  fromfolder)

  tofolderfull=  EC2CWD + '/' + tofolder  if tofolder.find(EC2CWD) == -1 else tofolder

  #Zip folder before sending it
  file2= folder1+ '/' + file1+'.zip'
  os_zipfolder(fromfolder, file2)
  print(aws_ec2_put(file2, tofolder= tofolderfull, host=host, typecopy='all'))

  #   Need install sudo apt-get install zip unzip
  cmd1= '/usr/bin/unzip '+ tofolderfull+'/'+file1+'.zip ' + ' -d ' +  tofolderfull +'/'
  aws_ec2_cmd_ssh(cmdlist=  [cmd1], host=host)

def aws_ec2_put(fromfolder='d:/ file1.zip', tofolder='/home/notebook/aapackage/', host='', typecopy='code') :
  Copy python code, copy specific file, copy all folder content
  :param fromfolder: 1 file or 1 folder
  :param tofolder:
  :param host:


  sftp= aws_ec2_create_con('sftp', host, isprint=1)

  if fromfolder.find('.') > -1  :    # Copy 1 file
     if fromfolder.find(':') == -1 : print('Please put absolute path'); return 0

     fromfolder, file1= os_split_dir_file(fromfolder)
     tofull=  tofolder + '/'+ file1  if tofolder.find('.') == -1 else  tofolder

     sftp.put(fromfolder+'/'+file1, tofull)
     try :
          sftp.stat(tofull);   isexist= True
     except: isexist= False
     sftp.close();  return (isexist, tofull)

  else :  #Copy Folder to



    if typecopy == 'code' and fromfolder.find('.') == -1  :    #Local folder and code folder
      foldername= fromfolder
      fromfolder= DIRCWD+ '/' + foldername
      tempfolder= DIRCWD+'/zdisks3/toec2_notebook/'+foldername
      os_folder_delete(tempfolder)
      os_folder_copy(fromfolder,  tempfolder,  pattern1="*.py")
      sftp.put(tempfolder, tofolder)
      return 1

    if typecopy== 'all' :
      if fromfolder.find(':') == -1 : print('Please put absolute path'); return 0
      if fromfolder.find('.') > -1 :  #1 file
       fromfolder, file1= os_split_dir_file(fromfolder)
       tofull= tofolder + '/'+ file1  if tofolder.find('.') == -1 else tofolder
       tofolder, file2= os_split_dir_file(tofull)

       sftp.put(fromfolder+'/'+file1, tofull)

       try :
          sftp.stat(tofull);   isexist= True
       except: isexist= False
       print(isexist, tofull)
'''






'''
http://boto.cloudhackers.com/en/latest/s3_tut.html

Storing Large Data¶

At times the data you may want to store will be hundreds of megabytes or more in size. S3 allows you to split such files into smaller components. You upload each component in turn and then S3 combines them into the final object. While this is fairly straightforward, it requires a few extra steps to be taken. The example below makes use of the FileChunkIO module, so pip install FileChunkIO if it isn’t already installed.

import math, os
import boto
from filechunkio import FileChunkIO

# Connect to S3
c = boto.connect_s3()
b = c.get_bucket('mybucket')

# Get file info
source_path = 'path/to/your/file.ext'
source_size = os.stat(source_path).st_size

# Create a multipart upload request
mp = b.initiate_multipart_upload(os.path.basename(source_path))

# Use a chunk size of 50 MiB (feel free to change this)
chunk_size = 52428800
chunk_count = int(math.ceil(source_size / float(chunk_size)))

# Send the file parts, using FileChunkIO to create a file-like object
# that points to a certain byte range within the original file. We
# set bytes to never exceed the original file size.
for i in range(chunk_count):
    offset = chunk_size * i
    bytes = min(chunk_size, source_size - offset)
    with FileChunkIO(source_path, 'r', offset=offset,
                         bytes=bytes) as fp:
        mp.upload_part_from_file(fp, part_num=i + 1)

# Finish the upload
mp.complete_upload()
'''


########################### Google Drive    ###############################################################
def googledrive_get():
   '''
   https://github.com/ctberthiaume/gdcp
   ... I am using this now to transfer thousands of mp3 files from a ubuntu vps to google drive.


http://olivermarshall.net/how-to-upload-a-file-to-google-drive-from-the-command-line/
https://github.com/prasmussen/gdrive  : Super Complete

gdrive [global] upload [options] <path>

global:
  -c, --config <configDir>         Application path, default: /Users/<user>/.gdrive
  --refresh-token <refreshToken>   Oauth refresh token used to get access token (for advanced users)
  --access-token <accessToken>     Oauth access token, only recommended for short-lived requests because of short lifetime (for advanced users)

options:
  -r, --recursive           Upload directory recursively
  -p, --parent <parent>     Parent id, used to upload file to a specific directory, can be specified multiple times to give many parents
  --name <name>             Filename
  --no-progress             Hide progress
  --mime <mime>             Force mime type
  --share                   Share file
  --delete                  Delete local file when upload is successful
  --chunksize <chunksize>   Set chunk size in bytes, default: 8388608

   :return:
   '''
   pass

def googledrive_put():
  '''
  100 GB: 2USD,  1TB: 10USD
  https://gsuite.google.com/intl/en/pricing.html

  :return:
  '''
  pass


def googledrive_list():
   pass



def os_processify_fun(func):
    '''Decorator to run a function as a process.
    Be sure that every argument and the return value is *pickable*.
    The created process is joined, so the code does not  run in parallel.
    @processify

    def test():
      return os.getpid()

    @processify
    def test_deadlock():
      return range(30000)

   @processify
   def test_exception():
     raise RuntimeError('xyz')

   def test():
     print os.getpid()
     print test_function()
     print len(test_deadlock())
     test_exception()

   if __name__ == '__main__':
     test()

    '''
    import sys
    import traceback
    from functools import wraps
    from multiprocessing import Process, Queue

    def process_func(q, *args, **kwargs):
        try:
            ret = func(*args, **kwargs)
        except Exception:
            ex_type, ex_value, tb = sys.exc_info()
            error = ex_type, ex_value, ''.join(traceback.format_tb(tb))
            ret = None
        else:
            error = None

        q.put((ret, error))

    # register original function with different name
    # in sys.modules so it is pickable
    process_func.__name__ = func.__name__ + 'processify_func'
    setattr(sys.modules[__name__], process_func.__name__, process_func)

    @wraps(func)
    def wrapper(*args, **kwargs):
        q = Queue()
        p = Process(target=process_func, args=[q] + list(args), kwargs=kwargs)
        p.start()
        ret, error = q.get()
        #p.join()

        if error:
            ex_type, ex_value, tb_str = error
            message = '%s (in subprocess)\n%s' % (ex_value.message, tb_str)
            raise ex_type(message)

        return ret
    return wrapper


@os_processify_fun
def ztest_processify():
    return os.getpid()


















def date_getspecificdate(datelist, datetype1="yearend", outputype1="intdate", includelastdate=True, includefirstdate=False, ):
 vec2= []

 if isint(datelist[0] )  :  datelist= dateint_todatetime(datelist)

 t0 = datelist[0]
 if datetype1== "monthend" :
    for i,t in enumerate(datelist):
#      print(datetime_tostring([t0, t]))
      if t.month != t0.month : vec2.append([i-1,t0]);  # month has change
      t0= t

 if datetype1== "2monthend" :
    for i,t in enumerate(datelist):
      if t.month != t0.month and np.mod(t0.month,2) ==0 : vec2.append([i-1,t0]);  # month has change
      t0= t

 if datetype1== "3monthend" :
    for i,t in enumerate(datelist):
      if t.month != t0.month and np.mod(t0.month,3) ==0 : vec2.append([i-1,t0]);  # month has change
      t0= t

 if datetype1== "4monthend" :
    for i,t in enumerate(datelist):
      if t.month != t0.month and np.mod(t0.month,4) ==0 : vec2.append([i-1,t0]);  # month has change
      t0= t

 if datetype1== "6monthend" :
    for i,t in enumerate(datelist):
      if t.month != t0.month and np.mod(t0.month,6) ==0 : vec2.append([i-1,t0]);  # month has change
      t0= t

 if datetype1== "monthstart" :
    for i,t in enumerate(datelist):
      if t.month != t0.month : vec2.append([i,t]);   # month has change
      t0= t

 if datetype1== "yearstart" :
    vec2.append([0,t0])
    for i,t in enumerate(datelist):
      if t.year != t0.year: vec2.append([i,t]);   # month has change
      t0= t

 if datetype1== "yearend" :
    for i,t in enumerate(datelist):
      if t.year != t0.year : vec2.append([i-1,t0]);   # month has change
      t0= t

 if includelastdate :
      vec2.append([len(datelist)-3, datelist[-1]])

 if outputype1== "intdate" :
   vec2= np.array(vec2)
   vec2= np.array(vec2[:,0], dtype="int")
   return vec2
 else :
   return np.array(vec2)