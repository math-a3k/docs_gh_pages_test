# -*- coding: utf-8 -*-
#---------Various Utilities function for Python--------------------------------------
from __future__ import division
from __future__ import print_function
from future import standard_library
standard_library.install_aliases()
from builtins import next;      from builtins import map
from builtins import zip;       from builtins import str
from builtins import range;     from past.builtins import basestring
from past.utils import old_div; from builtins import object

import os, sys
if sys.platform.find('win') > -1 :
   print("")
   # from guidata import qthelpers  #Otherwise Error with Spyder Save

import datetime, time, arrow,  shutil,  IPython, gc, copy
import matplotlib.pyplot as plt
import numexpr as ne, numpy as np, pandas as pd, scipy as sci
import urllib3
from bs4 import BeautifulSoup
from numba import jit, float32

#####################################################################################################
import os, sys
#CFG   = {'plat': sys.platform[:3]+"-"+os.path.expanduser('~').split("\\")[-1].split("/")[-1], "ver": sys.version_info.major}
#DIRCWD= {'win-asus1': 'D:/_devs/Python01/project27/', 'win-unerry': 'G:/_devs/project27/' , 'lin-noel': '/home/noel/project27/', 'lin-ubuntu': '/home/ubuntu/project27/' }[CFG['plat']]
#print(os.environ)
#DIRCWD= os.environ["DIRCWD"];

import configmy; CFG, DIRCWD= configmy.get(config_file="_ROOT", output= ["_CFG", "DIRCWD"])
os.chdir(DIRCWD); sys.path.append(DIRCWD + '/aapackage')


__path__=     DIRCWD +'/aapackage/'
__version__=  "1.0.0"
__file__=     "util.py"








'''
Run Unit test   python util.py --do  test01


#### System Variables == os.environ, Admin right
setx DIRCWD "D:/_devs/Python01/project27/"   /M


This is actually already done for you in config._sections. Example:

$ cat test.ini
[DIRCWD]
win-asus1 = D:/_devs/Python01/project27/
key = item

[Second Section]
othervar = othervalue
otherkey = otheritem
And then:

from ConfigParser import ConfigParser;config = ConfigParser(); config.read('config.ini'); config_dict= {s:dict(config.items(s)) for s in config.sections()}
config_dict= []


PLOTLY_APIKEY=    'aBq2k5nrzLK0chV5T32x'
PLOTLY_USERNAME=  'kevinno'

import plotly.tools as tls
tls.set_credentials_file(username=PLOTLY_USERNAME, api_key=PLOTLY_APIKEY)
'''


####################################################################################################

'''
https://console.developers.google.com/iam-admin/projects
Google Drive API
https://googledrive.github.io/PyDrive/docs/_build/html/quickstart.html#authentication

https://console.developers.google.com/iam-admin/quotas?project=refreshing-code-142020

Create an empty file called __init__.py in all you folders. Then you can import using . as a folder separator. Documentation here.


#--------------------Utilities-------------------------------------------------------------------
# http://win32com.goermezer.de/content/blogsection/7/284/

WhenPackage is destroyed resintall and copy paste the data.
from saved package

ps://pythonconquerstheuniverse.wordpress.com/2009/10/15/python-packages/

import spyderlib
74	os.environ["SPYDER_PARENT_DIR"] = os.path.abspath(os.path.join(spyderlib.__file__, "../.."))
…	
199	from spyderlib.utils.programs import (run_python_script, is_module_installed,
200	                                      start_file, run_python_script_in_terminal)
201	from spyderlib.utils.iofuncs import load_session, save_session, reset_session


https://service.wi2.ne.jp/wi2net/SbjLogin/2/

klepto is built on top of dill, and can cache function inputs and outputs to memory (so that calculations don't need to be run twice), and it can seamlessly persist objects in the cache to disk or to a database.
The versions on github are the ones you want. https://github.com/uqfoundation/klepto or https://github.com/joblib/joblib. Klepto is newer, but has a much broader set of caching and archiving solutions than joblib. Joblib has been in production use longer, so it's better tested -- especially for parallel computing.
Here's an example of typical klepto workflow: https://github.com/uqfoundation/klepto/blob/master/tests/test_workflow.py
Here's another that has some numpy in it: https://github.com/uqfoundation/klepto/blob/master/tests/test_cache.py

'''

'''
Unpacking dictionaries

** unpacks dictionaries.

func(a=1, b=2, c=3)
is the same as

args = {'a': 1, 'b': 2, 'c':3}
func(**args)
It's useful if you have to construct parameters:

args = {'name': person.name}
if hasattr(person, "address"):
    args["address"] = person.address
func(**args)  # either expanded to func(name=person.name) or
              #                    func(name=person.name, address=person.address)
Packing parameters of a function

def setstyle(**styles):
    for key, value in styles.iteritems():      # styles is a regular dictionary
        setattr(someobject, key, value)
This lets you use the function like this:

setstyle(color="red", bold=False)

'''



############## #Serialize Python Session #######################################################
# https://github.com/uqfoundation/dill

class testclass(object):
   def __init__(self,x) :
       pass

   def z_autotest(self) :
     import sys, io
     # create file-like string to capture output
     codeOut = io.StringIO()
     codeErr = io.StringIO()

     code = """
         def f(x):
         x = x + 1
         return x
         print 'This is my output.'
     """

     # capture output and errors
     sys.stdout = codeOut
     sys.stderr = codeErr

     exec(code)

     # restore stdout and stderr
     sys.stdout = sys.__stdout__
     sys.stderr = sys.__stderr__

     print(f(4))
     s = codeErr.getvalue(); print("error:\n%s\n" % s)
     s = codeOut.getvalue(); print("output:\n%s" % s)
     codeOut.close()
     codeErr.close()



def session_load_function(name='test_20160815'):
 import dill
 n1= DIRCWD+ '/aaserialize/session/'+ name + '.pkl'

 dill.load_session(n1)
 print(n1)



def session_save_function(name='test'):
 import dill, pickle
 t1= date_now()
 n1= DIRCWD+ '/aaserialize/session/dill_session_'+ name + '_'+t1+'.pkl'

 dill.dump_session(n1)
 print(n1)


def py_save_obj_dill(obj1, keyname='', otherfolder=0) :
   import dill, pickle, numpy, pandas
   dir0, keyname= z_key_splitinto_dir_name(keyname)
   os_folder_create(DIRCWD+'/aaserialize/' + dir0)
   dir1= DIRCWD+'/aaserialize/' + dir0 + '/'+ keyname + '.pkl' if otherfolder==0 else keyname

   type_list= [numpy, pandas.core.series, dill.source.getmodule(int), dill.source.getmodule(str), dill.source.getmodule(float)]
   name_type= []

   type1= dill.source.getmodule(type(obj1))
   name1= ''
   if not name1 in type_list and not type1 in type_list :
     with open( dir1, 'wb') as f:
       dill.dumps(obj1, protocol=pickle.HIGHEST_PROTOCOL)
   else :
       print('Primitive type not dill saved')
   return dir1


def session_spyder_showall():
  ls= os_file_listall(DIRCWD+ '/aaserialize/session/', pattern="session*.pkl", dirlevel=2)
  lsname= ls[:,0]
  for s in lsname:
    print(s)

def session_guispyder_save(filename):
 from spyderlib.utils.iofuncs import save_session
 save_session(filename+'.session.tar')

def session_guispyder_load(filename):
 from spyderlib.spyder import load_session
 load_session(filename+'.session.tar')
 '''
  spyderlib/spyder.lib
    #---- Sessions
    def load_session(self, filename=None):
        """Load session"""
        if filename is None:
            self.redirect_internalshell_stdio(False)
            filename, _selfilter = getopenfilename(self, _("Open session"),
                        getcwd(), _("Spyder sessions")+" (*.session.tar)")
            self.redirect_internalshell_stdio(True)
            if not filename:
                return
        if self.close():
            self.next_session_name = filename


 '''


def session_load(filename, dict1=None, towhere='main'):
  ''' .spydata file,  dict1: already provided Dict,  towhere= main, function, dict '''
  from spyderlib.utils.iofuncs import load_dictionary
  print('Loading : ', end=' ')
  if dict1 is None :
    if filename.find('.') == -1 : #Use default folder
        dir0, keyname= z_key_splitinto_dir_name(filename)
        os_folder_create(DIRCWD+ '/aaserialize/session/' + dir0)
        filename=DIRCWD + '/aaserialize/session/' + dir0 + '/' + keyname + '.spydata'
    print(filename)
    data = load_dictionary(filename)

  if towhere=='main' :  #Load in interpreter module
    module = sys.modules['__main__']
    for k, v in list(data[0].items()):
       setattr(module, k, v)
       print(k, end=' ')

  elif towhere=='function' :
      from inspect import getmembers, stack
      globals1 = dict(getmembers(stack()[1][0]))["f_globals"]

  elif towhere=='dict' :
    return data


def session_save(filename="/folder1/name1", globals1=None):
    '''Need to pass globals() Cannot Get Save data to .spydata file

   BIG issue with Import, Impor FULL MODULE ----> BIG ISSUE
     BIG ISSUE with DICT, USE LIST INSTEAD
        If you try to put this code in a module and import the function then you will have to pass globals() to the function explicitly as the globals() in the function is not the IPython global namespace. However, you can put the above code inside your ~/.ipython/profile_PROFILE/startup/startup.ipy file and it will work as expected.
       PROFILE is the name of the profile that you plan to start IPython with.
    '''
    #from spyderlib.utils.iofuncs import save_dictionary
    #from spyderlib.widgets.dicteditorutils import globalsfilter
    #from spyderlib.plugins.variableexplorer import VariableExplorer
    #from spyderlib.baseconfig import get_supported_types

    # Local imports

    import spyderlib.utils.iofuncs as iofunc
    import spyderlib.widgets.dicteditorutils as spyutil
    import spyderlib.plugins.variableexplorer as spyvar
    import spyderlib.baseconfig as spybas

    if globals1 is None :  #Extract Global from the caller, Be careful can have issue
      print('Error, Please 2nd Arguments    (...,globals()')
      return None
      #from inspect import getmembers, stack
      #globals1 = dict(getmembers(stack()[1][0]))["f_globals"]
    
    settings = spyvar.VariableExplorer.get_settings()
    data = spyutil.globalsfilter(globals1, check_all=True, filters=tuple(spybas.get_supported_types()['picklable']),
                         exclude_private=settings['exclude_private'],
                         exclude_uppercase=settings['exclude_uppercase'],
                         exclude_capitalized=settings['exclude_capitalized'],
                         exclude_unsupported=settings['exclude_unsupported'],
                         excluded_names=settings['excluded_names']+['settings','In'])

    if filename.find('.') == -1 : #Use default folder
        dir0, keyname= z_key_splitinto_dir_name(filename)
        os_folder_create(DIRCWD+ '/aaserialize/session/' + dir0)
        filename= DIRCWD+'/aaserialize/session/' + dir0 + '/'+ keyname + '.spydata'

    iofunc.save_dictionary(data,filename)
    os.chdir(DIRCWD)
    print(('Saved: '+filename))


'''
#Used for Spyder Load /Session
def spyder_save_gui_session(name1) :
 path= DIRCWD+'/'+name1
 os.makedirs(path)

 savepath = path+'/'+name1+'.spydata'
 session_save(savepath)
 spyder_savesession( path+'/'+name1+'.session.tar')

 shutil.make_archive(name1, 'zip',  root_dir=path, base_dir=None)
 shutil.rmtree(path)

def spyder_load_gui_session(name1) :
 import zipfile
 path= DIRCWD
 path2= DIRCWD+'/'+name1
 #shutil.unpack_archive(path+'/'+name1+'.zip')
 
 with zipfile.ZipFile(path+'/'+name1+'.zip', 'r') as myzip:
  myzip.extractall(  path2)
 
 spyder_loadsession(path2 + '/'+name1+'.session.tar')
 session_load(path2 + '/' + name1 + '.spydata')
 shutil.rmtree(path2)
'''



#####################################################################################
 #-------- Python General---------------------------------------------------------
#  Import File
# runfile('D:/_devs/Python01/project27/stockMarket/google_intraday.py', wdir='D:/_devs/Python01/project27/stockMarket')

def aa_unicode_ascii_utf8_issue():
   '''Take All csv in a folder and provide Table, Column Schema, type

 METHOD FOR Unicode / ASCII issue
1. Decode early
Decode to <type 'unicode'> ASAP
df['PREF_NAME']=       df['PREF_NAME'].apply(to_unicode)

2. Unicode everywhere


3. Encode late
>>> f = open('/tmp/ivan_out.txt','w')
>>> f.write(ivan_uni.encode('utf-8'))

Important methods
s.decode(encoding)  <type 'str'> to <type 'unicode'>
u.encode(encoding)  <type 'unicode'> to <type 'str'>

http://farmdev.com/talks/unicode/

   '''


def isexist(a) :
 try:
   a; return True
 except NameError: 
  return False

def isfloat(x):
  try:    
   v=   False if x == np.inf else True
   float(x)
   return v
  except :    return False
      
def isint(x): return isinstance(x, ( int, np.int, np.int64, np.int32 ) )

def a_isanaconda():
 import sys; 
 txt= sys.version
 if txt.find('Continuum') > 0 : return True
 else: return False

if a_isanaconda() :  #Import the Packages
   pass
  

def a_run_ipython(cmd1):
 ''' Execute Ipython Command in python code
     run -i :  run including current interprete variable
 '''
 IPython.get_ipython().magic(cmd1)

def a_autoreload() :
 a_run_ipython("load_ext autoreload"); a_run_ipython("autoreload 2")

def a_start_log(id1='', folder='aaserialize/log/')  :
 a_run_ipython('logstart -o -r -t ' + folder + 'log' + str(id1) + '_' + a_get_platform() + '_' + date_now() + ' rotate')

def a_cleanmemory():
  import gc;  gc.collect() 

def a_module_codesample(module_str='pandas') :
  dir1= 'D:/_devs/Python01/project27/docs/'
  file1= dir1+'/' + module_str + '/codesample.py'
  txt= os_file_read(file1)
  os_gui_popup_show(txt)

def a_module_doc(module_str='pandas') :
  dir1= 'D:/_devs/Python01/project27/docs/'
  file1= dir1+'/' + module_str + '/doc.py'
  txt= os_file_read(file1)
  os_gui_popup_show(txt)

def a_module_generatedoc(module_str="pandas", fileout=''):
  ''' #  getmodule_doc("jedi", r"D:\_devs\Python01\aapackage\doc.txt")'''
  from .automaton import codeanalysis as ca
  #pathout= DIRCWD+'/docs/'+ module1.__name__
  pathout= DIRCWD+'/docs/'+ module_str
  if not os.path.exists(pathout) : os.makedirs(pathout)
  ca.getmodule_doc(module_str, file2=pathout+'/signature.py')

def a_info_conda_jupyter():
 s= '''
 CONDA COMMAND :
  'roll back' an installation, downgrading
     conda list  --revisions
     conda install --revision [revision number]         #Roll back to this number

  conda update PackageName

  conda uninstall PackageName


 IPYPARALLEL :
    Enable tab in Jupyter:   ipcluster nbextension enable
    Disable :                ipcluster nbextension enable
    Start 3 clsuter          ipcluster start -n 3
        <Stop them           ipclsuter stop

 JUPYTER Notebook :
      jupyter notebook      :  Enable Jupyter Server
      Close : Ctrl + C
      %connect_info      : To get Connection Details of Ipytho notebook


 '''
 print(s)

def a_run_cmd(cmd1):os_process_run(cmd1, shell=True)
# cmd("ipconfig")

def a_help():
 str= """
  PYCHARM shortcut :
    Highlight the Text  +  Alt +G   : Google Search
    Atl+C :  Doc   Alt+X : Doc     Alt+W : Doc Internet
    Ctrl+W : Refactor Name
    Shift + Shift  : Search Everywhere
    TouchRigh of Altgr + TouchRigh of Altgr: Definition of method

  #Save Session       session_save('/kaggle/supermarket_02', globals())

  #Load Session       session_load('/kaggle/kaggle_supermarket_01')


 1) %load_ext autoreload     #Reload the packages
    %autoreload 2

 2) Install PIP:
     If Permission access,Exit Spyder,  Use CMD  
       D:\WinPython-64-2710\scripts>pip install scikit-learn --upgrade -
 
     !pip install NAME --upgrade --no-deps  #NO dependencies installed.     
     
     !pip install git+git://github.com/fchollet/keras.git --upgrade --no-deps
     !pip install https://github.com//pymc3/archive/master.zip  
     !pip install  https://github.com/jpype-py3 /zipball/master  or  /tarball/master   

     !pip install c:\this_is_here\numpy-1.10.4+mkl-cp27-cp27m-win_amd64.whl
     !pip freeze   to get the list of package isntall

     !pip install -h    Help on pip  

 3) Windows CMD: 
    a_run_cmd("ipconfig")

    A: [enter] Changes the default drive from C to A.
   cd \furniture\chairs.  Moves you to the directory  'FURNITURE'
   cd .. Moves you up one level in the path.
   cd \  Takes you back to the root directory (c: in this case).


 4) import sys  
    reload(sys)  
    sys.setdefaultencoding('utf8')
    sys.getdefaultencoding()

 5) 
    !pip install packagename
    The ! prefix is a short-hand for the %sc command to run a shell command.

    You can also use the !! prefix which is a short-hand for the %sx command to execute a shell command 
    and capture its output (saved into the _ variable by default).

  6) Compile a modified package
    1) Put into a folder
    2) Add the folder to the Python Path (Spyder Python path)
 
 
    3) Compile in Spyder using, (full directory)
       !!python D:\\_devs\\Python01\\project27\\scikit_learn\\sklearn\\setup.py install

    4) Project is built here:
      D:\_devs\Python01\project27\build\lib.win-amd64-2.7
      http://scikit-learn.org/dev/developers/contributing.html#git-repo
python setup.py develop

Convert Python 2 to Python 3
import lib2to3

!2to3 D:\_devs\Python01\project\aapackage\codeanalysis.py

D:\_devs\Python01\project\zjavajar

# os.system('cd D:\_app\visualstudio13\Common7\IDE') #Execute any command
# os.path.abspath('memo.txt') #get absolute path
# os.path.exists('memo.txt')
# os.path.exists('memo.txt')
# os.getenv('PATH')    #Get environ variable

# sitepackage= r'D:\_devs\Python01\WinPython-64-2710\python-2.7.10.amd64\Lib\site-packages/'




  """
 os_gui_popup_show(str)
 #print( str)

def a_info_system():
 import platform; print((platform.platform() + '\n'))
 # conda install -c anaconda psutil=5.0.0
 # https://github.com/nicolargo/glances
 import sys; print(("Python", sys.version))
 a_info_packagelist()

def a_info_packagelist() :
 import pip
 installed_packages = pip.get_installed_distributions()
 installed_packages_list = sorted(["%s==%s" % (i.key, i.version) for i in installed_packages])
 for p in  installed_packages_list : print(p)
 # def aa_getlistofpackage() : np.__config__.show(); a_info_packagelist()

def a_get_pythonversion():    return sys.version_info[0]


#################### Printers ###################################################################
def print_object(vv, txt='')  :
 ''' #Print Object Table  '''
 print(("\n\n" + txt +"\n"))
 sh= np.shape(vv)
 kkmax, iimax= sh[0], sh[1]
 for k in range(0, kkmax):
   aux= ""
   for i in range(0,iimax) :
     if vv[k,0] != None :
      aux+=  str(vv[k,i])    + "," 
      
   if vv[k,0] != None :   
       print (aux )

def print_object_tofile(vv, txt, file1='d:/regression_output.py')  :
 ''' #Print to file Object   Table   '''
 with open(file1, mode='a')  as file1 :
  file1.write("\n\n" + txt +"\n" )  
  sh= np.shape(vv)
  kkmax, iimax= sh[0], sh[1]
  for k in range(0, kkmax):
    aux= ""
    for i in range(0,iimax) :
     if vv[k,0] != None :
      aux+=  str(vv[k,i])    + "," 
      
    if vv[k,0] != None :   
       #print (aux )
       file1.write(aux + 	"\n")

def print_ProgressBar (iteration, total, prefix = '', suffix = '', decimals = 1, barLength = 100):
    """# Print iterations progress
     Call in a loop to create terminal progress bar
    @params:
        iteration   - Required  : current iteration (Int)
        total       - Required  : total iterations (Int)
        prefix      - Optional  : prefix string (Str)
        suffix      - Optional  : suffix string (Str)
        decimals    - Optional  : positive number of decimals in percent complete (Int)
        barLength   - Optional  : character length of bar (Int)
    """
    formatStr = "{0:." + str(decimals) + "f}"
    percent = formatStr.format(100 * (iteration / float(total)))
    filledLength = int(round(barLength * iteration / float(total)))
    bar = '█' * filledLength + '-' * (barLength - filledLength)
    sys.stdout.write('\r%s |%s| %s%s %s' % (prefix, bar, percent, '%', suffix)),
    if iteration == total:
        sys.stdout.write('\n')
    sys.stdout.flush()
    '''
    for item in items:
     # Do stuff...
     # Update Progress Bar
     i += 1
     printProgress(i, l, prefix = 'Progress:', suffix = 'Complete', barLength = 50)
    '''


###################### OS- ######################################################################
def os_zip_checkintegrity(filezip1):
  import zipfile
  zip_file = zipfile.ZipFile(filezip1)
  try :
    ret= zip_file.testzip()
    if ret is not None:
        print("First bad file in zip: %s" % ret)
        return  False
    else: return True
  except Exception as e: return False

def os_zipfile(folderin, folderzipname, iscompress=True):
  import os, zipfile
  compress = zipfile.ZIP_DEFLATED if iscompress else  zipfile.ZIP_STORED
  zf= zipfile.ZipFile(folderzipname, "w", compress, allowZip64=True)

  for dirname, subdirs, files in os.walk(folderin):
        zf.write(dirname)
        for filename in files:  zf.write(os.path.join(dirname, filename))
  zf.close()

  r= os_zip_checkintegrity(folderzipname)
  if r : return  folderin, folderzipname
  else :
     print('Corrupt File')
     return  folderin, False

def os_zipfolder(dir_tozip='/zdisks3/output', zipname='/zdisk3/output.zip', dir_prefix=None, iscompress=True):
 '''
 shutil.make_archive('/zdisks3/results/output', 'zip',
                     root_dir=/zdisks3/results/',
                     base_dir='output')

 os_zipfolder('zdisk/test/aapackage', 'zdisk/test/aapackage.zip', 'zdisk/test')'''

 dir_tozip= dir_tozip if dir_tozip[-1] != '/' else dir_tozip[:-1]
 # dir_prefix= dir_prefix if dir_prefix[-1] != '/' else dir_prefix[:-1]

 if dir_prefix is None :
   dir_tozip, dir_prefix= '/'.join(dir_tozip.split('/')[:-1]), dir_tozip.split('/')[-1]

 import shutil
 shutil.make_archive(zipname.replace('.zip',''), 'zip', dir_tozip, base_dir=dir_prefix)


 '''
 shutil.make_archive(base_name, format[, root_dir[, base_dir[, verbose[, dry_run[, owner[, group[, logger]]]]]]])
Create an archive file (eg. zip or tar) and returns its name.

base_name is the name of the file to create, including the path, minus any format-specific extension. format is the archive format: one of “zip” (if the zlib module or external zip executable is available), “tar”, “gztar” (if the zlib module is available), or “bztar” (if the bz2 module is available).

root_dir is a directory that will be the root directory of the archive; ie. we typically chdir into root_dir before creating the archive.

base_dir is the directory where we start archiving from; ie. base_dir will be the common prefix of all files and directories in the archive.

root_dir and base_dir both default to the current directory.

owner and group are used when creating a tar archive. By default, uses the current owner and group.

logger must be an object compatible with PEP 282, usually an instance of logging.Logger.
  '''

 '''
 import zipfile
 compress = zipfile.ZIP_DEFLATED if iscompress else  zipfile.ZIP_STORED
 zf= zipfile.ZipFile(zipname, "w", compress, allowZip64=True)
 for dirname, subdirs, files in os.walk(dir_tozip):
    zf.write(dirname)
    for filename in files:  zf.write(os.path.join(dirname, filename))
 zf.close()
 '''
 r= os_zip_checkintegrity(zipname)
 if r : return  zipname
 else :
    print('Corrupt File'); return False

def os_zipextractall(filezip_or_dir="folder1/*.zip", tofolderextract='zdisk/test', isprint=1):
   '''os_zipextractall( 'aapackage.zip','zdisk/test/'      )  '''
   import zipfile

   if filezip_or_dir.find("*") > - 1 :    #  Many Zip
      ziplist1= os_file_listall(filezip_or_dir[:filezip_or_dir.find("*")],'*.zip')
      fileziplist_full=   ziplist1[2]

   else :                                 # Only 1
    filename=    os_file_getname(filezip_or_dir)
    fileziplist_full= [filezip_or_dir]


   # if os.path.exists( foldernew2  ) :      #Either File or Folder exists
   #   print('Renaming Folder ' + foldernew + ' with _')
   #   # lastfolder, beforelast= tofolderextract.split('/')[]
   #   os_folder_copy(foldernew2,  foldernew2+'_' )


   for filezip in fileziplist_full :
     filezip_name=    os_file_getname(filezip)
     zip_ref = zipfile.ZipFile(filezip, 'r')
     zip_ref.extractall(tofolderextract)   #Will create the path
     zip_ref.close()
     isok= os.path.exists( tofolderextract  )
     if not isok : print('Error: ' + filezip_name)

   if isok : return  tofolderextract
   else :    return -1

def os_folder_copy(src, dst, symlinks=False, pattern1="*.py", fun_file_toignore=None):
   '''
       callable(src, names) -> ignored_names
       'src' parameter, which is the directory being visited by copytree(), and
       'names' which is the list of `src` contents, as returned by os.listdir():

    Since copytree() is called recursively, the callable will be called once for each directory that is copied.
    It returns a  list of names relative to the `src` directory that should not be copied.
   :param fun_ignore:
   '''
   import shutil, errno, fnmatch
   def fun_file_toignore(src, names) :
       pattern= '!' + pattern1
       file_toignore= fnmatch.filter(names, pattern)
       return file_toignore

   try:
        shutil.copytree(src, dst, symlinks=False, ignore=fun_file_toignore)
   except OSError as exc: # python >2.5
        if exc.errno == errno.ENOTDIR:
            shutil.copy(src, dst,symlinks=False, ignore=fun_file_toignore)
        else: raise

def os_folder_create(directory) :
   DIR0= os.getcwd()
   if not os.path.exists(directory): os.makedirs(directory)
   os.chdir(DIR0)

def os_folder_robocopy(from_folder='', to_folder='', my_log='H:/robocopy_log.txt'):
    """
    Copy files to working directory
    robocopy <Source> <Destination> [<File>[ ...]] [<Options>]
    We want to copy the files to a fast SSD drive
    """

    if os.path.isdir(from_folder) &  os.path.isdir(to_folder):
        subprocess.call(["robocopy", from_folder, to_folder, "/LOG:%s" % my_log])
    else:
        print("Paths not entered correctly")




def os_file_replace(source_file_path, pattern, substring):
    from tempfile import mkstemp
    from shutil import move
    from os import remove

    fh, target_file_path = mkstemp()
    with open(target_file_path, 'w') as target_file:
        with open(source_file_path, 'r') as source_file:
            for line in source_file:
                target_file.write(line.replace(pattern, substring))
    remove(source_file_path)
    move(target_file_path, source_file_path)


def os_file_replacestring1(findStr, repStr, filePath):
    "replaces all findStr by repStr in file filePath"
    import fileinput  
    file1= fileinput.FileInput(filePath, inplace=True, backup='.bak')
    for line in file1:
       line= line.replace(findStr,  repStr)
       sys.stdout.write(line)
    file1.close()
    print(("OK: "+format(filePath)))


def os_file_replacestring2(findstr, replacestr, some_dir, pattern="*.*", dirlevel=1  ):
  ''' #fil_replacestring_files("logo.png", "logonew.png", r"D:/__Alpaca__details/aiportfolio",    pattern="*.html", dirlevel=5  )
  '''
  list_file= os_file_listall(some_dir, pattern=pattern, dirlevel=dirlevel)
  list_file= list_file[2]
  for file1 in list_file : os_file_replacestring1(findstr, replacestr, file1)


def os_file_getname(path):
    import ntpath
    head, tail = ntpath.split(path)
    return tail or ntpath.basename(head)

def os_file_getpath(path):
    import ntpath
    head, tail = ntpath.split(path)
    return head

def os_file_gettext(file1):
 with open(file1, 'r',encoding='UTF-8',) as f:      
  return f.read()
 #def os_file_listall(some_dir, pattern="*.*", dirlevel=1):
 #  return listallfile(some_dir, pattern=pattern, dirlevel=dirlevel)

def os_file_listall(dir1, pattern="*.*", dirlevel=1, onlyfolder=0):
  '''
   # DIRCWD=r"D:\_devs\Python01\project"
   # aa= listallfile(DIRCWD, "*.*", 2)
   # aa[0][30];   aa[2][30]

   :param dir1:
   :param pattern:
   :param dirlevel:
   :param onlyfolder:
   :return:
  '''
  import fnmatch; import os; import numpy as np;  matches = []
  dir1 = dir1.rstrip(os.path.sep)
  num_sep = dir1.count(os.path.sep)

  if onlyfolder :
   for root, dirs, files in os.walk(dir1):
    num_sep_this = root.count(os.path.sep)
    if num_sep + dirlevel <= num_sep_this: del dirs[:]
    matches.append([]); matches.append([]); matches.append([]);   # Filename, DirName
    for dirs in fnmatch.filter(dirs, pattern):
      matches[0].append(os.path.splitext(dirs)[0])
      matches[1].append(os.path.splitext(root)[0])
      matches[2].append(os.path.join(root, dirs))
   return np.array(matches)

  for root, dirs, files in os.walk(dir1):
    num_sep_this = root.count(os.path.sep)
    if num_sep + dirlevel <= num_sep_this: del dirs[:]
    matches.append([]); matches.append([]); matches.append([]);   # Filename, DirName
    for files in fnmatch.filter(files, pattern):
      matches[0].append(os.path.splitext(files)[0])   
      matches[1].append(os.path.splitext(files)[1])  
      matches[2].append(os.path.join(root, files))   
  return np.array(matches)

def os_file_rename(some_dir, pattern="*.*", pattern2="", dirlevel=1):
  import fnmatch; import os; import numpy as np; import re;  matches = []
  some_dir = some_dir.rstrip(os.path.sep)
  assert os.path.exists(some_dir)
  num_sep = some_dir.count(os.path.sep)
  for root, dirs, files in os.walk(some_dir):
    num_sep_this = root.count(os.path.sep)
    if num_sep + dirlevel <= num_sep_this: del dirs[:]
    matches.append([]); matches.append([]); matches.append([]);   # Filename, DirName
    for files in fnmatch.filter(files, pattern):
        
     #replace pattern by pattern2        
      nfile= re.sub(pattern, pattern2, files) 
      os.path.abspath(root)
      os.rename(files, nfile)
 
      matches[0].append(os.path.splitext(nfile)[0])   
      matches[1].append(os.path.splitext(nfile)[1])  
      matches[2].append(os.path.join(root, nfile))   
  return np.array(matches).T

def os_gui_popup_show(txt):
  def os_gui_popup_show2(txt):
    from tkinter import Tk, Scrollbar,  Text, mainloop, RIGHT, END, LEFT, Y
    root = Tk()
    S = Scrollbar(root)
    T = Text(root, height=50, width=90)
    S.pack(side=RIGHT, fill=Y)
    T.pack(side=LEFT, fill=Y)
    S.config(command=T.yview)
    T.config(yscrollcommand=S.set)
    T.insert(END, txt)
    root.attributes('-topmost', True) # note - before topmost
    mainloop(  )
  os_gui_popup_show2(txt)
  # import _thread
  # _thread.start_new_thread(os_gui_popup_show2, (txt,))  #issues

def os_print_tofile(vv, file1,  mode1='a'):  # print into a file='a
    '''
    Here is a list of the different modes of opening a file:
r
Opens a file for reading only. The file pointer is placed at the beginning of the file. This is the default mode.

rb

Opens a file for reading only in binary format. The file pointer is placed at the beginning of the file. This is the default mode.
r+

Opens a file for both reading and writing. The file pointer will be at the beginning of the file.
rb+

Opens a file for both reading and writing in binary format. The file pointer will be at the beginning of the file.
w

Opens a file for writing only. Overwrites the file if the file exists. If the file does not exist, creates a new file for writing.
wb

Opens a file for writing only in binary format. Overwrites the file if the file exists. If the file does not exist, creates a new file for writing.
w+

Opens a file for both writing and reading. Overwrites the existing file if the file exists. If the file does not exist, creates a new file for reading and writing.
wb+

Opens a file for both writing and reading in binary format. Overwrites the existing file if the file exists. If the file does not exist, creates a new file for reading and writing.
a

Opens a file for appending. The file pointer is at the end of the file if the file exists. That is, the file is in the append mode. If the file does not exist, it creates a new file for writing.
ab

Opens a file for appending in binary format. The file pointer is at the end of the file if the file exists. That is, the file is in the append mode. If the file does not exist, it creates a new file for writing.
a+

Opens a file for both appending and reading. The file pointer is at the end of the file if the file exists. The file opens in the append mode. If the file does not exist, it creates a new file for reading and writing.
ab+

Opens a file for both appending and reading in binary format. The file pointer is at the end of the file if the file exists. The file opens in the append mode. If the file does not exist, it creates a new file for reading and writing.
To open a text file, use:
fh = open("hello.txt", "r")

To read a text file, use:
print fh.read()

To read one line at a time, use:
print fh.readline()

To read a list of lines use:
print fh.readlines()

To write to a file, use:
fh = open("hello.txt", "w")
lines_of_text = ["a line of text", "another line of text", "a third line"]
fh.writelines(lines_of_text)
fh.close()

To append to file, use:
fh = open("Hello.txt", "a")
fh.close()

    '''
    with open(file1, mode1) as text_file:  text_file.write(str(vv))

def os_path_norm(pth): #Normalize path for Python directory
    ''' #r"D:\_devs\Python01\project\03-Connect_Java_CPP_Excel\PyBindGen\examples" '''
    if a_get_pythonversion()==2:
     ind = pth.find(":")
     if ind > -1 :
       a, b = pth[:ind], pth[ind + 1:].encode("string-escape").replace("\\x", "/")
       return "{}://{}".format(a, b.lstrip("\\//").replace("\\\\", "/"))        
     else: return pth
    else: 
      pth = pth.encode("unicode-escape").replace(b"\\x", b"/")
      return pth.replace(b"\\\\", b"/").decode("utf-8")

def os_path_change(path1): path1= normpath(path1); os.chdir(path1)    #Change Working directory path

def os_path_current(): return DIRCWD

def os_file_exist(file1): return os.path.exists(file1)

def os_file_size(file1): return os.path.getsize(file1)

def os_file_read(file1):
 fh = open(file1,"r")
 return fh.read()


def os_file_isame(file1, file2) :
   import filecmp
   return filecmp.cmp(file1 , file2 )




def os_file_get_file_extension(file_path):
    """
    >>> get_file_extension("/a/b/c")
    ''
    >>> get_file_extension("/a/b/c.tar.xz")
    'xz'
    """
    _ext = os.path.splitext(file_path)[-1]
    if _ext:
        return _ext[1:] if _ext.startswith('.') else _ext

    return ""





def os_file_normpath(path):
    """Normalize path.
    - eliminating double slashes, etc. (os.path.normpath)
    - ensure paths contain ~[user]/ expanded.

    :param path: Path string :: str
    """
    return os.path.normpath(os.path.expanduser(path) if '~' in path else path)


def os_folder_is_path(path_or_stream):
    """
    Is given object `path_or_stream` a file path?
    :param path_or_stream: file path or stream, file/file-like object
    :return: True if `path_or_stream` is a file path
    """
    return isinstance(path_or_stream, str)


def os_file_get_path_from_stream(maybe_stream):
    """
    Try to get file path from given stream `stream`.

    :param maybe_stream: A file or file-like object
    :return: Path of given file or file-like object or None

    >>> __file__ == get_path_from_stream(__file__)
    True
    >>> __file__ == get_path_from_stream(open(__file__, 'r'))
    True
    >>> strm = anyconfig.compat.StringIO()
    >>> get_path_from_stream(strm) is None
    True
    """
    if os_folder_is_path(maybe_stream):
        return maybe_stream  # It's path.

    maybe_path = getattr(maybe_stream, "name", None)
    if maybe_path is not None:
        maybe_path = os.path.abspath(maybe_path)

    return maybe_path


def os_file_try_to_get_extension(path_or_strm):
    """
    Try to get file extension from given path or file object.
    :return: File extension or None
    """
    path = os_file_get_path_from_stream(path_or_strm)
    if path is None:
        return None

    return os_file_get_file_extension(path) or None


def os_file_are_same_file_types(paths):
    """
    Are given (maybe) file paths same type (extension) ?
    :param paths: A list of file path or file(-like) objects

    >>> are_same_file_types([])
    False
    >>> are_same_file_types(["a.conf"])
    True
    >>> are_same_file_types(["a.yml", "b.json"])
    False
    >>> strm = anyconfig.compat.StringIO()
    >>> are_same_file_types(["a.yml", "b.yml", strm])
    False
    """
    if not paths:   return False
    ext = os_file_try_to_get_extension(paths[0])
    if ext is None: return False

    return all(os_file_try_to_get_extension(p) == ext for p in paths[1:])


def os_file_norm_paths(paths, marker='*'):
    """
    :param paths:
        A glob path pattern string, or a list consists of path strings or glob
        path pattern strings or file objects
    :param marker: Glob marker character or string, e.g. '*'
    :return: List of path strings
    >>> norm_paths([])
    []
    >>> norm_paths("/usr/lib/a/b.conf /etc/a/b.conf /run/a/b.conf".split())
    ['/usr/lib/a/b.conf', '/etc/a/b.conf', '/run/a/b.conf']
    >>> paths_s = os.path.join(os.path.dirname(__file__), "u*.py")
    >>> ref = sglob(paths_s)
    >>> ref = ["/etc/a.conf"] + ref
    >>> assert norm_paths(["/etc/a.conf", paths_s]) == ref
    >>> strm = anyconfig.compat.StringIO()
    >>> assert norm_paths(["/etc/a.conf", strm]) == ["/etc/a.conf", strm]
    """
    def sglob(files_pattern):
      """
      glob.glob alternative of which results sorted always.
      """
      return sorted(glob.glob(files_pattern))

    def _norm_paths_itr(paths, marker='*'):
     """Iterator version of :func:`norm_paths`.
     """
     for path in paths:
        if os_folder_is_path(path):
            if marker in path:  # glob path pattern
                for ppath in sglob(path):
                    yield ppath
            else:
                yield path  # a simple file path
        else:  # A file or file-like object
            yield path

    if os_folder_is_path(paths) and marker in paths:
        return sglob(paths)

    return list(_norm_paths_itr(paths, marker=marker))




def os_file_mergeall(nfile, dir1, pattern1, deepness=2):
 ll= os_file_listall(dir1,pattern1,deepness)
 with open(nfile, mode='a', encoding='UTF-8') as nfile1:
  txt=''; ii=0
  for l in ll[2]:
   txt= '\n\n\n\n' + os_file_gettext(l)
   nfile1.write(txt) 
 nfile1.close()


def os_file_extracttext(output_file, dir1, pattern1="*.html", htmltag='p', deepness=2):
 ''' Extract text from html '''
 ll= os_file_listall(dir1, pattern1,5)

 with open(output_file, mode='a', encoding='UTF-8') as output_file1:
    txt=''; ii=0
    for l in ll[2]:
      page= os_file_gettext(l)
      soup = BeautifulSoup(page, "lxml")
      txt2 = ' \n\n'.join([p.text for p in soup.find_all(htmltag)])
   
    txt= '\n\n\n\n' + txt2.strip()
    output_file1.write(txt)


def os_path_append(p1, p2=None, p3=None, p4=None):
 sys.path.append(p1)
 if p2 is not None:  sys.path.append(p2)
 if p3 is not None:  sys.path.append(p3)
 if p4 is not None:  sys.path.append(p4)



########### WAIT BEFORE LAUNCH #########################################################################
def os_wait_cpu(priority=300, cpu_min=50) :
 from time import sleep; import psutil, arrow
 aux= psutil.cpu_percent()
 while aux > cpu_min :
    print("CPU:", aux, arrow.utcnow().to('Japan').format())
    sleep( priority );  aux=  psutil.cpu_percent(); sleep(10); aux= 0.5*(aux + psutil.cpu_percent() )
 print("Starting script:", aux, arrow.utcnow().to('Japan').format())


def os_split_dir_file(dirfile) :
    lkey= dirfile.split('/')
    if len(lkey)== 1 :dir1=""
    else :            dir1= '/'.join(lkey[:-1]); dirfile= lkey[-1]
    return dir1, dirfile

def os_process_run(cmd_list=['program', 'arg1', 'arg2'], capture_output=False):
    import subprocess
    #cmd_list= os_path_norm(cmd_list)
    PIPE=subprocess.PIPE
    STDOUT=subprocess.STDOUT
    proc = subprocess.Popen(cmd_list, stdout=PIPE, stderr=STDOUT, shell=False)  #Always put to False

    if capture_output :
     stdout, stderr = proc.communicate()
     err_code = proc.returncode
     print("Console Msg: \n")
     print((str(stdout))) #,"utf-8"))
     print(("\nConsole Error: \n"+ str(stderr) ))
     #    return stdout, stderr, int(err_code)


def os_process_2():
   pass
   '''
Ex: Dialog (2-way) with a Popen()
p = subprocess.Popen('Your Command Here',
                 stdout=subprocess.PIPE,
                 stderr=subprocess.STDOUT,
                 stdin=PIPE,
                 shell=True,
                 bufsize=0)
p.stdin.write('START\n')
out = p.stdout.readline()
while out:
  line = out
  line = line.rstrip("\n")

  if "WHATEVER1" in line:
      pr = 1
      p.stdin.write('DO 1\n')
      out = p.stdout.readline()
      continue

  if "WHATEVER2" in line:
      pr = 2
      p.stdin.write('DO 2\n')
      out = p.stdout.readline()
      continue

out = p.stdout.readline()
p.wait()
'''





#######################  Python Interpreter ###################################################
def py_importfromfile(modulename, dir1):
 #Import module from file:  (Although this has been deprecated in Python 3.4.)
 vv= a_get_pythonversion()
 if vv==3:
  from importlib.machinery import SourceFileLoader
  foo = SourceFileLoader("module.name", "/path/to/file.py").load_module()
  foo.MyClass()
 elif vv==2 :
  import imp
  foo = imp.load_source('module.name', '/path/to/file.py')
  foo.MyClass()

def py_memorysize(o, ids, hint=" deep_getsizeof(df_pd, set()) "):
    """ deep_getsizeof(df_pd, set())
    Find the memory footprint of a Python object
    The sys.getsizeof function does a shallow size of only. It counts each
    object inside a container as pointer only regardless of how big it
    """
    from collections import Mapping, Container
    from sys import getsizeof

    d = py_memorysize
    if id(o) in ids:
        return 0

    r = getsizeof(o)
    ids.add(id(o))

    if isinstance(o, str) or isinstance(0, str):
        r= r

    if isinstance(o, Mapping):
        r= r + sum(d(k, ids) + d(v, ids) for k, v in o.items())

    if isinstance(o, Container):
        r=  r + sum(d(x, ids) for x in o)

    return r * 0.0000001


def save(obj, folder='/folder1/keyname', isabsolutpath=0 ) :
 return py_save_obj(obj, folder=folder, isabsolutpath=isabsolutpath)

def load(folder='/folder1/keyname', isabsolutpath=0 ) :
 return py_load_obj(folder=folder, isabsolutpath=isabsolutpath)

def save_test(folder='/folder1/keyname', isabsolutpath=0  ) :
 ztest=  py_load_obj(folder=folder, isabsolutpath=isabsolutpath)
 print(('Load object Type:' + str(type(ztest)) ))
 del ztest; gc.collect()


def py_save_obj(obj, folder='/folder1/keyname', isabsolutpath=0):
    import pickle
    if isabsolutpath==0  and folder.find('.pkl') == -1 : #Local Path
       dir0, keyname= z_key_splitinto_dir_name(folder)
       os_folder_create(DIRCWD+'/aaserialize/' + dir0)
       dir1= DIRCWD+'/aaserialize/' + dir0 + '/'+ keyname + '.pkl'
    else :
       dir1= folder

    with open( dir1, 'wb') as f:
        pickle.dump(obj, f, pickle.HIGHEST_PROTOCOL)
    return dir1
    
def py_load_obj(folder='/folder1/keyname', isabsolutpath=0, encoding1='utf-8'):
    '''def load_obj(name, encoding1='utf-8' ):
         with open('D:/_devs/Python01/aaserialize/' + name + '.pkl', 'rb') as f:
            return pickle.load(f, encoding=encoding1)
    '''
    import pickle
    if isabsolutpath==0 and folder.find('.pkl') == -1 :
       dir0, keyname= z_key_splitinto_dir_name(folder)
       os_folder_create(DIRCWD+'/aaserialize/' + dir0)
       dir1= DIRCWD+'/aaserialize/' + dir0 + '/'+ keyname + '.pkl'
    else :
       dir1= folder

    with open(dir1, 'rb') as f:
        return pickle.load(f)

def z_key_splitinto_dir_name(keyname) :
    lkey= keyname.split('/')
    if len(lkey)== 1 :dir1=""
    else :            dir1= '/'.join(lkey[:-1]); keyname= lkey[-1]
    return dir1, keyname




############# Object ########################################################################
def sql_getdate():
  pass

'''
solver1= load_obj('batch/solver_ELVIS_USD_9assets_3_regime_perf_20160906_059')

solver1.x
solver1.convergence
._calculate_population_energies
solver1.next()  #One generation evolution
solver1.solve()  #Solve the pronlem

solver1.population           # Get the list of population sol
solver1.population_energies  # Population energy

aux= solver1.population


solver1.next()  #One generation evolution


def np_runsolver(name1, niter):

pfi= copy.deepcopy( get_class_that_defined_method(solver1.func) )

obj_getclass_property(pfi) 

solver1.x



id : int
date:
time:
name: str
solver
xbest
fbest
population
population_energies
params1
params2
params3
params4
params5
details

storage_optim= np.empty((1000, 15), dtype= np.object)

 
  id= ['id','date', 'time','name','solver','xbest', 'fbest', 'population','pop_energies',
  'params1', 'params2', 'params3', 'params4', 'params5', 'details']      )                                 
                                  
def tostore('store/'+dbname, dbid, vv) :                                  
 storage_optim=  load_obj(dbname)
 
 storage_optim[0,0]= 0
 storage_optim[0,1]= date_now()
 storage_optim[0,2]= copy.deepcopy(solver1)

 copyfile(dbname)
 save_obj(storage_optim, 'store/'+dbname)


aux= (solver1.population ,  solver1.population_energies  )
save_obj(aux, 'batch/elvis_usd_9assets_3_regime_perf_best_population_01')


'''


############### Object Class Introspection ##################################################
def obj_getclass_of_method(meth):
    import inspect
    for cls in inspect.getmro(meth.__self__.__class__):
        if meth.__name__ in cls.__dict__: 
            return cls
    return None


def obj_getclass_property(pfi) :
 for property, value in vars(pfi).items():
    print(property, ": ", value)



################# XML / HTML processing  ####################################################
'''
https://pypi.python.org/pypi/RapidXml/
http://pugixml.org/benchmark.html

'''



#####################################################################################
#-------- PDF processing ------------------------------------------------------------
def print_topdf() :
 import datetime
 import numpy as np
 from matplotlib.backends.backend_pdf import PdfPages
 import matplotlib.pyplot as plt

 # Create the PdfPages object to which we will save the pages:
 # The with statement makes sure that the PdfPages object is closed properly at
 # the end of the block, even if an Exception occurs.
 with PdfPages('multipage_pdf.pdf') as pdf:
    plt.figure(figsize=(3, 3))
    plt.plot(list(range(7)), [3, 1, 4, 1, 5, 9, 2], 'r-o')
    plt.title('Page One')
    pdf.savefig()  # saves the current figure into a pdf page
    plt.close()

    plt.rc('text', usetex=True)
    plt.figure(figsize=(8, 6))
    x = np.arange(0, 5, 0.1)
    plt.plot(x, np.sin(x), 'b-')
    plt.title('Page Two')
    pdf.savefig()
    plt.close()

    plt.rc('text', usetex=False)
    fig = plt.figure(figsize=(4, 5))
    plt.plot(x, x*x, 'ko')
    plt.title('Page Three')
    pdf.savefig(fig)  # or you can pass a Figure object to pdf.savefig
    plt.close()

    # We can also set the file's metadata via the PdfPages object:
    d = pdf.infodict()
    d['Title'] = 'Multipage PDF Example'
    d['Author'] = 'Jouni K. Sepp\xe4nen'
    d['Subject'] = 'How to create a multipage pdf file and set its metadata'
    d['Keywords'] = 'PdfPages multipage keywords author title subject'
    d['CreationDate'] = datetime.datetime(2009, 11, 13)
    d['ModDate'] = datetime.datetime.today()




#####################################################################################
#--------CSV processing -------------------------------------------------------------
#Put Excel and CSV into Database / Extract CSV from database

def os_config_setfile(dict_params, outfile, mode1='w+') :
  with open(outfile, mode=mode1) as f1 :
     for key, item in dict_params.items() :
       if isinstance(item, str)  or isinstance(item, str) :
          f1.write( str(key) +'= ' + "'"+ str(item) + "'" + ' \n')
       else :
          f1.write( str(key) +'= ' + str(item) + ' \n')

  f1.close()
  print(outfile)


def os_config_getfile(file1) :
   with open(file1, mode='r')  as f1 :
      ll= f1.readlines()

   for x in ll :
      print(x, end=' ')


def os_csv_process(file1):
 print(  '''
import csv
with open(file2, 'r',encoding='UTF-8',) as f:      
  reader = csv.reader(f,  delimiter='/' )
  kanjidict=  dict(reader) 
  
 return kanjidict  

import csv
with csv.reader(open(file2, 'r',encoding='UTF-8'),  delimiter=',' ) as reader :
  for row in reader:

# pip install fastcsv

with fastcsv.Reader(io.open(CSV_FILE)) as reader:
    for row in reader:
        pass


with fastcsv.Writer(io.open(CSV_FILE, 'w', encoding='cp932')) as writer:
    writer.writerow(row)

import csv
FUNDING = 'data/funding.csv'

def read_funding_data(path):
    with open(path, 'rU') as data:
        reader = csv.DictReader(data)
        for row in reader:
            yield row

if __name__ == "__main__":
    for idx, row in enumerate(read_funding_data(FUNDING)):
        if idx > 10: break
        print "%(company)s (%(numEmps)s employees) raised %(raisedAmt)s on %(fundedDate)s" % row

A couple of key points with the code above:
Always wrap the CSV reader in a function that returns a generator (via the yield statement).
Open the file in universal newline mode with 'rU' for backwards compatibility.
Use context managers with [callable] as [name] to ensure that the handle to the file is closed automatically.
Use the csv.DictReader class only when headers are present, otherwise just use csv.reader. (You can pass a list of fieldnames, but you'll see its better just to use a namedtuple as we discuss below).
This code allows you to treat the data source as just another iterator or list in your code. In fact if you do the following:

data = read_funding_data(FUNDING)
print repr(data)

Encoding reading :
These encoders translate the native file encoding to UTF-8, which the csv module can read because it's 8-bit safe.
try:
    import unicodecsv as csv
except ImportError:
    import warnings
    warnings.warn("can't import `unicodecsv` encoding errors may occur")
    import csv

pip install unicodecsv

from collections import namedtuple

fields = ("permalink","company","numEmps", "category","city","state","fundedDate", "raisedAmt","raisedCurrency","round")
FundingRecord = namedtuple('FundingRecord', fields)

def read_funding_data(path):
    with open(path, 'rU') as data:
        data.readline()            # Skip the header
        reader = csv.reader(data)  # Create a regular tuple reader
        for row in map(FundingRecord._make, reader):
            yield row

if __name__ == "__main__":
    for row in read_funding_data(FUNDING):
        print row
        break

https://districtdatalabs.silvrback.com/simple-csv-data-wrangling-with-python

from datetime import datetime

fields = ("permalink","company","numEmps", "category","city","state","fundedDate", "raisedAmt","raisedCurrency","round")

class FundingRecord(namedtuple('FundingRecord_', fields)):

    @classmethod
    def parse(klass, row):
        row = list(row)                                # Make row mutable
        row[2] = int(row[2]) if row[2] else None       # Convert "numEmps" to an integer
        row[6] = datetime.strptime(row[6], "%d-%b-%y") # Parse the "fundedDate"
        row[7] = int(row[7])                           # Convert "raisedAmt" to an integer
        return klass(*row)

    def __str__(self):
        date = self.fundedDate.strftime("%d %b, %Y")
        return "%s raised %i in round %s on %s" % (self.company, self.raisedAmt, self.round, date)

def read_funding_data(path):
    with open(path, 'rU') as data:
        data.readline()            # Skip the header
        reader = csv.reader(data)  # Create a regular tuple reader
        for row in map(FundingRecord.parse, reader):
            yield row

if __name__ == "__main__":
    for row in read_funding_data(FUNDING):
        print row
        break
''')


def pd_toexcel(df, outfile='file.xlsx', sheet_name='sheet1', append=1, returnfile=1):
   '''
# Create a Pandas Excel writer using XlsxWriter as the engine.
writer = pd.ExcelWriter('test.xlsx', engine='xlsxwriter')
df.to_excel(writer, sheet_name='Sheet1')
writer.save()

# Get the xlsxwriter objects from the dataframe writer object.
workbook  = writer.book
worksheet = writer.sheets['Sheet1']

# Add some cell formats.
format1 = workbook.add_format({'num_format': '#,##0.00'})
format2 = workbook.add_format({'num_format': '0%'})
format3 = workbook.add_format({'num_format': 'h:mm:ss AM/PM'})

# Set the column width and format.
worksheet.set_column('B:B', 18, format1)

# Set the format but not the column width.
worksheet.set_column('C:C', None, format2)

worksheet.set_column('D:D', 16, format3)

# Close the Pandas Excel writer and output the Excel file.
writer.save()

from openpyxl import load_workbook
wb = load_workbook(outfile)
ws = wb.active
ws.title = 'Table 1'

tableshape = np.shape(table)
alph = list(string.ascii_uppercase)

for i in range(tableshape[0]):
    for j in range(tableshape[1]):
        ws[alph[i]+str(j+1)] = table[i, j]

for cell in ws['A'] + ws[1]:
    cell.style = 'Pandas'

wb.save('Scores.xlsx')

   '''

   if append and os_file_exist(outfile) :
     from openpyxl import load_workbook
     book = load_workbook(outfile)
     writer = pd.ExcelWriter(outfile, engine='openpyxl')
     writer.book = book
     writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
     df.to_excel(writer, sheet_name=sheet_name)
     writer.save()
     if returnfile : return outfile
   else :
     writer = pd.ExcelWriter(outfile, engine='xlsxwriter')
     df.to_excel(writer, sheet_name=sheet_name, )
     writer.save()
     if returnfile : return outfile

def pd_toexcel_many(outfile='file1.xlsx', df1=None, df2=None, df3=None, df4=None, df5=None, df6=None) :
  pd_toexcel(df1, outfile,  sheet_name='df1')
  if df2 is not None :   pd_toexcel(df2, outfile,  sheet_name='df2')
  if df3 is not None :   pd_toexcel(df3, outfile,  sheet_name='df3')
  if df4 is not None :   pd_toexcel(df4, outfile,  sheet_name='df4')
  if df5 is not None :   pd_toexcel(df5, outfile,  sheet_name='df5')
  if df6 is not None :   pd_toexcel(df6, outfile,  sheet_name='df6')



############# STRING- ########################################################################
def find_fuzzy(xstring, list_string):
 ''' if xstring matches partially, add to the list   '''
 return [ xi for xi in list_string if xi.find(xstring) > -1 ]


def str_match_fuzzy(xstring, list_string):
 ''' if any of list_strinf elt matches partially xstring '''
 for  xi in list_string:
   if xstring.find(xi) > -1:  return True
 return False


def str_parse_stringcalendar(cal):
    '''----------Parse Calendar  --------'''
    cal2 = cal.split('\n')
    cal3 = [x for x in cal2 if x != '']
    import dateutil
    cal4 = []
    for x in cal3:
        i0 = x.find(':')
        cal4.append([datetime_toint(dateutil.parser.parse(x[:i0])), x[i0 + 1:]])
    cal4 = np.array(cal4)
    cal5 = np.array(sortcol(cal4, 0), dtype=str)
    for x in cal5: print(x[0], ":", x[1].strip())
    return cal5

def str_make_unicode(input, errors='replace'):
    ttype= type(input) 
    if ttype != str  : 
        input =  input.decode('utf-8', errors=errors);  return input
    else:  return input

def str_empty_string_array(x,y=1):
 if y==1:  return ["" for x in range(x)]
 else: return [["" for row in range(0,x)] for col in range(0,y)]

def str_empty_string_array_numpy(nx, ny=1):
 arr = np.empty((nx, ny), dtype=object)
 arr[:, :] = ''
 return arr

def str_isfloat(value):
  try:    float(value);    return True
  except :    return False

def str_is_azchar(x):
  try:    float(x);    return True
  except :    return False  
    
def str_is_az09char(x):
  try:    float(x);    return True
  except :    return False  

def str_reindent(s, numSpaces): #change indentation of multine string
    '''
   if args:
       aux= name1+'.'+obj.__name__ +'('+ str(args) +')  \n' + str(inspect.getdoc(obj))
       aux= aux.replace('\n', '\n       ')
       aux= aux.rstrip()
       aux= aux + ' \n'
       wi( aux)
    '''
    s = string.split(s, '\n')
    s = [(numSpaces * ' ') + string.lstrip(line) for line in s]
    s = string.join(s, '\n')
    return s

def str_split2(delimiters, string, maxsplit=0):  #Split into Sub-Sentence
    import re
    regexPattern = '|'.join(map(re.escape, delimiters))
    return re.split(regexPattern, string, maxsplit)

def str_split_pattern(sep2, ll, maxsplit=0):  #Find Sentence Pattern
 import re
 regexPat = '|'.join(sep2) 
 regex3 = re.compile( '(' + regexPat + r')|(?:(?!'+ regexPat +').)*', re.S)
 #re.compile(r'(word1|word2|word3)|(?:(?!word1|word2|word3).)*', re.S)
 ll= regex3.sub(lambda m: m.group(1) if m.group(1) else "P", ll)
 return ll

def pd_str_isascii(x):
  try :
    x.decode('ascii'); return True
  except: return false


def str_to_utf8(x):
  """ Do it before saving/output to external printer """
  return x.encode('utf-8')


def str_to_unicode(x, encoding='utf-8'):
  ''' Do it First after Loading some text '''
  if isinstance(x, str):
    if not isinstance(x, str): return str(x, encoding)
  else : return x


'''
http://www.ibm.com/developerworks/library/l-pyint/

spam.upper().lower()

spam.strip()
spam.lstrip()
spam.rstrip()  Performs both lstrip() and rstrip() on string

count(str, beg= 0,end=len(string))  Counts how many times str occurs in string or in a substring of
find(str, beg=0 end=len(string))  str occurs in string or in a substring of strindex if found  -1 otherwise.
replace(old, new [, max])   Replaces all occurrences of old in string with new or at most max occurrences if max given.

isdecimal()   Returns true if a unicode string contains only decimal characters and false otherwise.
isalnum()  Returns true if string has at least 1 character and all characters are alphanumeric and false otherwise.
salpha()   Returns true if string has at least 1 character and all characters are alphabetic and false otherwise.
isdigit()  Returns true if string contains only digits and false otherwise.
islower()  Returns true if string has at least 1 cased character and all cased characters are in lowercase and false otherwise.
isnumeric() Returns true if a unicode string contains only numeric characters and false otherwise.
isspace()  Returns true if string contains only whitespace characters and false otherwise.
istitle()  Returns true if string is properly "titlecased" and false otherwise.

lower()   Converts all uppercase letters in string to lowercase.
capitalize()   Capitalizes first letter of string
center(width, fillchar)  Returns space-padded string with string centered  w
title()  Returns "titlecased" version of string, that is, all words begin with uppercase and the rest are lowercase.
upper()   Converts lowercase letters in string to uppercase.

join(seq)  Merges (concatenates) the string representations of elements in sequence seq into a string, with separator string.
	
split(str="", num=string.count(str))   Splits string according to delimiter str  and returns list of substrings;
splitlines( num=string.count('\n'))  Splits string at all (or num) NEWLINEs and returns a list of each line with NEWLINEs removed.

ljust(width[, fillchar])   Returns a space-padded string with the original string left-justified to a total of width columns.

maketrans()  Returns a translation table to be used in translate function.

index(str, beg=0, end=len(string))   Same as find(), but raises an exception if str not found.
rfind(str, beg=0,end=len(string))   Same as find(), but search backwards in string.

rindex( str, beg=0, end=len(string))   Same as index(), but search backwards in string.

rjust(width,[, fillchar])  Returns a space-padded string with string right-jus to a total of width columns.

startswith(str, beg=0,end=len(string))
Determines if string or a substring of string (if starting index beg and ending index end are given) starts with substring str; returns true if so and false otherwise.

decode(encoding='UTF-8',errors='strict')
Decodes the string using the codec registered for encoding. encoding defaults to the default string encoding.
	
encode(encoding='UTF-8',errors='strict')
Returns encoded string version of string; on error, default is to raise a ValueError

endswith(suffix, beg=0, end=len(string))
Determines if string or a substring of string  ends with suffix; returns true if so and false otherwise.

expandtabs(tabsize=8)
Expands tabs in string to multiple spaces; defaults to 8 spaces per tab if tabsize not provided.
	
zfill (width)
Returns original string leftpadded with zeros to a total of width characters; intended for numbers, zfill() retains any sign given (less one zero).

'''


################### LIST UTIL / Array  ############################################################
def np_minimize(fun_obj, x0=[0.0], argext=(0,0), bounds1=[(0.03, 0.20), (10, 150)], method='Powell' ) :
   def penalty(vv):      #Constraints Penalty
     penalty=0.0  
     for i,x in enumerate(vv) :
       penalty+= 5000000*( -min(x-bounds1[i][0], 0.0) + max(x-bounds1[i][1],0.0))   
     return   penalty

   def loss(vv, argext):
     return fun_obj(vv,argext) + penalty(vv)

   res= sci.optimize.minimize(loss, x0, args=argext, method=method,   bounds=bounds1, tol= 0.001)
   return res

def np_minimizeDE(fun_obj, bounds, name1, maxiter=10, popsize=5, solver=None):
  solver= sci.optimize._differentialevolution.DifferentialEvolutionSolver(fun_obj, bounds=bounds, popsize=popsize) 
  imin=0      
      
  name1= '/batch/solver_'+ name1 
  fbest0=1500000.0
  for i in range(imin, imin+maxiter):
    xbest, fbest = next(solver)              
    print(0,i, fbest, xbest)
    res= (copy.deepcopy(solver), i, xbest, fbest)  
    try :
         save(res, name1+date_now()+'_'+np_int_tostr(i))
    except :  pass      
    if np.mod(i+1, 11)==0 :
            if np.abs(fbest - fbest0) < 0.001 : break;
            fbest0= fbest  

def np_remove_NA_INF_2d(X):
 im, jm= np.shape(X)
 for i in range(0, im) :
   for j in range(0,jm) :
    if np.isnan(X[i,j]) or np.isinf(X[i,j]): X[i,j]= X[i-1,j]
 return X
  
def np_addcolumn(arr, nbcol):
  sh= np.shape(arr);  
  vv= np.zeros((sh[0], nbcol))
  arr2= np.column_stack((arr, vv))
  return arr2

def np_addrow(arr, nbrow):
  sh= np.shape(arr);
  if len(sh)  > 1 :  
   vv= np.zeros((nbrow, sh[1]))
   arr2= np.row_stack((arr, vv))
   return arr2
  else:
   return np.append(arr, np.zeros(nbrow))

def np_int_tostr(i):
  if i < 10 : return '00'+str(i)
  elif i < 100 : return '0'+str(i)
  else: return str(i)


def np_dictordered_create():
   from collections import OrderedDict
   return OrderedDict()

def np_list_unique(seq) :   
 from sets import Set
 set = Set(seq)
 return list(set)


def np_list_tofreqdict(l1, wweight=[]) :
   # Get list frequency
   dd= dict()
   if len(wweight) == 0 :
     for x in l1 :
       try :    dd[x]+= 1
       except : dd[x]=  1
     return dd
   else :
     for ii, x in enumerate(l1) :
        try :     dd[x]+= wweight[ii]
        except :  dd[x]=  wweight[ii]
     return dd



# used to flatten a list or tupel [1,2[3,4],[5,[6,7]]] -> [1,2,3,4,5,6,7]
def np_list_flatten(seq):
    l = []
    for elt in seq:
        t = type(elt)
        if t is tuple or t is list:
            for elt2 in flatten(elt):  l.append(elt2)
        else:   l.append(elt)
    return l



def np_dict_tolist(dd , withkey=0) :
   if withkey :
     return [ [key,val]  for key, val in list(dd.items()) ]
   else:
     return [ val  for _, val in list(dd.items()) ]

def np_dict_tostr_val(dd) :
    return ','.join([ str(val)  for _, val in list(dd.items()) ])

def np_dict_tostr_key(dd) :
    return ','.join([ str(key)  for key,_ in list(dd.items()) ])



def np_removelist(x0, xremove=[]) :
  xnew=[]
  for x in x0 :
    if np_findfirst(x, xremove) < 0 : xnew.append(x)
  return xnew

def np_transform2d_int_1d(m2d, onlyhalf=False) :
  imax, jmax= np.shape(m2d)
  v1d= np.zeros((imax*jmax, 4))  ; k=0
  for i  in range(0,imax):
    for j in range(i+1, jmax):
       v1d[k,0]= i; v1d[k,1]=j
       v1d[k,2]= m2d[i,j]
       v1d[k,3]= np.abs(m2d[i,j])
       k+=1
  v1d= v1d[v1d[:,2] !=0]
  return np_sortbycol(v1d,3, asc=False)

def np_mergelist(x0, x1) :
  xnew= list(x0)
  for x in x1 :
    xnew.append(x)
  return list(xnew)

def np_enumerate2(vec_1d):
 v2= np.empty((len(vec_1d), 2))
 for k,x in enumerate(vec_1d): v2[k,0]= k; v2[k,1]= x 
 return v2

#Pivot Table from List data    
def np_pivottable_count(mylist):
    mydict={}.fromkeys(mylist,0)
    for e in mylist:
        mydict[e]= mydict[e] + 1   # Map Reduce function 
    ll2= np_dict_tolist(mydict)
    ll2= sorted(ll2, key = lambda x: int(x[1]), reverse=True)
    return ll2

def np_nan_helper(y):
    """ Input:  - y, 1d numpy array with possible NaNs
        Output - nans, logical indices of NaNs - index, a function, with signature indices= index(logical_indices),
              to convert logical indices of NaNs to 'equivalent' indices
    """
    return np.isnan(y), lambda z: z.nonzero()[0]

def np_interpolate_nan(y):
 nans, x= np_nan_helper(y)
 y[nans]= np.interp(x(nans), x(~nans), y[~nans])
 return y

def np_and1(x,y, x3=None, x4=None, x5=None, x6=None, x7=None, x8=None):
  if not x8 is None :  return np.logical_and.reduce((x8,x7, x6, x5, x4,x3,x,y))
  if not x7 is None :  return np.logical_and.reduce((x7, x6, x5, x4,x3,x,y))
  if not x6 is None:  return np.logical_and.reduce((x6, x5, x4,x3,x,y))
  if not x5 is None:  return np.logical_and.reduce((x5, x4,x3,x,y))
  if not x4 is None :  return np.logical_and.reduce((x4,x3,x,y))
  if not x3 is None :  return np.logical_and.reduce((x3,x,y))

def np_sortcol(arr, colid, asc=1):
 """ df.sort(['A', 'B'], ascending=[1, 0])  """
 df = pd.DataFrame(arr)
 arr= df.sort_values(colid, ascending=asc)   
 return arr.values

def np_sort(arr, colid, asc=1):
 """ df.sort(['A', 'B'], ascending=[1, 0])  """
 df = pd.DataFrame(arr)
 arr= df.sort_values(colid, ascending=asc)   
 return arr.values

def np_ma(vv, n):
  '''Moving average '''
  return np.convolve(vv, np.ones((n,))/n)[(n-1):]


@jit(float32[:,:](float32[:,:]))
def np_cleanmatrix(m):
  m= np.nan_to_num(m)
  imax, jmax= np.shape(m)
  for i in range(0,imax):
    for j in range(0,jmax):
       if abs(m[i,j]) > 300000.0 : m[i,j]= 0.0
  return m

def np_torecarray(arr, colname) :
  return  np.core.records.fromarrays(arr.T, names=','.join(colname), formats = ','.join(['f8'] * np.shape(arr)[1]))

def np_sortbycolumn(arr, colid, asc=True): 
 df = pd.DataFrame(arr)
 arr= df.sort_values(colid, ascending=asc)   
 return arr.values

def np_sortbycol(arr, colid, asc=True): 
 if len(np.shape(arr)) > 1 :
  df = pd.DataFrame(arr)
  arr= df.sort_values(colid, ascending=asc)   
  return arr.values
 else :
  return np.reshape(arr, (1,len(arr)))  
    
# if colid==0 : return arr[np.argsort(arr[:])]
# else : return arr[np.argsort(arr[:, colid])]

def min_kpos(arr, kth):
   ''' return kth mininimun '''
   return np.partition(arr, kth-1)[kth-1]

def max_kpos(arr, kth):
   ''' return kth mininimun '''
   n= len(arr)
   return np.partition(arr, n-kth+1)[n-kth+1-1]


@jit(nopython=True, nogil=True)    #Only numericalpostioon
def np_findfirst(item, vec):
    """return the index of the first occurence of item in vec"""
    for i in range(len(vec)):
        if item == vec[i]: return i
    return -1
    
@jit(nopython=True, nogil=True) 
def np_find(item, vec):
    """return the index of the first occurence of item in vec"""
    for i in range(len(vec)):
        if item == vec[i]: return i
    return -1


def find(item, vec):
    """return the index of the first occurence of item in vec"""
    for i in range(len(vec)):
        if item == vec[i]: return i
    return -1

def findnone(vec):
    """return the index of the first occurence of item in vec"""
    for i in range(len(vec)):
        if vec[i]==None : return i
    return -1

def findx(item, vec):
    """return the index of the first occurence of item in vec"""
    try :    
      if type(vec)== list : i2= vec.index(item)
      else :   
        i=np.where(vec==item)[0]
        i2= i[0] if len(i) > 0 else -1
    except: 
       i2= -1
    return i2

def finds(itemlist, vec):
  """return the index of the first occurence of item in vec"""
  idlist= []  
  for x in itemlist :
    ix=-1    
    for i in range(len(vec)):
        if x == vec[i]:  idlist.append(i); ix=i
    if ix==-1: idlist.append(-1)       
  if idlist== [] : return -1 
  else : return idlist

def findhigher(x, vec):
    """return the index of the first occurence of item in vec"""
    for i in range(len(vec)):
        if vec[i] > x : return i
    return -1

def findlower(x, vec):
    """return the index of the first occurence of item in vec"""
    for i in range(len(vec)):
        if vec[i] < x: return i
    return -1

  
import operator
def np_find_minpos(values):
 min_index, min_value = min(enumerate(values), key=operator.itemgetter(1))
 return min_index, min_value
 
def np_find_maxpos(values):
 max_index, max_value = max(enumerate(values), key=operator.itemgetter(1))
 return max_index, max_value

def np_find_maxpos_2nd(numbers):
    count = 0
    m1 = m2 = float('-inf')
    for i,x in enumerate(numbers):
        count += 1
        if x > m2:
            if x >= m1:
                m1, m2 = x, m1            
            else:
                m2 = x; i2= i
    return i2, m2 if count >= 2 else None

def np_findlocalmax2(v, trig):
 n=len(v)  
 v2= np.zeros((n,8))
 tmax,_= np_find_maxpos(v)
 if n < 3:
   max_index, max_value= np_find_maxpos(v)
   v2=  [[max_index, max_value]] 
   return v2
 else:
  for i,x in enumerate(v):
    if i < n-1 and i > 0 :
     if x > v[i-1] and x > v[i+1] : 
           v2[i,0]= i;  v2[i,1]=x
  v2 =np_sortbycolumn(v2,1,asc=False)
 
  #Specify the max :
 for k in range(0, len(v2)) :
  kmax= v2[k,0]
  kmaxl=  findhigher(v2[k,1], v[:kmax][::-1])  #Find same level of max
  kmaxr=  findhigher(v2[k,1], v[kmax+1:])

  kmaxl= 0 if kmaxl==-1 else kmax-kmaxl
  kmaxr= n if kmaxr==-1 else kmaxr+kmax
  
  v2[k,2]= np.abs(kmaxr-kmaxl)   #Range 
  v2[k,3]= np.abs(kmax-tmax)   #Range of the Max After
  v2[k,4]= 0  #Range of the Max Before
  v2[k,5]= kmax-kmaxl
  v2[k,6]= kmaxr-kmax

 v2= v2[np.logical_and( v2[:,5] > trig , v2[:,6] > trig ) ]
 v2 =np_sortbycolumn(v2,0,asc=1)
 return v2

def np_findlocalmin2(v, trig):
 n=len(v)  
 v2= np.zeros((n,8))
 tmin,_= np_find_minpos(v)
 if n < 3:
   max_index, max_value= np_find_minpos(v)
   v2= [[max_index, max_value]]
   return v2
 else:
  for i,x in enumerate(v):
    if i < n-1 and i > 0 :
     if x < v[i-1] and x < v[i+1] :      
           v2[i,0]= i;  v2[i,1]=x
  v2 =np_sortbycolumn(v2,1,asc=False)
 
 #Classification of the Local min
  for k in range(0, len(v2)) :
   if v2[k,1] != 0.0 :
    kmin= v2[k,0]
    kminl=  findlower(v2[k,1], v[:kmin][::-1])  #Find same level of min
    kminr=  findlower(v2[k,1], v[kmin+1:])

    kminl= 0 if kminl==-1 else kmin-kminl
    kminr= n if kminr==-1 else kminr+kmin

    v2[k,2]= np.abs(kminr-kminl)   #Range 
    v2[k,3]= np.abs(kmin-tmin)   #Range of the min After
    v2[k,4]= 0  #Range of the min After
    v2[k,5]= kmin-kminl
    v2[k,6]= kminr-kmin
 v2= v2[np.logical_and( v2[:,5] > trig , v2[:,6] > trig ) ]
 v2 =np_sortbycolumn(v2,0,asc=1)
 return v2

def np_findlocalmax(v):
 n=len(v)   
 v2= np.zeros((n,2))
 if n > 2:
  for i,x in enumerate(v):
    if i < n-1 and i > 0 :
     if x > v[i-1] and x > v[i+1] : 
           v2[i,0]= i;  v2[i,1]=x
  v2 =np_sortbycolumn(v2,1,asc=False)
  return v2
 else : 
   max_index, max_value= np_find_maxpos(v)
   return  [[max_index, max_value]]

def np_findlocalmin(v):
 n=len(v)   
 v2= np.zeros((n,2))
 if n > 2:
  for i,x in enumerate(v):
    if i < n-1 and i > 0 :
     if x < v[i-1] and x < v[i+1] :      
           v2[i,0]= i;  v2[i,1]=x
 
  v2 =np_sortbycolumn(v2[:i2],0,asc=True)
  return v2
 else : 
   
   max_index, max_value= np_find_minpos(v)
   return  [[max_index, max_value]]

def np_stack(v1, v2=None, v3=None, v4=None, v5=None):
   sh= np.shape(v1)
   if sh[0] < sh[1] :    
     v1= np.row_stack((v1,v2))
     if v3 != None : v1= np.row_stack((v1,v3))
     if v4 != None : v1= np.row_stack((v1,v4))
     if v5 != None : v1= np.row_stack((v1,v5))
   else :
     v1= np.column_stack((v1,v2))
     if v3 != None : v1= np.column_stack((v1,v3))
     if v4 != None : v1= np.column_stack((v1,v4))
     if v5 != None : v1= np.column_stack((v1,v5))       
       
   return v1

def np_uniquerows(a):
    a = np.ascontiguousarray(a)
    unique_a = np.unique(a.view([('', a.dtype)]*a.shape[1]))
    return unique_a.view(a.dtype).reshape((unique_a.shape[0], a.shape[1]))

def np_remove_zeros(vv, axis1=1):
   return vv[~np.all(vv == 0, axis=axis1)]

def np_sort(vv): 
 return vv[np.lexsort(np.transpose(vv)[::-1])]      #Sort the array by different column

def np_memory_array_adress(x):
    # This function returns the memory block address of an array.# b = a.copy(); id(b) == aid
    return x.__array_interface__['data'][0]

    


def np_pivotable_create(table, left, top, value):
    """
    Creates a cross-tab or pivot table from a normalised input table. Use this
    function to 'denormalize' a table of normalized records.

    * The table argument can be a list of dictionaries or a Table object.
    (http://aspn.activestate.com/ASPN/Cookbook/Python/Recipe/334621)
    * The left argument is a tuple of headings which are displayed down the
    left side of the new table.
    * The top argument is a tuple of headings which are displayed across the
    top of the new table.
    Tuples are used so that multiple element headings and columns can be used.

    E.g. To transform the list (listOfDicts):

    Name,   Year,  Value
    -----------------------
    'Simon', 2004, 32
    'Simon', 2005, 128
    'Russel', 2004, 64
    'Eric', 2004, 52
    'Russel', 2005, 32

    into the new list:

    'Name',   2004, 2005
    ------------------------
    'Simon',  32,     128
    'Russel',  64,     32
    'Eric',   52,     NA

    you would call pivot with the arguments:

    newList = pivot(listOfDicts, ('Name',), ('Year',), 'Value')

    """
    rs = {}
    ysort = []
    xsort = []
    for row in table:
        yaxis = tuple([row[c] for c in left])       # e.g. yaxis = ('Simon',)
        if yaxis not in ysort: ysort.append(yaxis)
        xaxis = tuple([row[c] for c in top])        # e.g. xaxis = ('2004',)
        if xaxis not in xsort: xsort.append(xaxis)
        try:
            rs[yaxis]
        except KeyError:
            rs[yaxis] = {}
        if xaxis not in rs[yaxis]:
            rs[yaxis][xaxis] = 0
        rs[yaxis][xaxis] += row[value]


    # In the following loop we take care of missing data,
    # e.g 'Eric' has a value in 2004 but not in 2005
    for key in rs:
        if len(rs[key]) > len(xsort):
            for var in xsort:
                if var not in list(rs[key].keys()):
                    rs[key][var] = ''

    headings = list(left)
    xsort.sort()
    headings.extend(xsort)

    t = []

    # The lists 'sortedkeys' and 'sortedvalues' make sure that
    # even if the field 'top' is unordered, data will be transposed correctly.
    # E.g. in the example above the table rows are not ordered by the year
    for left in ysort:
        row = list(left)
        sortedkeys = list(rs[left].keys())
        sortedkeys.sort()
        sortedvalues = list(map(rs[left].get, sortedkeys))
        row.extend(sortedvalues)
        t.append(dict(list(zip(headings,row))))
    return t




############### PANDA UTIL  ################################################################


def pd_info(df, doreturn=1) :
 df.info()
 dtype0= df.dtypes.to_dict()
 for i in df.columns :
   print(i,  dtype0[i], type(df[i].values[0]),  df[i].values[0])
 if doreturn : return dtype0

def pd_info_memsize(df, memusage=0):
 df.info(memory_usage='deep')
 pd_info(df)
 if memusage==1 :return df.memory_usage().sum()



def pd_row_findlast(df, colid=0, emptyrowid=None) :
 for ii, rr in df.iterrows() :
   if rr[colid] is emptyrowid : return ii


# selection
def pd_row_select(df, **conditions):
    '''Select rows from a df according to conditions
    pdselect(data, a=2, b__lt=3) __gt __ge __lte  __in  __not_in
    will select all rows where 'a' is 2 and 'b' is less than 3
    '''
    if type(df) == pd.Series:   df = pd.DataFrame({'value': df})
    for c, value in list(conditions.items()):
        if c in df.columns:
            c = (df[c]==value)
        elif c.endswith('__neq') or c.endswith('__not_eq'):
            if c.endswith('__neq'):
                c =c[:-len('__neq')]
            elif c.endswith('__not_eq'):
                c =c[:-len('__not_eq')]
            c =df[c]!=value
        elif c.endswith('__gt'):
            c =c[:-len('__gt')]
            c =df[c] > value
        elif c.endswith('__ge'):
            c =c[:-len('__ge')]
            c =df[c] >= value
        elif c.endswith('__gte'):
            c =c[:-len('__gte')]
            c =df[c] >= value
        elif c.endswith('__lt'):
            c =c[:-len('__lt')]
            c =df[c] < value
        elif c.endswith('__le'):
            c =c[:-len('__le')]
            c =df[c] <= value
        elif c.endswith('__lte'):
            c =c[:-len('__lte')]
            c =df[c] <= value
        elif c.endswith('__in'):
            c =c[:-len('__in')]
            c = np.in1d(df[c], value)
        elif c.endswith('__not_in'):
            c =c[:-len('__not_in')]
            c = np.in1d(df[c], value, invert=True)
        else:
            raise ValueError("Cannot process condition '{}'".format(c))
        df = df[c]
    return df

def pd_csv_randomread(filename, nsample=10000, filemaxline=-1, dtype=None) :
 if filemaxline==-1 : n = sum(1 for line in open(filename)) - 1 #number of records in file (excludes header)
 else : n= filemaxline
 skip = np.sort(np.random.randint(1,n+1,n- nsample)) #the 0-indexed header will not be included in the skip list
 df= pd.read_csv(filename, skiprows=skip, dtype=dtype)
 return df



# Creation
def pd_array_todataframe(array, colname=None, index1=None, dotranspose=False):
   sh= np.shape(array)
   if len(sh) > 1 :
     if sh[0] < sh[1] and dotranspose :   #  masset x time , need Transpose
         return pd.DataFrame(data= array.T, index= index1, columns=colname)
     else :
         return pd.DataFrame(data= array, index= index1, columns=colname)
   else :   #1d vector
     return pd.DataFrame(data= np.array(array).reshape(-1, 1), index= index1, columns=colname)

def pd_dataframe_toarray(df):
  col1= df.index
  array1= df.reset_index().values[1:,:]
  column_name= df.columns
  return column_name, col1, array1

def pd_createdf(array1, col1=None, idx1=None) :
  return  pd.DataFrame(data=array1, index=idx1, columns=col1)

def pd_create_colmapdict_nametoint(df) :
  ''' 'close' ---> 5    '''
  col= df.columns.values; dict1={}
  for k,x in enumerate(col) :
    dict1[x]= k
  return dict1



# Extract
def pd_extract_col_idx_val(df):
 return df.columns.values,  df.index.values, df.values


def pd_extract_col_uniquevalue_tocsv(df, colname='', csvfile='') :
 ''' Write one column into a file   '''
 a= df[colname].unique()
 a= np.array(a)
 pd.DataFrame( a, columns=[colname]).to_csv( csvfile )
 print(csvfile)



def pd_split_col_idx_val(df):
 return df.columns.values,  df.index.values, df.values

def pd_splitdf_inlist(df, colid, type1="dict"):
    ''' Split df into dictionnary of dict/list '''
    UniqueNames = df.sym.unique()
    if type1=='dict' :
      dfDict = {elem: pd.DataFrame for elem in UniqueNames}
      for key in list(DataFrameDict.keys()):
        dfDict[key] = df[df[colid] == key]
      return dfDict

    if type1 == "list":
      l1 = []
      for key in UniqueNames:
        l1.append(df[df[colid] == key])
      return l1

def pd_find(df, regex_pattern='*', col_restrict=[], isnumeric=False, doreturnposition=False) :
 ''' Find string / numeric values inside df columns, return position where found
     col_restrict : restrict to these columns '''
 dtype0= df.dtypes.to_dict()
 col0=   df.columns if col_restrict==[] else col_restrict

 if not isnumeric :   #object string columns
   colx= [col for col in col0 if str(dtype0[col])== 'object'   ]
   print(("Searching Cols: "+ str(colx)))
   for i, coli in enumerate(colx) :
     dfm= df[coli].str.contains(regex_pattern, na=False, regex=True)
     if i==0 :  mask = dfm
     else    :  mask=  np.column_stack((mask, dfm))

 else :
   numval= regex_pattern
   colx= [col for col in col0 if str(dtype0[col]).find('float') > -1 or str(dtype0[col]).find('int') > -1 ]
   print(("Searching Cols: "+ str(colx)))
   for i, coli in enumerate(colx) :
     dfm= (df[coli]== numval).values        #df[col].loc[ df[col]== numval  ]
     if i==0 :  mask = dfm
     else    :  mask=  np.column_stack((mask, dfm))

 # print mask
 if len(mask.shape) < 2 :  mask= mask.values.reshape(-1,1)

 gc.collect()
 if doreturnposition :
   locate= np_dictordered_create()        #Position in Dict_Column --> Indice_j
   for j in range(0, mask.shape[1])  :
     pos_tempj= np.array([ i for i in range(0,mask.shape[0])   if mask[i,j] ], dtype= np.int32)
     locate[colx[j]] = pos_tempj

   # locate= np.array([ (colx[j], i) for i in xrange(0,mask.shape[0]) for j in xrange(0, mask.shape[1])  if mask[i,j] ])
   return df.loc[mask.any(axis=1)], locate
 else :
   return df.loc[mask.any(axis=1)]



#dtypes
'''
def pd_dtypes_tocategory(df, columns=[], targetype='category'):
   for col in columns : df[col]= df[col].astype(targetype)
   return df
'''
def pd_dtypes_totype2(df, columns=[], targetype='category'):
   for col in columns : df[col]= df[col].astype(targetype)
   return df


def pd_dtypes(df, returnasdict=0) :
 from collections import OrderedDict
 dtype0= OrderedDict(df.dtypes.apply(lambda x: x.name))
 ss='''{ '''
 for i, col in enumerate(df.columns.values) :
   ss+= "'"+col+ "':" + "'"+ dtype0[col]+"', "
   if i%3==0 and i> -1 : ss+= '\n'
 ss= ss[0:-2] + ' }'

 if returnasdict : return eval(ss)
 print(ss)
 print('''\n df.astype(typedict)  Pandas 'object' : 'category', 'unicode' , 'str'  'boolean',
    	float16, float32, int8, int16, int32,uint8, uint16, uint32 ''')


def pd_df_todict2(df1, colkey='table', excludekey=[''], onlyfirstelt= True) :
      df= df.drop_duplicates(colkey).reset_index(level=0, drop=True)
      dict0 = {}
      for i in range(0, len(df)):
         id0 = df.iloc[i,0]
         val0= df.iloc[i,1]
         if id0 not in excludekey :
            dict0.setdefault(id0, [])
            if onlyfirstelt :  dict0[id0]= val0
            else:              dict0[id0].append(val0)
      return dict0

      
def pd_df_todict(df, colkey='machine_code', colval='adress') :
   dict0 = {}
   for ii, row in df.iterrows() :
     try :
        dict0[ row[colkey] ]=  row[colval]
     except :
        pass
         
   return dict0   
      
  


def pd_col_addfrom_dfmap(df, dfmap, colkey, colval, df_colused, df_colnew, exceptval=-1, inplace= True) :
  ''' Add new columns based on df_map:  In Place Modification of df
    df:     Dataframe of transactions.
    dfmap:  FSMaster Dataframe
      colkey: colum used for dict key.  machine_code
      colval: colum used for dict val.  adress
      
    df_colused  :     "machine_code"
    exception val:  -1 or ''
  '''
  map_dict= pd_df_todict(dfmap, colkey= colkey , colval=colval )  
  
  def map_dict_fun( rowi ) :
    try :     return map_dict[  rowi[ df_colused ]  ]
    except :  return exceptval
    
  df[df_colnew]=  df.apply( lambda x : map_dict_fun(x) , axis= 1) 
    
  if inplace : return None
  else  :      return df
  

   
   
'''
def pd_dtypes_getdict(df=None, csvfile=None) :
   if df is not None :   return df.dtypes.to_dict()
   elif csvfile is not None :
      df= pd.read_csv(csvfile, nrows= 1000)
      return df.dtypes.to_dict()


def pd_dtypes_getblaze(df1) :
 from collections import OrderedDict
 x= str(OrderedDict(df1.dtypes.apply(lambda x: x.name)))
 x=x.replace("',", "':")
 x=x.replace("(", "")
 x=x.replace(")", "")
 x=x.replace("OrderedDict[", "{")
 x=x.replace("]", "}")
 print(" string 5char: |S5 object in Pandas,  object_bool, string_, unicode
	float16, float32, float64, int8, int16, int32, int64,	uint8, uint16, uint32, uint64")
 return x
'''


'''
 # return df1.dtypes.apply(lambda x: x.name).to_dict()
 from collections import OrderedDict
 ds=OrderedDict()
 for i,x in enumerate(df1.columns.values) :
   ds[x]=  df1.dtypes[i].name
 print ds
'''



# Apply / Transform df
def pd_applyfun_col(df, newcol, ff, use_colname="all/[colname]") :
   ''' use all Columns to compute values '''
   if use_colname=="all/[colname]" : df[newcol]= ff(df.values)
   else :  df[newcol]= ff(df[use_colname].values)
   return df

'''
def pd_cleanquote(q):
 col= q.columns.values
 for kid in col:
   if kid not in ['date', 'day','month','year'] :
      q[kid]= pd.to_numeric(q[kid], errors='coerce').values  #Put NA on string
 
 q= q.fillna(method='pad')  
 return q
'''

def pd_date_intersection(qlist) :
 date0= set(qlist[0]['date'].values)
 for qi in qlist :
   qs= set(qi['date'].values)
   date0 = set.intersection(date0, qs)
 date0= list(date0); date0= sorted(date0)
 return date0

def pd_is_categorical(z):
    if isinstance(z, pd.Categorical): return True
    try:
        return isinstance(z.values, pd.Categorical)
    except:
        return False

def pd_str_encoding_change(df, cols, fromenc='iso-8859-1', toenc='utf-8'):
    #  Western European: 'cp1252'
    for col in cols:
        df[col] = df[col].str.decode(fromenc).str.encode(toenc)
    return df

def pd_str_unicode_tostr(df, targetype=str) :
 '''
 https://www.azavea.com/blog/2014/03/24/solving-unicode-problems-in-python-2-7/
 Nearly every Unicode problem can be solved by the proper application of these tools;
 they will help you build an airlock to keep the inside of your code nice and clean:

encode(): Gets you from Unicode -> bytes
decode(): Gets you from bytes -> Unicode
codecs.open(encoding=”utf-8″): Read and write files directly to/from Unicode (you can use any encoding,
 not just utf-8, but utf-8 is most common).
u”: Makes your string literals into Unicode objects rather than byte sequences.
Warning: Don’t use encode() on bytes or decode() on Unicode objects

>>> uni_greeting % utf8_name
Traceback (most recent call last):
 File "<stdin>", line 1, in <module>
UnicodeDecodeError: 'ascii' codec can't decode byte 0xc3 in position 3: ordinal not in range(128)
# Solution:
>>> uni_greeting % utf8_name.decode('utf-8')
u'Hi, my name is Josxe9.'

 '''
 return pd_dtypes_type1_totype2(df, fromtype=str, targetype=str)

def pd_dtypes_type1_totype2(df, fromtype=str, targetype=str) :
 for ii in df.columns :
  if isinstance(df[ii].values[0], fromtype) :
     df[ii]= df[ii].astype(targetype)
 return df



#----Insert/update col row  -----------------------------------------
def pd_resetindex(df):
  df.index = list(np.arange(0,len(df.index)))
  return df

def pd_insertdatecol(df_insider,  format1="%Y-%m-%d %H:%M:%S:%f") :
 df_insider= pd_addcol(df_insider, 'dateinsert')
 df_insider['dateinsert']=  df_insider['dateinsert'].astype(str)
 df_insider.dateinsert.values[:]  =  date_nowtime('str', format1= format1)
 return df_insider

def pd_replacevalues(df,  matrix):
 ''' Matrix replaces df.values  '''
 imax, jmax= np.shape(vec)
 colname= df.columns.values
 for j in jmax :
   df.loc[colname[j]] = matrix[:,j]

 return df

def pd_removerow(df, row_list_index=[23, 45]) :
 return df.drop(row_list_index)

def pd_removecol(df1, name1):
 return df1.drop(name1, axis=1)


def pd_insertrow(df, rowval, index1=None, isreset=1):
 df2=  pd_array_todataframe(rowval, df.columns.values, index1 )
 df=   df.append(df2, ignore_index=True)
 #if isreset : df.reset_index(inplace=True)
 return df


#---- h5 In / Out  ------------------------------------------------
def pd_h5_cleanbeforesave(df):
   '''Clean Column type before Saving in HDFS: Unicode, Datetime  '''
   #Unicode String :
   df= pd_resetindex(df)   #Reset Index 0 to 100000
   df= pd_str_unicode_tostr(df, targetype=str)

   '''
   for col in  ['date'] :
   # df_pd[col] = df_pd[col].astype(str)
   df_pd[col] = df_pd[col].apply(lambda x: x.encode('utf-8').strip())

   df_pd.to_hdf('test.h5','df',format='table',mode='w',data_columns=True,encoding='latin1')

   '''
   return df

def pd_h5_addtable(df, tablename, dbfile='F:\temp_pandas.h5') :
  store = pd.HDFStore(dbfile)
  if find(tablename, list(store.keys())) > 0 :
     # tablename=tablename + '_1';
     print('Table Exist, change table name')
  else :
     store.append(tablename, df); store.close()


def pd_h5_tableinfo(filenameh5, table):
    store= pd.HDFStore(filenameh5)
    return store.get_storer(table).table


def pd_h5_dumpinfo(dbfile='E:\_data\stock\intraday_google.h5'):
  store = pd.HDFStore(dbfile)
  extract=[]; errsym=[]
  for symbol in list(store.keys()):
     try:
       df= pd.read_hdf(dbfile, symbol)
       extract.append([symbol[1:], df.shape[1],
                       df.shape[0], datetime_tostring(df.index.values[0]),
                        datetime_tostring(df.index.values[-1]) ])

     except: errsym.append(symbol)
  return np.array(extract), errsym

def pd_h5_save(df, filenameh5='E:/_data/_data_outlier.h5', key='data'):
 ''' File is release after saving it'''
 store = pd.HDFStore(filenameh5)  
 store.append(key, df); store.close()

def pd_h5_load(filenameh5='E:/_data/_data_outlier.h5', table_id='data', exportype="pandas", rowstart=-1, rowend=-1, cols=[]):
  if rowend==-1 :  df=  pd.read_hdf(filenameh5, table_id)
  else :           df=  pd.read_hdf(filenameh5, table_id,  start=rowstart, end=rowend)
  if exportype=='pandas':   return df
  elif exportype=='numpy' : return df.values

def pd_h5_fromcsv_tohdfs(dircsv='dir1/dir2/', filepattern='*.csv', tofilehdfs='file1.h5', tablename='df', col_category=[],
                         dtype0=None,  encoding='utf-8', chunksize= 2000000, mode='a',format='table', complib=None):

  csvlist=  os_file_listall(dircsv, filepattern)
  csvlist=  csvlist[2]   # 2: Full_path + filename

  store = pd.HDFStore(tofilehdfs)
  max_size= {}
  for ii, tablei_file in enumerate(csvlist) :

    #Inference of Type, String size from top 1000 records......
    if ii== 0 :
       if dtype0 is None :
         df_i= pd.read_csv(tablei_file, nrows=1000, sep=',') #Use Top 1000 to estimate size....
         dtype0= df_i.dtypes.to_dict()
         col_list=[]
         # for col, x  in dtype0.items():
         #   if x == np.dtype('O') :    #Object == date, string, need to convert to string....
         #      col_list.append(col)
             # df_i[col] =    df_i[col].map(lambda x:  str(str(x).encode(encoding)))
             # max_size[col]= max(15, df_i[col].str.len().max())

    #ISSUE: DO NOT USE encoding='utf-8' when reading the Japanese Character.....
    list01= pd.read_csv(tablei_file,  chunksize=chunksize, dtype= dtype0, sep=',')  #, encoding=encoding)
    for k, df_i in enumerate(list01) :
      print('.', end=' ')
      for col in col_category :
          df_i[col] =    df_i[col].astype('category')

      #for col in col_list :
         # df_i[col] = df_i[col].map(lambda x:  str(str(x).encode(encoding)))
      store.append(tablename, df_i,  mode=mode, format=format, complib=complib)  #,  min_itemsize=max_size)

  store.close(); del store; print('\n')
  return os_file_exist(tofilehdfs)


def pd_np_toh5file(numpyarr, fileout="file.h5", table1='data'):
 pd= pd.DataFrame(numpyarr);  st= pd.HDFStore(fileout);  st.append(table1, pd); del pd



##############Date Manipulation#######################################################################
from dateutil import parser

def date_allinfo():
   '''

   https://aboutsimon.com/blog/2016/08/04/datetime-vs-Arrow-vs-Pendulum-vs-Delorean-vs-udatetime.html


   '''


def date_convert(t1, fromtype, totype) :
   try :
      n= len(t1)
   except : t1= [t1]; n=1
   t0= t1[0]

   if isinstance(t0, str) :
      pass

   elif isinstance(t0, int) :
      pass

   elif isinstance(t0, datetime) :
      pass

   elif isinstance(t0, np.datetime64) :
     pass

   for t in t1 :
      t2= _dateconvert_from(t, type1)   #  to Datetime
      t3= _dateconvert_from(t, totype)   # Datetime  to target type
      tlist.append(t3)
   return tlist

def datetime_tostring(datelist1):
 if isinstance(datelist1, datetime.date) :      return datelist1.strftime("%Y%m%d")
 if isinstance(datelist1, np.datetime64 ) :
   t= pd.to_datetime(str(datelist1)) ;  return t.strftime("%Y%m%d") 
  
 date2 = []
 for t in datelist1:
     date2.append(t.strftime("%Y%m%d"))
 return date2

def date_remove_bdays(from_date, add_days):
    isint1= isint(from_date)  
    if isint1 :  from_date= dateint_todatetime(from_date)
    business_days_to_add = add_days
    current_date = from_date
    while business_days_to_add < 0:
        current_date += datetime.timedelta(days=-1)
        weekday = current_date.weekday()
        if weekday >= 5: # sunday = 6
            continue
        business_days_to_add += 1
    if isint1 : return datetime_toint(current_date) 
    else :      return   current_date

def date_add_bdays(from_date, add_days):
    isint1= isint(from_date)
    if isint1 :  from_date= dateint_todatetime(from_date)
    business_days_to_add = add_days
    current_date = from_date
    while business_days_to_add > 0:
        current_date += datetime.timedelta(days=1)
        weekday = current_date.weekday()
        if weekday >= 5: # sunday = 6
            continue
        business_days_to_add -= 1
    if isint1 : return datetime_toint(current_date) 
    else :      return   current_date

def datenumpy_todatetime(tt, islocaltime=True):
   #  http://stackoverflow.com/questions/29753060/how-to-convert-numpy-datetime64-into-datetime
   if type(tt)==np.datetime64:
      if islocaltime:
         return datetime.datetime.fromtimestamp(tt.astype('O') / 1e9)
      else:
         return datetime.datetime.utcfromtimestamp(tt.astype('O') / 1e9)
   elif type(tt[0])==np.datetime64:
      if islocaltime:
         v=[datetime.datetime.fromtimestamp(t.astype('O') / 1e9) for t in tt]
      else:
         v=[datetime.datetime.utcfromtimestamp(t.astype('O') / 1e9) for t in tt]
      return v
   else:
      return tt  # datetime case

def datetime_tonumpydate(t, islocaltime=True):
   #  http://stackoverflow.com/questions/29753060/how-to-convert-numpy-datetime64-into-datetime
   return np.datetime64(t)

def datestring_todatetime(datelist1, format1= "%Y%m%d") :
 if isinstance(datelist1, str)  :  return parser.parse(datelist1)
 date2 = []
 for s in datelist1:
     date2.append(parser.parse(s))
     #date2.append(datetime.datetime.strptime(s, format1))
 return date2    

def datetime_toint(datelist1):
 if isinstance(datelist1, datetime.date) :      return int(datelist1.strftime("%Y%m%d"))
 date2 = []
 for t in datelist1:
     date2.append(int(t.strftime("%Y%m%d")))
 return date2

def date_holiday():
   '''
   https://jakevdp.github.io/blog/2015/07/23/learning-seattles-work-habits-from-bicycle-counts/

from pandas.tseries.holiday import USFederalHolidayCalendar
cal = USFederalHolidayCalendar()
holidays = cal.holidays('2012', '2016', return_name=True)
holidays.head()

holidays_all = pd.concat([holidays, "Day Before " + holidays.shift(-1, 'D'),  "Day After " + holidays.shift(1, 'D')])
holidays_all = holidays_all.sort_index()
holidays_all.head()

from pandas.tseries.offsets import CustomBusinessDay
from pandas.tseries.holiday import USFederalHolidayCalendar
bday_us = CustomBusinessDay(calendar=USFederalHolidayCalendar())
dateref[-1] - bday_us- bday_us

   '''

def date_add_bday(dint, nbday, country='us') :
  d= dateint_todatetime(dint)

  if country=='us'  :
     from pandas.tseries.offsets import CustomBusinessDay
     from pandas.tseries.holiday import USFederalHolidayCalendar
     bday_us = CustomBusinessDay(calendar=USFederalHolidayCalendar())
     d= d + bday_us*nbday
  else :
     d= d+ pd.tseries.offsets.BDay(nbday)

  return datetime_toint(d)

def dateint_todatetime(datelist1) :
 if isinstance(datelist1, int)  :  return parser.parse(str(datelist1))
 date2 = []
 for s in datelist1:
     date2.append(parser.parse(str(s)))
     #date2.append(datetime.datetime.strptime(s, format1))
 return date2    

def date_diffinday(intdate1, intdate2):
  dt= dateint_todatetime(intdate2) - dateint_todatetime(intdate1)
  return dt.days

def date_diffinyear(startdate, enddate):
 return date_as_float(startdate) - date_as_float(enddate)

def date_diffinbday( intd2, intd1)   :
  d1= dateint_todatetime(intd1)
  d2= dateint_todatetime(intd2)
  d1= d1.date()
  d2= d2.date()

  return np.busday_count(d1, d2)

def date_gencalendar(start='2010-01-01', end='2010-01-15', country='us') :
  from pandas.tseries.holiday import USFederalHolidayCalendar
  from pandas.tseries.offsets import CustomBusinessDay
  us_bd = CustomBusinessDay(calendar=USFederalHolidayCalendar())
  return np.arrray(pd.DatetimeIndex(start=start,end=end, freq=us_bd))

def date_finddateid(date1, dateref) :
  i= np_findfirst(date1, dateref)
  if i==-1 : i= np_findfirst(date1+1, dateref) 
  if i==-1 : i= np_findfirst(date1-1, dateref)     
  if i==-1 : i= np_findfirst(date1+2, dateref)    
  if i==-1 : i= np_findfirst(date1-2, dateref) 
  if i==-1 : i= np_findfirst(date1+3, dateref)    
  if i==-1 : i= np_findfirst(date1-3, dateref) 
  if i==-1 : i= np_findfirst(date1+5, dateref)    
  if i==-1 : i= np_findfirst(date1-5, dateref)  
  if i==-1 : i= np_findfirst(date1+7, dateref)    
  if i==-1 : i= np_findfirst(date1-7, dateref)  
  return i 
   
def datestring_toint(datelist1) :
 if isinstance(datelist1, str) :      return int(datelist1)    
 date2 = []
 for s in datelist1:   date2.append(int(s))
 return date2   

def date_now(i=0):
 from datetime import datetime
 d= datetime.now()
 if i > 0 : d= date_add_bdays(d,i)
 else:      d= date_remove_bdays(d,i)
 return str(datetime_toint(d))

def date_nowtime(type1='str', format1= "%Y-%m-%d %H:%M:%S:%f"):
 ''' str / stamp /  '''
 from datetime import datetime
 #d= datetime.now()
 d=datetime.today()  #today = datetime.today().strftime('%Y%m%d_%H%M%S%f')
 if   type1== 'str' :     return d.strftime(format1)
 elif type1== 'stamp' : return d.strftime('%Y%m%d_%H%M%S%f')
 else : return d

def date_tofloat(dt):
    size_of_day = 1. / 366.
    size_of_second = size_of_day / (24. * 60. * 60.)
    days_from_jan1 = dt - datetime.datetime(dt.year, 1, 1)
    if not isleap(dt.year) and days_from_jan1.days >= 31+28:
        days_from_jan1 += timedelta(1)
    return dt.year + days_from_jan1.days * size_of_day + days_from_jan1.seconds * size_of_second

def date_generatedatetime(start="20100101", nbday=10, end=""):
  from dateutil.rrule import DAILY, rrule, MO, TU, WE, TH, FR
  start= datestring_todatetime(start)
  if end=="" :  end = date_add_bdays(start,nbday-1) #  + datetime.timedelta(days=nbday) 
  date_list= list(rrule(DAILY, dtstart=start, until=end, byweekday=(MO,TU,WE,TH,FR)))
  
  return np.array(date_list)





############################# Utilities for Numerical Calc ######################################
def np_numexpr_vec_calc(filename, expr, i0=0, imax=1000, fileout='E:\_data\_QUASI_SOBOL_gaussian_xx3.h5'  ):
 ''' New= xx*xx  over very large series
 #numexpr_vect_calc(filename, 0, imax=16384*4096, "xx*xx", 'E:\_data\_QUASI_SOBOL_gaussian_xx3.h5'  ):
'''
 pdframe=  pd.read_hdf(filename,'data', start=i0, stop=imax)    #from file
 xx= pdframe.values;  del pdframe    #to numpy vector
 xx= ne.evaluate(expr)  
 pdf =pd.DataFrame(xx); del xx  
# filexx3=   'E:\_data\_QUASI_SOBOL_gaussian_xx3.h5'   
 store = pd.HDFStore(fileout) 
 store.append('data', pdf); del pdf



def np_numexpr_tohdfs(filename, expr,  i0=0, imax=1000, fileout='E:\_data\_QUASI_SOBOL_gaussian_xx3.h5'  ):
 pdframe=  pd.read_hdf(filename,'data', start=i0, stop=imax)    #from file
 xx= pdframe.values;  del pdframe   
 xx= ne.evaluate(expr)  
 pdf =pd.DataFrame(xx); del xx    # filexx3=   'E:\_data\_QUASI_SOBOL_gaussian_xx3.h5' 
 store = pd.HDFStore(fileout);  store.append('data', pdf); del pdf

#numexpr_vect_calc(filename, 0, imax=16384*4096, "xx*xx", 'E:\_data\_QUASI_SOBOL_gaussian_xx3.h5'  ):


# yy1= getrandom_tonumpy('E:\_data\_QUASI_SOBOL_gaussian_xx2.h5', 16384, 4096)
#----------------------------------------------------------------------------



##################### Statistics    #################################################################
def np_comoment(xx,yy,nsample, kx,ky) :
#   cx= ne.evaluate("sum(xx)") /  (nsample);   cy= ne.evaluate("sum( yy)")  /  (nsample)
#   cxy= ne.evaluate("sum((xx-cx)**kx * (yy-cy)**ky)") / (nsample)
   cxy= ne.evaluate("sum((xx)**kx * (yy)**ky)") / (nsample)
   return cxy 

def np_acf(data):
    #Autocorrelation
    n = len(data)
    mean = np.mean(data)
    c0 = np.sum((data - mean) ** 2) / float(n)

    def r(h):
        acf_lag = ((data[:n - h] - mean) * (data[h:] - mean)).sum() / float(n) / c0
        return acf_lag
    x = np.arange(n) # Avoiding lag 0 calculation
    acf_coeffs = np.asarray(list(map(r, x)))
    return acf_coeffs
 


##################### Plot Utilities ################################################################
def plot_XY(xx, yy, zcolor=None, tsize=None, title1='', xlabel='', ylabel='', figsize=(8, 6), dpi=75, savefile='') :

  #Color change
  if zcolor is None :  c= [[0, 0, 0]]
  elif isinstance(zcolor, int) : zcolor= zcolor
  else :
     aux= np.array(zcolor, dtype= np.float64)
     c= np.abs(aux)
  cmhot = plt.get_cmap("Blues")

  #Marker size
  if tsize is None :  tsize=  50
  elif isinstance(tsize, int) : tsize= tsize
  else :
     aux= np.array(tsize, dtype= np.float64)
     tsize=  np.abs(aux)
     tsize=(tsize - np.min(tsize)) / (np.max(tsize) - np.min(tsize)) * 130 + 1

  #Plot
  fig, ax1 = plt.subplots(nrows=1, ncols=1)

  fig.set_size_inches(figsize[0], figsize[1] )
  fig.set_dpi(dpi)
  fig.tight_layout()

  ax1.scatter(xx, yy, c=c, cmap=cmhot, s=tsize, alpha=0.5)
  ax1.set_xlabel(xlabel, fontsize=11)
  ax1.set_ylabel(ylabel, fontsize=11)
  ax1.set_title(title1)
  ax1.grid(True)
  # fig.autoscale(enable=True, axis='both')
  #fig.colorbar(ax1)

  if savefile != '' :
     plt.savefig(savefile, dpi=dpi)
     plt.clf()
  else :
     plt.show()


def plot_heatmap( frame,   ax=None, cmap=None,   vmin=None,vmax=None, interpolation='nearest'  ):
    from matplotlib import pyplot as plt
    if ax is None:
        ax = plt.gca()
    ax.set_xticks(np.arange(frame.shape[1]))
    ax.set_xticklabels(frame.columns, rotation='vertical')

    ax.set_yticks(np.arange(frame.shape[0]))
    ax.set_yticklabels(frame.index)
    ax.grid(False)
    ax.set_aspect('auto')
    ax.imshow(frame.values, interpolation=interpolation, cmap=cmap, vmin=vmin, vmax=vmax, aspect='auto')
    return



##########
def np_map_dict_to_bq_schema(source_dict, schema, dest_dict):
    '''
     new_dict = {}
     map_dict_to_bq_schema (my_dict, schema, new_dict)

     new_dict = {}
     map_dict_to_bq_schema (my_dict, schema, new_dict)

    :param source_dict:
    :param schema:
    :param dest_dict:
    :return:
    '''
    #iterate every field from current schema
    for field in schema['fields']:
        #only work in existant values
        if field['name'] in source_dict:
            #nested field
            if field['type'].lower()=='record' and 'fields' in field:
                #list
                if 'mode' in field and field['mode'].lower()=='repeated':
                    dest_dict[field['name']] = []
                    for item in source_dict[field['name']]:
                        new_item = {}
                        np_map_dict_to_bq_schema( item, field, new_item )
                        dest_dict[field['name']].append(new_item)
                #record
                else:
                    dest_dict[field['name']] = {}
                    np_map_dict_to_bq_schema( source_dict[field['name']], field, dest_dict[field['name']] )
            #list
            elif 'mode' in field and field['mode'].lower()=='repeated':
                dest_dict[field['name']] = []
                for item in source_dict[field['name']]:
                    dest_dict[field['name']].append(item)
            #plain field
            else:
                dest_dict[field['name']]=source_dict[field['name']]

                format_value_bq(source_dict[field['name']], field['type'])




########################### Google Drive    ###############################################################
def googledrive_get():
   '''
   https://github.com/ctberthiaume/gdcp
   ... I am using this now to transfer thousands of mp3 files from a ubuntu vps to google drive.


http://olivermarshall.net/how-to-upload-a-file-to-google-drive-from-the-command-line/
https://github.com/prasmussen/gdrive  : Super Complete

gdrive [global] upload [options] <path>

global:
  -c, --config <configDir>         Application path, default: /Users/<user>/.gdrive
  --refresh-token <refreshToken>   Oauth refresh token used to get access token (for advanced users)
  --access-token <accessToken>     Oauth access token, only recommended for short-lived requests because of short lifetime (for advanced users)

options:
  -r, --recursive           Upload directory recursively
  -p, --parent <parent>     Parent id, used to upload file to a specific directory, can be specified multiple times to give many parents
  --name <name>             Filename
  --no-progress             Hide progress
  --mime <mime>             Force mime type
  --share                   Share file
  --delete                  Delete local file when upload is successful
  --chunksize <chunksize>   Set chunk size in bytes, default: 8388608

   :return:
   '''
   pass

def googledrive_put():
  '''
  100 GB: 2USD,  1TB: 10USD
  https://gsuite.google.com/intl/en/pricing.html

  :return:
  '''
  pass


def googledrive_list():
   pass



def os_processify_fun(func):
    '''Decorator to run a function as a process.
    Be sure that every argument and the return value is *pickable*.
    The created process is joined, so the code does not  run in parallel.
    @processify

    def test():
      return os.getpid()

    @processify
    def test_deadlock():
      return range(30000)

   @processify
   def test_exception():
     raise RuntimeError('xyz')

   def test():
     print os.getpid()
     print test_function()
     print len(test_deadlock())
     test_exception()

   if __name__ == '__main__':
     test()

    '''
    import sys
    import traceback
    from functools import wraps
    from multiprocessing import Process, Queue

    def process_func(q, *args, **kwargs):
        try:
            ret = func(*args, **kwargs)
        except Exception:
            ex_type, ex_value, tb = sys.exc_info()
            error = ex_type, ex_value, ''.join(traceback.format_tb(tb))
            ret = None
        else:
            error = None

        q.put((ret, error))

    # register original function with different name
    # in sys.modules so it is pickable
    process_func.__name__ = func.__name__ + 'processify_func'
    setattr(sys.modules[__name__], process_func.__name__, process_func)

    @wraps(func)
    def wrapper(*args, **kwargs):
        q = Queue()
        p = Process(target=process_func, args=[q] + list(args), kwargs=kwargs)
        p.start()
        ret, error = q.get()
        #p.join()

        if error:
            ex_type, ex_value, tb_str = error
            message = '%s (in subprocess)\n%s' % (ex_value.message, tb_str)
            raise ex_type(message)

        return ret
    return wrapper


@os_processify_fun
def ztest_processify():
    return os.getpid()




def date_getspecificdate(datelist, datetype1="yearend", outputype1="intdate", includelastdate=True, includefirstdate=False, ):
 vec2= []

 if isint(datelist[0] )  :  datelist= dateint_todatetime(datelist)

 t0 = datelist[0]
 if datetype1== "monthend" :
    for i,t in enumerate(datelist):
#      print(datetime_tostring([t0, t]))
      if t.month != t0.month : vec2.append([i-1,t0]);  # month has change
      t0= t

 if datetype1== "2monthend" :
    for i,t in enumerate(datelist):
      if t.month != t0.month and np.mod(t0.month,2) ==0 : vec2.append([i-1,t0]);  # month has change
      t0= t

 if datetype1== "3monthend" :
    for i,t in enumerate(datelist):
      if t.month != t0.month and np.mod(t0.month,3) ==0 : vec2.append([i-1,t0]);  # month has change
      t0= t

 if datetype1== "4monthend" :
    for i,t in enumerate(datelist):
      if t.month != t0.month and np.mod(t0.month,4) ==0 : vec2.append([i-1,t0]);  # month has change
      t0= t

 if datetype1== "6monthend" :
    for i,t in enumerate(datelist):
      if t.month != t0.month and np.mod(t0.month,6) ==0 : vec2.append([i-1,t0]);  # month has change
      t0= t

 if datetype1== "monthstart" :
    for i,t in enumerate(datelist):
      if t.month != t0.month : vec2.append([i,t]);   # month has change
      t0= t

 if datetype1== "yearstart" :
    vec2.append([0,t0])
    for i,t in enumerate(datelist):
      if t.year != t0.year: vec2.append([i,t]);   # month has change
      t0= t

 if datetype1== "yearend" :
    for i,t in enumerate(datelist):
      if t.year != t0.year : vec2.append([i-1,t0]);   # month has change
      t0= t

 if includelastdate :
      vec2.append([len(datelist)-3, datelist[-1]])

 if outputype1== "intdate" :
   vec2= np.array(vec2)
   vec2= np.array(vec2[:,0], dtype="int")
   return vec2
 else :
   return np.array(vec2)



def py_exception_print():
    import linecache
    exc_type, exc_obj, tb = sys.exc_info()
    f = tb.tb_frame
    lineno = tb.tb_lineno
    filename = f.f_code.co_filename
    linecache.checkcache(filename)
    line = linecache.getline(filename, lineno, f.f_globals)
    print('EXCEPTION IN ({}, LINE {} "{}"): {}'.format(filename, lineno, line.strip(), exc_obj))


def py_log_write(LOGFILE, prefix):
 import arrow, os
    ##########################################################################################################
 #LOGFILE =     DIRCWD + '/aapackage/'+ 'ztest_all.txt';
 print(LOGFILE)
 DATENOW =     arrow.utcnow().to('Japan').format("YYYYMMDDHHmm")
 UNIQUE_ID=    prefix +"_"+ DATENOW +"_"+ str(np.random.randint(10**5, 10**6,  dtype='int64'))
 sys.stdout =  open( LOGFILE, 'a')
 print("\n\n"+UNIQUE_ID+" ###################### Start:" + arrow.utcnow().to('Japan').format()  + "###########################") ; sys.stdout.flush() ; print(os)
 return UNIQUE_ID
 ##########################################################################################################








##############################################################################################
############################ UNIT TEST #######################################################
if __name__ == '__main__' :
 import argparse;  ppa = argparse.ArgumentParser()       # Command Line input
 ppa.add_argument('--do', type=str, default= 'action',  help='test / test02')
 arg = ppa.parse_args()


if __name__ == '__main__' and arg.do == "test":
 print(__file__)
 try:
  import util;  UNIQUE_ID= util.py_log_write( DIRCWD + '/aapackage/ztest_log_all.txt', "util")

  ###########################################################################################
  import numpy as np, pandas as pd, scipy as sci
  import util; print(util); print("")
  #util.a_info_system()
  util.a_isanaconda()
  util.date_allinfo()

  vv  =   np.random.rand(1,10)
  mm  =   np.random.rand(100,5)
  df1  =  pd.DataFrame(mm, columns=["aa", "bb", 'c', 'd', 'e'] )

  # util.pd_createdf(mm, ["aa", "bb", 'c', 'd', 'e'],  )
  print(util.np_sort(vv))

  util.save(df1, "ztest_df")
  df2= util.load("ztest_df")












  ###########################################################################################
  print("\n\n"+ UNIQUE_ID +" ###################### End:" + arrow.utcnow().to('Japan').format() + "###########################") ; sys.stdout.flush()
 except Exception as e : util.py_exception_print()









'''
  try :

  except Exception as e: print(e)


import numpy as np, arrow
UNIQUE_ID=    str(np.random.randint(10**14, 10**15,  dtype='int64'))


print("\n\n###################### Start util # :" + arrow.utcnow().to('Japan').format() + "###########################") ; sys.stdout.flush()
 
 
'''











