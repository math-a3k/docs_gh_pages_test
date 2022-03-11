# coding=utf-8
from __future__ import division
from future import standard_library
standard_library.install_aliases()
from builtins import next
from builtins import map
from builtins import zip
from builtins import str
from builtins import range
from past.builtins import basestring
from past.utils import old_div
from builtins import object
# -*- coding: utf-8 -*-
#---------Various Utilities function for Python--------------------------------------
import os, sys
# if sys.platform.find('win') > -1 :
#  from guidata import qthelpers  #Otherwise Erro with Spyder Save

import datetime, time, arrow,  shutil
import matplotlib.pyplot as plt
import numexpr as ne, numpy as np, pandas as pd, scipy as sci



import util as util










################   SQL UTIL  ###############################################################
'''
from sqlalchemy import create_engine
engine = create_engine("postgresql://u:p@host/database")
'''
import sqlalchemy as sql

def sql_create_dbengine(type1='',  dbname='', login='', password='', url='localhost', port=5432) :
   ''' Return SQL Alchemy Connector

# psycopg2
engine = create_engine('postgresql+psycopg2://scott:tiger@localhost/mydatabase')

# MySQL-connector-python  Official one
engine = create_engine('mysql+mysqlconnector://scott:tiger@localhost/foo')
conda install -c anaconda mysql-connector-python=2.0.4
engine = create_engine('postgresql://%s:%s@localhost:5432/%s' %(myusername, mypassword, mydatabase))

engine = create_engine('sqlite:///  folder/foo.db')

   '''
   if type1=='postgres' :        # psycopg2
      engine = sql.create_engine('postgresql+psycopg2://' + login + ':' + password + '@' + url + ':'+ str(port) + '/'+ dbname)

   elif type1=='mysql' :        # MYSQL
      engine = sql.create_engine('mysql+mysqlconnector://' + login + ':' + password + '@' + url + ':'+ str(port) + '/'+ dbname)

   elif type1=='sqlite' :        # SQLlite
      engine = sql.create_engine('sqlite:/// ' + url+ '/'+ dbname)

   return engine


def sql_query(sqlr='SELECT ticker,shortratio,sector1_id, FROM stockfundamental',  dbengine=None, output='df', dburl='sqlite:///aaserialize/store/finviz.db'):
 '''
 :param sqlr:       'SELECT ticker,shortratio,sector1_id, FROM stockfundamental'
 :param output:     df   /   file1.csv
 :param dburl:      'sqlite:///aaserialize/store/finviz.db'
 :param dbengine:   dbengine = sql.create_engine('postgresql+psycopg2://postgres:postgres@localhost/coke')
 :return:
 '''
 if dbengine is None :   dbengine = sql.create_engine(dburl)
 df= pd.read_sql_query( sqlr, dbengine)

 if output=='df' : return df
 elif output.find('.csv') > -1 :  df.to_csv(output)

def sql_get_dbschema(dburl='sqlite:///aapackage/store/yahoo.db', dbengine=None, isprint=0):
    """function sql_get_dbschema
    Args:
        dburl='sqlite:   
        dbengine:   
        isprint:   
    Returns:
        
    """
    if dbengine is None :
        dbengine = sql.create_engine(dburl)
    inspector = sql.inspect(dbengine)
    l1= []
    for table_name in inspector.get_table_names():
        aux= ""
        for column in inspector.get_columns(table_name):
            l1.append([table_name, column['name']])
            aux= aux + ', ' + column['name']
        if isprint:   print((table_name + ", Col: "+aux +"\n"))
    return np.array(l1)

def sql_delete_table(name, dbengine):
 """function sql_delete_table
 Args:
     name:   
     dbengine:   
 Returns:
     
 """
 pd.io.sql.execute('DROP TABLE IF EXISTS '+name, dbengine)
 pd.io.sql.execute('VACUUM', dbengine)




def sql_insert_excel(file1='.xls', dbengine=None, dbtype='') :
 ''' http://flask-excel.readthedocs.io/en/latest/
 https://pythonhosted.org/pyexcel/tutorial_data_conversion.html#import-excel-sheet-into-a-database-table
 from sqlalchemy import create_engine
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy import Column , Integer, String, Float, Date
from sqlalchemy.orm import sessionmaker
engine = create_engine("sqlite:///birth.db")
Base = declarative_base()
Session = sessionmaker(bind=engine)

class BirthRegister(Base):
...     __tablename__='birth'
...     id=Column(Integer, primary_key=True)
...     name=Column(String)
...     weight=Column(Float)
...     birth=Column(Date)

Base.metadata.create_all(engine)

https://www.digitalocean.com/community/tutorials/how-to-use-celery-with-rabbitmq-to-queue-tasks-on-an-ubuntu-vps
import os
import pyexcel
import datetime

from sqlalchemy import create_engine
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy import Column, Integer, String, Float, Date
from sqlalchemy.orm import sessionmaker


engine = create_engine("sqlite:///birth.db")
Base = declarative_base()
Session = sessionmaker(bind=engine)


# here is the destination table
class BirthRegister(Base):
    __tablename__ = 'birth'
    id = Column(Integer, primary_key=True)
    name = Column(String)
    weight = Column(Float)
    birth = Column(Date)


Base.metadata.create_all(engine)


# create fixture
data = [
    ["name", "weight", "birth"],
    ["Adam", 3.4, datetime.date(2015, 2, 3)],
    ["Smith", 4.2, datetime.date(2014, 11, 12)]
]
pyexcel.save_as(array=data,
                dest_file_name="birth.xls")

# import the xls file
session = Session()  # obtain a sql session
pyexcel.save_as(file_name="birth.xls",
                name_columns_by_row=0,
                dest_session=session,
                dest_table=BirthRegister)

# verify results
sheet = pyexcel.get_sheet(session=session, table=BirthRegister)
print(sheet)

session.close()
os.unlink('birth.db')
os.unlink("birth.xls")

 This code uses the openpyxl package for playing around with excel using Python code
 to convert complete excel workbook (all sheets) to an SQLite database
 The code assumes that the first row of every sheet is the column name
 Every sheet is stored in a separate table
 The sheet name is assigned as the table name for every sheet
 '''
 import sqlite3, re
 from openpyxl import load_workbook

 def slugify(text, lower=1):
    if lower == 1: text = text.strip().lower()
    text = re.sub(r'[^\w _-]+', '', text)
    text = re.sub(r'[- ]+', '_', text)
    return text

 # dbengine= sqlite3.connect(dbname)
 wb = load_workbook(filename=file1)  # r'abc.xlsx'
 sheets = wb.get_sheet_names()

 for sheet in sheets:   #Insert Sheet as new table
    ws = wb[sheet];     columns= []
    query = 'CREATE TABLE ' + str(slugify(sheet)) + '(ID INTEGER PRIMARY KEY AUTOINCREMENT'
    for row in ws.rows[0]:
        query += ', ' + slugify(row.value) + ' TEXT'
        columns.append(slugify(row.value))
    query += ');'
    dbengine.execute(query)


    tup = []
    for i, rows in enumerate(ws):
        tuprow = []
        if i == 0: continue
        for row in rows:
            tuprow.append(str(row.value).strip()) if str(row.value).strip() != 'None' else tuprow.append('')
        tup.append(tuple(tuprow))

    #Insert tuples
    insQuery1 = 'INSERT INTO ' + str(slugify(sheet)) + '('
    insQuery2 = ''
    for col in columns:
        insQuery1 += col + ', '
        insQuery2 += '?, '
    insQuery1 = insQuery1[:-2] + ') VALUES('
    insQuery2 = insQuery2[:-2] + ')'
    insQuery = insQuery1 + insQuery2

    dbengine.executemany(insQuery, tup)
    dbengine.commit()
 dbengine.close()

def sql_insert_df(df, dbtable, dbengine, col_drop=['id'], verbose=1) :
 """function sql_insert_df
 Args:
     df:   
     dbtable:   
     dbengine:   
     col_drop:   
     verbose:   
 Returns:
     
 """
 for c in df.columns:    #Remove columns
   if c in col_drop:  df = df.drop(c, axis=1)

 list1 = df.to_dict(orient='records')  # The orient='records' is the key of this, it allows to align with the format mentioned in the doc to insert in bulks.

 '''
 #list1= list1[0]
 list2=[]
 for i in xrange(0, len(list1)) :
    listx= list1[i]     
    for col in col_drop :
       del listx[col] #remove the ID, so SQL can create it, as well timeStamp
    list2.append(listx)
 '''

 if verbose : print(list1)

 metadata = sql.schema.MetaData(bind=dbengine,reflect=True)
 table = sql.Table(dbtable, metadata, autoload=True)

 Session = sql.orm.session.sessionmaker(bind=dbengine)
 session = Session()

 # Insert the dataframe into the database in one bulk
 res= dbengine.execute(table.insert(), list1)

 session.commit(); session.close()
 return res

def sql_insert_csv(csvfile, dbtable, dbengine, col_drop=[]) :
 """function sql_insert_csv
 Args:
     csvfile:   
     dbtable:   
     dbengine:   
     col_drop:   
 Returns:
     
 """
 start = datetime.now()
 chunksize = 20000
 j = 0
 index_start = 1

 for df in pd.read_csv(csvfile, chunksize=chunksize, iterator=True, encoding='utf-8'):

   df.index += index_start
   df = df.rename(columns={c: c.replace(' ', '') for c in df.columns}) # Remove spaces from columns
   #df['CreatedDate'] = pd.to_datetime(df['CreatedDate']) # Convert to datetimes
   #df['ClosedDate'] =  pd.to_datetime(df['ClosedDate'])
   for c in df.columns:    #Remove columns
      if c in col_drop:  df = df.drop(c, axis=1)


   ################  df to SQL Alchemy ######################################
   listToWrite = df.to_dict(orient='records')  # The orient='records' is the key of this, it allows to align with the format mentioned in the doc to insert in bulks.

   metadata = sql.schema.MetaData(bind=dbengine,reflect=True)
   table =    sql.Table(dbtable, metadata, autoload=True)
   Session = sql.orm.session.sessionmaker(bind=dbengine)
   session = Session()

   # Inser the dataframe into the database in one bulk
   dbengine.execute(table.insert(), listToWrite)
   session.commit();   session.close()

   #     df.to_sql('data', disk_engine, if_exists='append')
   j+=1
   index_start = df.index[-1] + 1
   print('{} seconds: completed {} rows'.format((datetime.now() - start).seconds, j*chunksize))




 ''' Batch Mode :
 #  CSV to SQL Lite ----------------------------------------------------------------------
disk_engine = create_engine('sqlite:///311_8M.db') # Initializes database with filename 311_8M.db in current directory

start = dt.datetime.now()
chunksize = 20000
j = 0
index_start = 1

for df in pd.read_csv('311_100M.csv', chunksize=chunksize, iterator=True, encoding='utf-8'):

    df = df.rename(columns={c: c.replace(' ', '') for c in df.columns}) # Remove spaces from columns
    df['CreatedDate'] = pd.to_datetime(df['CreatedDate']) # Convert to datetimes
    df['ClosedDate'] = pd.to_datetime(df['ClosedDate'])

    df.index += index_start

    # Remove the un-interesting columns
    columns = ['Agency', 'CreatedDate', 'ClosedDate', 'ComplaintType', 'Descriptor',
               'CreatedDate', 'ClosedDate', 'TimeToCompletion','City']

    for c in df.columns:
        if c not in columns:
            df = df.drop(c, axis=1)

    j+=1
    print '{} seconds: completed {} rows'.format((dt.datetime.now() - start).seconds, j*chunksize)

    df.to_sql('data', disk_engine, if_exists='append')
    index_start = df.index[-1] + 1




 '''

def sql_insert_csv2(csvfile='', dbtable='', columns=[], dbengine=None, nrows= 10000):
    """
    Upload data to a temporary table first using PANDAs to identify optimal data-types for columns
    PANDAS is not speed-efficient as it uses INSERT commands rather than COPY e.g. it took COPY 16mins average
    to get a 15GB CSV into the database (door-to-door) whereas pandas.to_sql took 50mins
    """

    dbtable +='_temp'
    counter = 0
    for i in os.listdir(csvfile):
        # Cycle through all CSVs and upload a small chunk to make sure everything is OK
        if counter < 1:
            if i.endswith(".csv") & i.startswith("..."):
                print(("Reading CSV: %s into PANDAs data-frame" % i))

                # First 1,000,000 rows
                #df = pd.read_csv(os.path.join(csvfile, i), nrows=1000000, header=None, sep='~') #sep=None; automatically find by sniffing

                # Upload whole file
                df = pd.read_csv(os.path.join(csvfile, i), nrows= nrows, header=None, sep='~') #sep=None; automatically find by sniffing
                # My dates were in columns 2 and 3
                # The column names were not present in the original CSVs
                df.columns =  columns
                print("CSV read-in successfully")
                print(df.shape)
                print(("Uploading %s to SQL Table: %s" % (i, dbtable)))
                df.to_sql(dbtable, engine, if_exists='append', index=False)
                counter += 1
                print(("Successfully uploaded: %d" % counter))




def sql_postgres_create_table(mytable='', database='', username='', password='' ):
    """ Create table copying the structure of the temp table created using pandas  Timer to benchmark """
    # Connect
    import psycopg2
    con = psycopg2.connect(database=database, user=username, password=password)
    cur = con.cursor()
    if con:
        print(('Connected: %s' % con))
    else:
        print('Connection lost')
        sys.exit(1)

    try:
        # Check if table exists already
        cur.execute("""
                    SELECT relname FROM pg_class WHERE relname = '{0}';
                    """.format(my_table))
        table_test = cur.fetchone()[0]
    except Exception as e:
        print(('Table %s does not exist' % mytable))
        table_test = None

    if table_test:
        print(('%s already exists' % mytable))
    else:
        print(('Creating table: %s' % mytable))
        try:
            # Copy structure and no data (1=2 is false)
            cur.execute("""
                        CREATE TABLE {0} AS SELECT * FROM {1} WHERE 1=2;
                        """.format(my_table, my_table+'_temp'))
            con.commit()
            print('Table created successfully')
        except psycopg2.DatabaseError as e:
            if con:
                con.rollback()
            print(('Error %s' % e))
            sys.exit(1)
    con.close()

def sql_postgres_insert_csv(path_2_csv='', my_table=''):
    """  Use the PostgreSQL COPY command to bulk-copy the CSVs into the newly created table """
    # Connect
    con = psycopg2.connect(database=mydatabase, user=myusername, password=mypassword)
    cur = con.cursor()
    if con:
        print(('Connected: %s' % con))
    else:
        print('Connection lost')
        sys.exit(1)

    copy_sql = """
               COPY %s FROM stdin DELIMITERS '~' CSV;
               """ % my_table
    counter = 0
    start_time = time.time()

    for i in os.listdir(path_2_csv):
        if i.endswith(".csv") and i.startswith("..."):
            print(("Uploading %s to %s" % (i, mytable)))
            with open(os.path.join(path_2_csv, i), 'r') as f:
                cur.copy_expert(sql=copy_sql, file=f)
                con.commit()
                counter += 1
                print(("Successfully uploaded %d CSVs" % counter))
                current_speed = ((time.time()-start_time)/60)/counter
                print(("Average speed is %.2f minutes per database" % current_speed))
    con.close()
    end_time = time.time()
    print(("Total duration of COPY: %.2f minutes" % ((end_time - start_time)/60)))

def sql_postgres_query_to_csv(sqlr='SELECT ticker,shortratio,sector1_id, FROM stockfundamental', csv_out=''):
    """ Submit query to created PostgreSQL database and output results to a CSV  """
    import psycopg2
    con = psycopg2.connect(database=mydatabase, user=myusername, password=mypassword)
    cur = con.cursor()
    if con:   print(('Connected: %s' % con))
    else:
       print('Connection lost'); return  -1

    output_query = "COPY ({0}) TO STDOUT WITH CSV HEADER".format(sqlr)
    with open(csv_out, 'w') as f:
        cur.copy_expert(output_query, f)
        print(("Successfully submitted results to: %s" % csv_out))
    con.close()

def sql_postgres_pivot():
   '''
Enabling the Crosstab Function
As we previously mentioned, the crosstab function is part of a PostgreSQL extension called tablefunc. To call the crosstab function,
you must first enable the tablefunc extension by executing the following SQL command:
CREATE extension tablefunc;

SELECT *
FROM crosstab( 'select student, subject, evaluation_result from evaluations order by 1,2')
AS final_result(Student TEXT, Geography NUMERIC,History NUMERIC,Language NUMERIC,Maths NUMERIC,Music NUMERIC);


##### Correct Even iF there are missing values :
http://www.vertabelo.com/blog/technical-articles/creating-pivot-tables-in-postgresql-using-the-crosstab-function

SELECT *
FROM crosstab( 'select student, subject, evaluation_result from evaluations
                where extract (month from evaluation_day) = 7 order by 1,2',
                'select name from subject order by 1')
     AS final_result(Student TEXT, Geography NUMERIC,History NUMERIC,Language NUMERIC,Maths NUMERIC,Music NUMERIC);


   '''
   pass


def sql_mysql_insert_excel():
 """function sql_mysql_insert_excel
 Args:
 Returns:
     
 """
 import MySQLdb, xlrd

 list= xlrd.open_workbook("prod.xls")
 sheet= list.sheet_by_index(0)

 database = MySQLdb.connect (host="localhost" , user="root" , passwd="" ,db="table")
 cursor = database.cursor()

 query= """INSERT INTO produits (idProduit, idCategorie, LibelleProduit, PrixProduit) VALUES (%s, %s, %s, %s)"""
 for r in range(1,sheet.nrows):
    idProduit = sheet.cell(r,0).value
    categorie = 999
    libelle=sheet.cell(r,1).value
    prix=sheet.cell(r,3).value #>>HERE THE PROBLEM the Imported Value <<<<

    values = (idProduit,categorie,libelle,prix)

    cursor.execute(query,values)

 cursor.close();
 database.commit()

 database.close()

 print("")
 print("All done !")

 columns= str(sheet.ncols)
 rows=str(sheet.nrows)
 print("i just import "+columns+" columns and " +rows+ " rows to MySQL DB")


def sql_pivotable(dbcon, ss='select  '):
  '''

 1) get the category

 2) Build the Pivot From category
  SELECT *
  FROM crosstab( 'select student, subject, evaluation_result from evaluations
                where extract (month from evaluation_day) = 7 order by 1,2',
                'select name from subject order by 1')
   AS final_result(Student TEXT, Geography NUMERIC,History NUMERIC,Language NUMERIC,Maths NUMERIC,Music NUMERIC);

  https://www.amazon.com/PostgreSQL-High-Performance-Gregory-Smith/dp/184951030X/ref=as_li_ss_tl?s=books&ie=UTF8&qid=1458352081&sr=1-6&keywords=postgres&linkCode=sl1&tag=postgres-bottom-20&linkId=c981783121cbd5542dc2b44a2297df57


http://blog.brakmic.com/data-science-for-losers-part-2/

Here we instruct Pandas to merge two tables by using certain primary keys from both when combining their rows into a new table. The parameter how instructs Pandas to use the inner-join which means it will only combine such rows which belong to both of the tables. Therefore we’ll not receive any NaN-rows. But in some cases this could be desirable. Then use the alternative options like left, right or outer.

Pivots with Tables from SQLAlchemy

And of course it’s possible to generate the same pivot tables with data that came from SQLAlchemy.
They’re nothing else but DataFrames all the way down. OK, not absolutely all the way down,
because there are also Series and NumPy arrays etc.,
but this is a little bit too much of knowledge for Losers like us. Maybe in some later articles.


 :return:
  '''

  ss= ''
  pass



