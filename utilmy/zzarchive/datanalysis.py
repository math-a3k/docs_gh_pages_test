# -*- coding: utf-8 -*-
'''    Data Analysis Utilities   '''
import sys, os, numpy as np, pandas as pd, copy, scipy as sci, matplotlib.pyplot as plt, math
import requests, re, arrow
from tabulate import tabulate
from datetime import datetime, timedelta
from calendar import isleap
from dateutil.parser import parse

from collections import OrderedDict
from attrdict import AttrDict as dict2


import statsmodels as sm

import kmodes
from kmodes.kmodes import KModes
import pylab as pl, itertools, sklearn as sk

from sklearn.decomposition import PCA
from sklearn import linear_model, covariance, model_selection
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import StandardScaler
from sklearn.neighbors import KNeighborsClassifier
from sklearn.svm import SVC
from sklearn.tree import DecisionTreeClassifier
from sklearn.naive_bayes import GaussianNB
from sklearn.discriminant_analysis import LinearDiscriminantAnalysis, QuadraticDiscriminantAnalysis
from sklearn.ensemble import RandomForestClassifier, AdaBoostClassifier,GradientBoostingClassifier, ExtraTreesClassifier
from sklearn.metrics import confusion_matrix
from sklearn.cluster import k_means, dbscan
from sklearn.decomposition import pca


###########################################################################################################################
# import os, sys
# CFG   = {'plat': sys.platform[:3]+"-"+os.path.expanduser('~').split("\\")[-1].split("/")[-1], "ver": sys.version_info.major}
# DIRCWD= {'win-asus1': 'D:/_devs/Python01/project27/', 'win-unerry': 'G:/_devs/project27/' , 'lin-noel': '/home/noel/project27/', 'lin-ubuntu': '/home/ubuntu/project27/' }[CFG['plat']]
# os.chdir(DIRCWD); sys.path.append(DIRCWD + '/aapackage')
# DIRCWD= os.environ["DIRCWD"]; os.chdir(DIRCWD); sys.path.append(DIRCWD + '/aapackage')

import configmy; CFG, DIRCWD= configmy.get(config_file="_ROOT", output= ["_CFG", "DIRCWD"])
os.chdir(DIRCWD); sys.path.append(DIRCWD + '/aapackage')



import util, datanalysis as da



__path__= DIRCWD +'/aapackage/'
__version__= "1.0.0"
__file__=    "datanalysis.py"


############# CONSTANT   ##################################################################################################
#type of data
"""
df : dataframe
Xmat :  Numpy Matrix values
Ytarget : Value to be predicted, Class
col: column,    row: row

"""







############# Pandas Processing   ######################################################################
def pd_filter_column(df_client_product, filter_val=[], iscol=1):
   '''
   # Remove Columns where Index Value is not in the filter_value
   # filter1= X_client['client_id'].values
   :param df_client_product:
   :param filter_val:
   :param iscol:
   :return:
   '''
   axis=1 if iscol==1 else 0
   col_delete1=[]
   for colname in df_client_product.index.values:  # !!!! row Delete
      if colname in filter_val :  col_delete1.append(colname)

   df2=df_client_product.drop(col_delete1, axis=axis, inplace=False)
   return df2

def pd_missing_show():
   pass

   '''
   https://blog.modeanalytics.com/python-data-visualization-libraries/


   Missing Data

     missingno
import missingno as msno
%matplotlib inline
msno.matrix(collisions.sample(250))
At a glance, date, time, the distribution of injuries, and the contribution factor of the first vehicle appear to be completely populated, while geographic information seems mostly complete, but spottier.

The sparkline at right summarizes the general shape of the data completeness and points out the maximum and minimum rows.

This visualization will comfortably accommodate up to 50 labelled variables. Past that range labels begin to overlap or become unreadable, and by default large displays omit them.


Heatmap
The missingno correlation heatmap lets you measure how strongly the presence of one variable positively or negatively affect the presence of another:
msno.heatmap(collisions)


https://github.com/ResidentMario/missingno

   '''

def pd_describe( df ):
   ''' Describe the tables
        
       
   '''
   coldes= [ "col", "coltype", "dtype", "count", "min", "max", "nb_na", "pct_na", "median",  "mean", "std",  "25%",  "75%",  "outlier"   ]


   def getstat(col, type1="num") :
      '''
         max, min, nb, nb_na, pct_na, median, qt_25, qt_75,
         nb, nb_unique, nb_na, freq_1st, freq_2th, freq_3th
         s.describe()
         count    3.0  mean     2.0 std      1.0
         min      1.0   25%      1.5  50%      2.0
         75%      2.5  max      3.0   
      '''
      ss=     list(df[col].describe().values)
      ss=     [ str(df[col].dtype) ] + ss
      nb_na=  df[col].isnull().sum()
      ntot=   len(df)
      ss=     ss + [nb_na, nb_na / ( ntot + 0.0) ]

      return  pd.Series(ss, [ "dtype", "count", "mean", "std", "min", "25%", "50%", "75%", "max", "nb_na", "pct_na"     ])

   dfdes=  pd.DataFrame( [], columns= coldes)
   cols=   df.columns
   for col in cols :
      dtype1 = str(df[col].dtype)
      if dtype1[0:3] in ["int", "flo"] :
         row1=   getstat(col, "num")
         dfdes=  pd.concat((dfdes, row1))

      if dtype1== "object" :
         pass




def pd_stack_dflist(df_list):
   df0= None
   for i, dfi in enumerate(df_list) :
      if df0 is None : df0= dfi
      else    :
         try:     df0= df0.append(dfi, ignore_index=True)
         except : print(('Error appending: '+str(i) ))
   return df0


def pd_validation_struct():
  pass
  '''
  https://github.com/jnmclarty/validada

  https://github.com/ResidentMario/checkpoints


  '''

def pd_checkpoint() :
  pass
'''
  Create Checkpoint on dataframe to save intermediate results
  https://github.com/ResidentMario/checkpoints
  To start, import checkpoints and enable it:

from checkpoints import checkpoints
checkpoints.enable()
This will augment your environment with pandas.Series.safe_map and pandas.DataFrame.safe_apply methods. Now suppose we create a Series of floats, except for one invalid entry smack in the middle:

import pandas as pd; import numpy as np
rand = pd.Series(np.random.random(100))
rand[50] = "____"
Suppose we want to remean this data. If we apply a naive map:

rand.map(lambda v: v - 0.5)

    TypeError: unsupported operand type(s) for -: 'str' and 'float'
Not only are the results up to that point lost, but we're also not actually told where the failure occurs! Using safe_map instead:

rand.safe_map(lambda v: v - 0.5)

    <ROOT>/checkpoint/checkpoints/checkpoints.py:96: UserWarning: Failure on index 50
    TypeError: unsupported operand type(s) for -: 'str' and 'float'


'''


'''
You can control how many decimal points of precision to display
In [11]:
pd.set_option('precision',2)

pd.set_option('float_format', '{:.2f}'.format)


Qtopian has a useful plugin called qgrid - https://github.com/quantopian/qgrid
Import it and install it.
In [19]:
import qgrid
qgrid.nbinstall()
Showing the data is straighforward.
In [22]:
qgrid.show_grid(SALES, remote_js=True)


SALES.groupby('name')['quantity'].sum().plot(kind="bar")


'''



############## Excel processing #######################################################################
######## Read file and extract data pattern:  .csv,  .txt, .xls  ##################################


def xl_setstyle(file1) :
 '''
   http://openpyxl.readthedocs.io/en/default/styles.html#cell-styles-and-named-styles
  import openpyxl.styles.builtins  as bi
  import openpyxl.styles.builtins

  col = ws.column_dimensions['A']
  col.font = Font(bold=True)

  for cell in ws['A'] + ws[1]:
    cell.style = 'data01'

  bd = Side(style='thick', color="000000")
  highlight.border = BORDER_NONE
  from openpyxl.styles import
 '''
 import openpyxl as xl
 try:
  from openpyxl.cell import get_column_letter as gcol
  from openpyxl.cell import column_index_from_string as gstr
 except :
    from openpyxl.utils import get_column_letter as gcol
    from openpyxl.utils import column_index_from_string as gstr
 from openpyxl import load_workbook

 try:      wb =      load_workbook(file1)
 except :  print(('File protected: '+file1))

 from openpyxl.styles import named_styles, Font, Border, Side
 data01 =      named_styles.NamedStyle(name="data01")
 data01.font = Font(name='Arial', bold=False, size=8, italic=False, color="000000")

 ws_names= wb.get_sheet_names()
 for wsname in ws_names :
   ws= wb.get_sheet_by_name(wsname)
   #cr= ws.get_squared_range('A1')
   for coli in range(1, 20) :

    #  issues, does not work
    col = ws.column_dimensions[gcol(coli)]
    col.font= Font(name='Arial', bold=False, size=8, italic=False, color="000000")
 wb.save(file1)



def xl_val(ws, colj, rowi) :
   from openpyxl.utils import get_column_letter as gcol
   try:
      return ws[gcol(colj) + str(rowi)].value
   except :
      return None

def isnull(x) :
 return x is None or x==''

def xl_get_rowcol(ws, i0, j0, imax, jmax) :
   rmat= []
   for j in range(j0, j0+jmax) :
     if isnull(xl_val(ws, j, i0)) and isnull(xl_val(ws, j+1, i0))  : #Stop conditions : column j+1 is empty
       return rmat
     rmatj=  [xl_val(ws, j, i)  for i in range(i0, i0+imax)  ]   # add all the rows
     rmat.append(rmatj)
   return rmat


def xl_getschema(dirxl="", filepattern='*.xlsx', dirlevel=1, outfile='.xlsx') :
 '''Take All excel in a folder and provide Table, Column Schema, type into master file
 '''

 def xl_is_data_block_start(ws, i, colj) :
  #Check if block of data start:  i,j+1,j+2   i+1 i+2 should be NO empty
  if isnull(xl_val(ws, colj, i))  : return False
  if (not isnull(xl_val(ws, colj, i))     and not isnull(xl_val(ws, colj+2, i)) and
      not isnull(xl_val(ws, colj+1, i+1)) and not isnull(xl_val(ws, colj+2, i+2))) :
    return True

 def xl_find_start_block(ws, colmin, rowmin, colmax, rowmax):
       #locate 1st non blank cell
       for j in range(colmin, colmax) :
         for i in range(rowmin, rowmax) :
          if xl_is_data_block_start(ws, i, j) : return i,j
       return rowmin, colmin


 def _xl_getschema(file1="*.xlsx",  output_istack=1) :
  ''' 1 file Get dataframe Schema: Table, Colnam, val1, val2 from Excel xlsx
  '''
  from openpyxl import load_workbook
  from openpyxl.utils import get_column_letter as gcol
  #filepath1, filen1=  util.os_split_dir_file(file1)
  filepath1= file1
  filen1=    util.os_file_getname(file1)
  filetype1= filen1[-4:]
  try:      wb= load_workbook(file1, read_only=True)
  except :  print(('File protected: '+file1));     return None

  print(file1)
  ws_names= wb.get_sheet_names()
  df_list=  []
  for wsname in ws_names :
    ws= wb.get_sheet_by_name(wsname)
    if ws.max_column > 2   and ws.max_row > 3 :   #Minimum table size

     print('    Reading: ',  ws.title, ws.max_row, ws.max_column)
     i0,j0= ws.min_row, ws.min_column
     if not isnull(xl_val(ws, i0+1, j0+1)) :
       i1, j1=    i0 + 2, ws.max_column
       lefttop, rightdown=   gcol(j0) + str(i0), gcol(j1) + str(i1)
       xmat =     np.array([[i.value for i in j] for j in ws[lefttop:rightdown]]).T

     else :
       #locate 1st non blank cell
       max_col, max_row= 50, 100
       i,j= xl_find_start_block(ws, ws.min_column, ws.min_row, max_col, max_row)
       xmat= np.array(xl_get_rowcol(ws, i, j, imax=3, jmax= 1000))

     cols=  ['uri', 'filetype', 'file', 'table', 'column', 'val1', 'val2']
     nlen=  xmat.shape[0]
     xfile= np.array( [[ filepath1, filetype1, filen1, str(wsname) ]]  * nlen)
     datai= np.column_stack( (xfile, xmat))

     df_list.append(pd.DataFrame(datai, columns= cols))

  if output_istack: return pd_stack_dflist(df_list)
  else :            return df_list


 filelist1= util.os_file_listall(dirxl, filepattern, dirlevel=dirlevel)
 df0= None
 for i, file1 in enumerate(filelist1[2]) :
   dfi= _xl_getschema(file1=file1,  output_istack=1)
   if df0 is None : df0= dfi
   else    :
      try :     df0= df0.append(dfi, ignore_index=True)
      except:   print(('Error on: ' + file1 ))

 if df0 is not None : util.pd_toexcel(df0, outfile=outfile, sheet_name='schema', returnfile=1)






######################  Pre Processing  ###############################################################
def str_to_unicode(x, encoding='utf-8'):
  if not isinstance(x, str):
      x = str(x, encoding)
      return x
  else :
    return x


def csv_dtypes_getdict(df=None, csvfile=None) :
   if df is not None :   return df.dtypes.to_dict()
   elif csvfile is not None :
      df= pd.read_csv(csvfile, nrows= 1000)
      return df.dtypes.to_dict()


def csv_fast_processing() :
   '''
   http://word.bitly.com/post/74069870671/optimizing-text-processing

import sys
from collections import defaultdict
OUT_FILES = defaultdict(dict)

open_outfiles()  # open all files I could possibly need

for line in sys.stdin:
    # 1. parse line for account_id and metric_type
    key = line.split(',')
    account_id = key[ACCOUNT_ID_INDEX][1:] # strip leading quote

    # 2. write to appropriate file for account_id and metric_type
    OUT_FILES[account_id][key[METRIC_TYPE_INDEX]].write(line)

   close_outfiles()  # close all the files we opened

   '''

def csv_col_schema_toexcel(dircsv="", filepattern='*.csv', outfile='.xlsx', returntable=1 ,maxrow=5000000, maxcol_pertable=90, maxstrlen='U80') :
 '''Take All csv in a folder and provide Table, Column Schema, type
 str(df[col].dtype)  USE str always, otherwise BIG Issue

METHOD FOR Unicode / ASCII issue
1. Decode early:  Decode to <type 'unicode'> ASAP
    df['PREF_NAME']=       df['PREF_NAME'].apply(to_unicode)
2. Unicode everywhere
3. Encode late :f = open('/tmp/ivan_out.txt','w')
                f.write(ivan_uni.encode('utf-8'))
 '''

 import gc
 filelist1= util.os_file_listall(dircsv, filepattern, dirlevel=0)
 csvlist= filelist1[0]  #name
 nsheet=  len( csvlist )
 nrows_schema= maxcol_pertable * nsheet

 type_guess_all= {}  #Dict of all new types
 schema= np.array(util.str_empty_string_array(15, nrows_schema), dtype=maxstrlen)
 jj=0
 for ii, tablei in enumerate(csvlist) :
   print((ii,filelist1[2][ii]))
   df= pd.read_csv(dircsv+'/'+tablei+'.csv', nrows= maxrow )  #, encoding='UTF-8')

   df_type1= df.dtypes.tolist()
   dfcol= df.columns.values
   #Transform to unicode EARLY
   for i, col  in enumerate(dfcol) :
      if  str(df_type1[i])== 'object' :
         df[col].fillna('', inplace=True)
         df[col]= df[col].apply(str_to_unicode)

   util.pd_toexcel(df[:100],outfile,  sheet_name= tablei)
   ncol= len(df.columns.values)

   #Stats on the dataframe
   df_uniquecount= df.apply(pd.Series.nunique).values
   df_count=       len(df.index) + np.zeros(ncol, dtype=np.int32)
   df_max=         df.apply(pd.Series.max).values
   df_min=         df.apply(pd.Series.min).values

   df_type2=       [ str(type(df.iloc[0,i])) for i in range(0, ncol) ]
   df_type_len=    [ df[col].map(len).max()  if str(df[col].dtype) == 'object' else 0  for col in df.columns.values ]
   #df_quantile= df.quantile(0.9, axis=1)

   #Type guessing
   df_type_guess_dict= df.dtypes.to_dict()
   df_type_guess= [''] * ncol
   for i, x in enumerate(df_type1) :
      df_type_guess[i]= df_type1[i]
      if str(x)== 'object' :
         ratio= float(df_uniquecount[i]) / df_count[i]
         if (df_uniquecount[i] < 1000 and ratio < 0.2) or ratio  < 0.01 :  #Less than 1% of all
             df_type_guess[i] = 'category'
      if str(x).find('int') > -1 :
         if   df_max[i] < 10000:       df_type_guess[i] = 'int16'
         elif df_max[i] < 4294967296: df_type_guess[i] = 'int32'
      if str(x).find('float') > -1 :
         if    df_max[i] < 10000.0 :      df_type_guess[i] = 'float16'
         elif  df_max[i] < 3.4028235e+38: df_type_guess[i] = 'float32'

      df_type_guess_dict[dfcol[i]]= str(df_type_guess[i]) #.replace('object', 'unicode')

   type_guess_all[tablei]= df_type_guess_dict
   #Schema: Table, column name, val into NUMPY array
   schema[jj:jj + ncol, 0]= tablei                # Table name
   schema[jj:jj + ncol, 1]= df.columns.values     # Col name
   schema[jj:jj + ncol, 2]= df.iloc[0,:].values   # Col 1st value
   schema[jj:jj + ncol, 3]= df_type1              # type Pandas
   schema[jj:jj + ncol, 4]= df_type2              # type numpy
   schema[jj:jj + ncol, 5]= df_type_guess         # Guess type (reduction memory)
   schema[jj:jj + ncol, 6]= df_type_len           # String length
   schema[jj:jj + ncol, 7]= df_uniquecount
   schema[jj:jj + ncol, 8]= df_count
   schema[jj:jj + ncol, 9]= df_max
   schema[jj:jj + ncol, 10]= df_min
   #schema[jj:jj + ncol, 11]= df_quantile
   schema[jj:jj + ncol, 12]= filelist1[2][ii]
   jj=jj + ncol + 1
   gc.collect()

 schema= pd.DataFrame(schema,
 columns=['table', 'column', 'val', 'df_type', 'np_type', 'guess_type', 'type_len',  'nb_unique', 'nb', 'max',
          'min',  'quantile_90', 'uri', 'col14', 'col15'])
 util.pd_toexcel(schema, outfile, sheet_name='schema')
 util.save(type_guess_all, outfile.replace('.xlsx','') + '_type_guess.pkl' )
 util.save(schema, outfile.replace('.xlsx','') + '_schema.pkl' )
 print(outfile); print(('_type_guess.pkl', '_schema.pkl'))
 if returntable : return schema, type_guess_all


def csv_col_get_dict_categoryfreq(dircsv, filepattern="*.csv", category_cols=[], maxline=-1, fileencoding="utf-8"):
  ''' Find Category Freq in large CSV Transaction Column   '''
  start = datetime.now()
  from collections import defaultdict   #Ordersed dict
  catdict={}               #List of Category to fill when parsing
  for colid in category_cols :
     catdict[colid] = defaultdict(int)

  csvlist=  util.os_file_listall(dircsv, filepattern)
  csvlist=  csvlist[:,2]
  kk=0
  for jj, filecsv in enumerate(csvlist) :
    maxline= 1000000000 if maxline==-1 else maxline
    for i, line in enumerate(open(filecsv, encoding=fileencoding)):
     if i > 0:
        kk+=1
        ll=  line.split(",")
        for colid in category_cols :
           catdict[colid][ ll[colid] ]+= 1

        if kk % 10000000 == 0:  print(i, datetime.now() - start)       #progress
     if kk > maxline : break

  return catdict


def csv_row_reduce_line(fromfile, tofile, condfilter, catval_tokeep, header=True, maxline=-1):
  ''' Reduce Data Row by filtering on some Category
    file_category=  in1+ "offers.csv"
    ncol= 8
    catval_tokeep=[ {} for i in xrange(0, ncol)]
    for i, line in enumerate(open(file_category)):
      ll=  line.split(",")
      catval_tokeep[3][  ll[1] ]  = 1  # Offer_file_col1 --> Transact_file_col_4
      catval_tokeep[4][  ll[3] ] =  1  # Offer_file_col3 --> Transact_file_col_4

  def condfilter(colk, catval_tokeep) :
    if colk[3] in catval_tokeep[3] or colk[4] in catval_tokeep[4]: return True
    else: return False
  '''
  start = datetime.now()
  maxline= 1000000000 if maxline==-1 else maxline

  with open(tofile, "wb") as outfile:
    jj_new = 0; i=0
    with open(fromfile) as f :
     if header :                        #1st line
        line= next(f);  outfile.write( ll)
     for i, line in enumerate(f):        #go through transactions file and reduce
        ll=  line.split(",")
        if condfilter(ll, catval_tokeep) :    #Condition  Filter : if category in offers dict
          outfile.write( ",".join(line) )
          jj_new += 1

        #progress
        if i % 5000000 == 0:  print(i, jj_new, datetime.now() - start)
        if i > maxline : break
  print(i, jj_new, datetime.now() - start)
  '''
  does not work, issue with character encoding....
      with open(fromfile, 'r') as f :
     with csv.reader(f,  delimiter=',' ) as reader :
      for ll in reader:
  '''


def csv_analysis() :
   '''
   https://csvkit.readthedocs.io/en/540/tutorial/1_getting_started.html

   sudo pip install csvkit

   :return:
   '''



def csv_row_reduce_line_manual(file_category, file_transact, file_reduced):
  ''' Reduce Data by filtering on some Category '''
  start = datetime.now()

  #Parse all categories and comps on offer in a dict
  offers_cat, offers_co = {}, {}
  for i, line in enumerate(open(file_category)):
    ll=  line.split(",")
    offers_cat[ ll[1] ] = 1
    offers_co[  ll[3] ] = 1


  #open output file
  with open(file_reduced, "wb") as outfile:
    #go through transactions file and reduce
    jj_new = 0
    for i, line in enumerate(open(file_transact)):
      if i == 0: outfile.write(line) #print header
      else:
        ll=  line.split(",")
        if ll[3] in offers_cat or ll[4] in offers_co:    #Condition  Filter : if category in offers dict
          outfile.write( line )
          jj_new += 1

      #progress
      if i % 5000000 == 0:  print(i, jj_new, datetime.now() - start)
  print(i, jj_new, datetime.now() - start)



def csv_row_mapreduce(dircsv="", outfile="", type_mapreduce='sum', nrow=1000000, chunk= 5000000) :
 '''Take All csv in a folder and provide Table, Column Schema'''
 filelist1= util.os_file_listall(dircsv,'*.csv')
 csvlist= filelist1[:,0]

 colname=""
 kchunk= int(nrow / chunk)

 dfout= pd.DataFrame([], columns=colname)
 for ii, tablei in enumerate(csvlist) :
   for kk in range(0, kchunk) :
    df= pd.read_csv(dircsv+'/'+tablei+'.csv', nrows=chunk, encoding='UTF-8')

 return dfout



def csv_pivotable(dircsv="", filepattern='*.csv', fileh5='.h5', leftX='col0', topY='col2', centerZ='coli', mapreduce='sum', chunksize= 500000, tablename='df'):
 ''' return df Pivot Table from series of csv file (transfer to d5 temporary)

Edit: you can groupby/sum from the store iteratively since this "map-reduces" over the chunks:

reduce(lambda x, y: x.add(y, fill_value=0),
       (df.groupby().sum() for df in store.select('df', chunksize=50000)))

 '''

 if dircsv != "" :
   util.pd_h5_fromcsv_tohdfs(dircsv, filepattern=filepattern, tofilehdfs=fileh5, tablename=tablename, chunksize= chunksize)

 store = pd.HDFStore(fileh5)
 if mapreduce== 'sum' :
   pivot0= None
   for df in store.select(tablename, chunksize=chunksize) :
      if pivot0 is None :
         pivot0= pd.DataFrame.pivot_table(df, values=centerZ, index=[leftX],  columns=[topY], aggfunc= np.sum,      fill_value=0)
      else :
         pivot_i= pd.DataFrame.pivot_table(df, values=centerZ, index=[leftX],  columns=[topY], aggfunc= np.sum,      fill_value=0)
         pivot0=  pd.concat([pivot0, pivot_i]).groupby(level=0).sum()

 if mapreduce== 'count' :
   pass

 return pivot0



def csv_bigcompute():
   pass




######################## DB related items #######################################################################
def db_getdata():
   pass

def db_sql():
   pass

def db_meta_add(metadb, dbname, new_table=('', []), schema=None, df_table_uri=None, df_table_columns=None) :
   ''' Create Meta database to store infos on the tables : csv, zip, HFS, Postgres
ALL_DB['japancoupon']= {}
ALL_DB['japancoupon']['schema']=    df_schema
ALL_DB['japancoupon']['df_table_uri']= df_schema_dictionnary
ALL_DB['japancoupon']['df_table_columns']= df_schema_dict
        DBname, db_schema, db_table_uri, db_table_columns(dict_table->colum_list),
   '''
   def pd_df_todict(df, colkey='table', firstelt= True) :
      df1= df.drop_duplicates(colkey).reset_index(level=0, drop=True)
      dict0 = {}
      for i in range(len(df)):
         id0 = df.iloc[i,0]
         val0= df.iloc[i,1]
         if id0 != '' :
           dict0.setdefault(id0, [])
           if firstelt :  dict0[id0]= val0
           else:          dict0[id0].append(val0)
      return dict0

   if schema is not None :
     metadb[dbname]['schema']= schema
     metadb[dbname]['table_columns']= pd_df_todict(schema[ ['table', 'column'] ])
     metadb[dbname]['table_uri']=     pd_df_todict(schema[ ['table', 'uri'] ])

   elif df_table_uri is not None :
     metadb[dbname]['table_uri']= pd_df_todict( df_table_uri )

   elif df_table_columns is not None :
     metadb[dbname]['table_columns']= pd_df_todict( df_table_columns, onlyfirstelt=False )  # table, colname, dtype

   if new_table[0] !='' :
      metadb[dbname][new_table[0]]=  new_table[1]

   return metadb


def db_meta_find(ALLDB, query='', filter_db=[],   filter_table=[], filter_column=[] ) :
  ''' Find string in all the meta table name, column
  db_meta_find(ALLDB, query='bottler', filter_db=['cokeon'],   filter_table=['table'], filter_column=['table'] )
  dbname: should be exact name
  fitler_table: partial match is ok
  fitler_column : partial name is ok
  return   (dbname, meta_table_name,  meta_table_filtered_by_row_containing query)
  '''
  rs=[]
  for iname, dbi in list(ALLDB.items()) :      #Search in All DB
    if iname in filter_db or len(filter_db)==0 :
      for jname, tablej in list(dbi.items()) :    # inside sub-meta table of DB
       if isinstance(tablej, pd.DataFrame) :  #Only Dataframe
        isok= util.str_match_fuzzy(jname, filter_table)
        if isok or len(filter_table)==0 :
          #print iname, jname
          #Each column of table
          colj= list(tablej.columns.values)
          if len(filter_column) > 0 :
             for colf in filter_column :
               aux= util.find_fuzzy(colf, colj)  #if colf name matches partially colj, OK
               colj= list(set(colj + aux))
               # colj= [ colji  for colji in colj if colji in filter_column ]

          df_new= util.pd_find(tablej, query, col_restrict=colj, isnumeric=False, doreturnposition=False)
          rs.append( (iname, jname, df_new ))   #DB name, meta_table, column_where_True

  return rs


######################  Study o   ###################################################################
def col_study_getcategorydict_freq(catedict) :
  ''' Generate Frequency of category : Id, Freq, Freqin%, CumSum%, ZScore
      given a dictionnary of category parsed previously
  '''
  catlist=[]
  for key,v in list(catedict.items()) :
     df= util.pd_array_todataframe(util.np_dict_tolist(v), ['category', 'freq'])
     df['freq_pct']= 100.0*df['freq'] / df['freq'].sum()
     df['freq_zscore']= df['freq'] / df['freq'].std()
     df= df.sort_values(by=['freq'], ascending=0)
     df['freq_cumpct']= 100.0*df['freq_pct'].cumsum()/df['freq_pct'].sum()
     df['rank']=  np.arange(0,len(df.index.values))
     catlist.append((key,df))
  return catlist

def col_feature_importance(Xcol, Ytarget) :
   ''' random forest for column importance '''
   pass

def pd_col_study_distribution_show(df, col_include=None, col_exclude=None, pars={'binsize':20}):
 '''  Perfom Full Study of the pandas columns'''
 if col_include is not None :
    features = [feature for feature in df.columns.values if  feature in col_include]
 elif  col_exclude is not None :
    features = [feature for feature in df.columns.values if not feature in col_exclude]

 for feature in features:
    values = df[feature].values
    nan_count = np.count_nonzero(np.isnan(values))
    values = sorted(values[~np.isnan(values)])
    print(('NaN count:', nan_count, 'Unique count:', len(np.unique(values))))
    print(('Max:', np.max(values), 'Min:', np.min(values)))
    print(('Median', np.median(values), 'Mean:', np.mean(values), 'Std:', np.std(values)))
    plot_Y(values, typeplot='.b', title='Values '+feature, figsize=(8,5))


    fit = sci.stats.norm.pdf(values, np.mean(values), np.std(values))  #this is a fitting indeed
    plt.title('Distribution Values '+feature)
    plt.plot(values,fit,'-g')
    plt.hist(values,normed=True, bins=pars['binsize'])      #use this to draw histogram of your data
    plt.show()

    plt.figure(figsize=(8,5))
    plt.title('Percentiles 5...95'+feature)
    plt.plot(list(range(5,100,5)), np.percentile(values, list(range(5,100,5))),'.b')
    plt.show()

    break

def col_study_summary(Xmat=[0.0, 0.0], Xcolname=['col1', 'col2'], Xcolselect=[9, 9], isprint=0):
   n, m=np.shape(Xmat)
   if Xcolselect== [9,9] :  Xcolselect= np.arange(0,m)
   if len(Xcolname) != m :  print('Error column size: ')  ; return None
   colanalysis=[]
   for icol in Xcolselect:
      Xraw_1unique=np.unique(Xmat[:, icol])
      vv= [Xcolname[icol], icol, len(Xraw_1unique), np.min(Xraw_1unique), np.max(Xraw_1unique), np.median(Xraw_1unique), np.mean(Xraw_1unique), np.std(Xraw_1unique)]
      colanalysis.append(vv)

   colanalysis= pd.DataFrame(colanalysis,
                 columns= ['Col_name', 'Col_id', 'Nb_Unique', 'MinVal', 'MaxVal', 'MedianVal', 'MeanVal', 'StdDev'])
   if isprint:
      print(('Nb_Samples:', np.shape(Xmat)[0], 'Nb Col:', len(Xcol)))
      print(colanalysis)
   return colanalysis

def pd_col_pair_plot(dfX, Xcolname_selectlist=None, dfY=None, Ycolname=None) :

  if dfY is None :  yy= dfX[Ycolname].values
  else :            yy= dfY[Ycolname].values

  for coli in Xcolname_selectlist :
    xx= dfX[coli].values
    title1= 'X: ' + str(coli) + ', Y: '+str(Ycolname[0])
    plt.scatter(xx, yy, s=1 )
    plt.autoscale(enable=True, axis='both', tight=None)
    #  plt.axis([-3, 3, -3, 3])  #gaussian
    plt.title(title1)
    plt.show()

def col_pair_correl(Xcol, Ytarget) :
   pass

def col_pair_interaction(Xcol, Ytarget):
   ''' random forest for pairwise interaction '''
   pass

def plot_col_pair(dfX, Xcolname_selectlist=None, dfY=None, Ycolname=None) :
  pd_col_pair_plot(dfX, Xcolname_selectlist=None, dfY=None, Ycolname=None)



######################  Transformation   ###########################################################
def tf_transform_catlabel_toint(Xmat):
   '''
     # ["paris", "paris", "tokyo", "amsterdam"]  --> 2 ,5,6
     # np.array(le.inverse_transform([2, 2, 1]))
     le = preprocessing.LabelEncoder()
     le.fit(["paris", "paris", "tokyo", "amsterdam"])
LabelEncoder()
list(le.classes_)
['amsterdam', 'paris', 'tokyo']
le.transform(["tokyo", "tokyo", "paris"])
array([2, 2, 1]...)
list(le.inverse_transform([2, 2, 1]))
['tokyo', 'tokyo', 'paris']
   '''
   le=sk.preprocessing.LabelEncoder()
   ncol=Xmat.shape[1]
   Xnew=np.zeros_like(Xmat)
   mapping_cat_int= {}
   
   for k in range(0, ncol):
      Xnew[:, k]=le.fit_transform(Xmat[:, k])
      mapping_cat_int[k] = le.get_params()

   return Xnew, mapping_cat_int

def tf_transform_pca(Xmat, dimpca=2, whiten=True):
   '''Project ndim data into dimpca sub-space  '''
   pca=PCA(n_components=dimpca, whiten=whiten).fit(Xmat)
   return pca.transform(Xmat)




######################  PLOT  ######################################################################
def plot_distance_heatmap(Xmat_dist, Xcolname):
    import matplotlib.pyplot as pyplot
    df= pd.DataFrame(Xmat_dist)
    df.columns= Xcolname
    df.index.name = "Col X"
    df.columns.name = "Col Y"

    fig = plt.figure()
    ax = fig.add_subplot(111)
    axim = ax.imshow(df.values,cmap = pyplot.get_cmap('RdYlGn'), interpolation = 'nearest')
    ax.set_xlabel(df.columns.name)
    ax.set_ylabel(df.index.name)
    ax.set_title("Pearson R Between Features")
    plt.colorbar(axim)


def plot_cluster_2D(X_2dim, target_class, target_names):
   ''' Plot 2d of Clustering Class,
       X2d: Nbsample x 2 dim  (projection on 2D sub-space)
   '''
   colors=itertools.cycle('rgbcmykw')
   target_ids=range(0, len(target_names))
   pl.figure()
   for i, c, label in zip(target_ids, colors, target_names):
      pl.scatter(X_2dim[target_class==i, 0], X_2dim[target_class==i, 1], c=c, label=label)
   pl.legend()
   pl.show()


def plot_cluster_tsne(Xmat, Xcluster_label=None, metric='euclidean', perplexity=50, ncomponent=2, savefile='',
                      isprecompute=False, returnval=True) :
 '''Plot High dimemnsionnal State using TSNE method
   'euclidean, 'minkowski', 'cityblock', 'seuclidean', 'sqeuclidean, 'cosine, 'correlation, 'hamming, 'jaccard, 'chebyshev,
   'canberra, 'braycurtis, 'mahalanobis', VI=None) 'yule, 'matching, 'dice, 'kulsinski, 'rogerstanimoto, 'russellrao, 'sokalmichener, 'sokalsneath,

   Xtsne= da.plot_cluster_tsne(Xtrain_dist, Xcluster_label=None, perplexity=40, ncomponent=2, isprecompute=True)

   Xtrain_dist= sci.spatial.distance.squareform(sci.spatial.distance.pdist(Xtrain_d,
               metric='cityblock', p=2, w=None, V=None, VI=None))
   '''
 from sklearn.manifold import TSNE
 if isprecompute : Xmat_dist= Xmat
 else : Xmat_dist= sci.spatial.distance.squareform(sci.spatial.distance.pdist(Xmat, metric=metric, p=ncomponent, w=None, V=None, VI=None))

 model = sk.manifold.TSNE(n_components=ncomponent,perplexity=perplexity, metric='precomputed', random_state=0)
 np.set_printoptions(suppress=True)
 X_tsne= model.fit_transform(Xmat_dist)

 # plot the result
 xx, yy =X_tsne[:, 0], X_tsne[:, 1]
 if Xcluster_label is None : Yclass= np.arange(0, X_tsne.shape[0])
 else:                       Yclass= Xcluster_label

 plot_XY(xx, yy, zcolor=Yclass, labels=Yclass, color_dot='plasma',savefile=savefile )

 if returnval : return X_tsne


def plot_cluster_pca(Xmat, Xcluster_label=None, metric='euclidean', dimpca=2, whiten=True, isprecompute=False, savefile='', doreturn=1) :
 from sklearn.decomposition import pca
 if isprecompute : Xmat_dist= Xmat
 else : Xmat_dist= sci.spatial.distance.squareform(sci.spatial.distance.pdist(Xmat, metric=metric, p=dimpca, w=None, V=None, VI=None))

 model = PCA(n_components=dimpca, whiten=whiten)
 X_pca= model.fit_transform(Xmat)

 # plot the result
 xx, yy =X_pca[:, 0], X_pca[:, 1]
 if Xcluster_label is None : Yclass= np.zeros( X_pca.shape[0])
 else:                       Yclass= Xcluster_label

 plot_XY(xx, yy, zcolor=Yclass, labels=Yclass, color_dot='plasma',savefile=savefile )

 if doreturn : return X_pca


def plot_cluster_hiearchy(Xmat_dist, p=30, truncate_mode=None, color_threshold=None,  get_leaves=True, orientation='top', labels=None,  count_sort=False, distance_sort=False, show_leaf_counts=True, do_plot=1, no_labels=False, leaf_font_size=None, leaf_rotation=None, leaf_label_func=None, show_contracted=False, link_color_func=None, ax=None,
                          above_threshold_color='b', annotate_above=0):
   from scipy.cluster.hierarchy import dendrogram, linkage
   from scipy.cluster.hierarchy import cophenet
   from scipy.spatial.distance import pdist

   ddata= dendrogram(Xmat_dist, p=30, truncate_mode= truncate_mode, color_threshold=color_threshold,
                     get_leaves=get_leaves, orientation='top', labels=None,
                     count_sort=False, distance_sort=False, show_leaf_counts=True,
                     no_plot= 1-do_plot, no_labels=False, leaf_font_size=None,
                     leaf_rotation=None, leaf_label_func=None,
                     show_contracted=False, link_color_func=None, ax=None,
                     above_threshold_color='b')

   if do_plot :
      annotate_above= 0
      plt.title('Hierarchical Clustering Dendrogram (truncated)')
      plt.xlabel('sample index or (sk_cluster size)')
      plt.ylabel('distance')
      for i, d, c in zip(ddata['icoord'], ddata['dcoord'], ddata['color_list']):
         x=0.5 * sum(i[1:3])
         y=d[1]
         if y > annotate_above  :
            plt.plot(x, y, 'o', c=c)
            plt.annotate("%.3g" % y, (x, y), xytext=(0, -5), textcoords='offset points', va='top', ha='center')
      if color_threshold:
         plt.axhline(y=color_threshold, c='k')
   return ddata


def plot_distribution_density(Xsample, kernel='gaussian', N=10, bandwith=1 / 10.0) :
  import statsmodels.api as sm
  from sklearn.neighbors import KernelDensity
  ''' from scipy.optimize import brentq
import statsmodels.api as sm
import numpy as np

# fit
kde = sm.nonparametric.KDEMultivariate()  # ... you already did this

# sample
u = np.random.random()

# 1-d root-finding
def fun_apply(x):
    return kde.cdf([x]) - u
sample_x = brentq(fun_apply, -99999999, 99999999)  # read brentq-docs about these constants
                                              # constants need to be sign-changing for the function
  '''

  fig, ax = plt.subplots()
  XN= len(Xsample)
  xmin, xmax= np.min(Xsample), np.max(Xsample)
  X_plot = np.linspace(xmin, xmax , XN)[:, np.newaxis]
  bins =   np.linspace(xmin, xmax , N)

  # Xhist, Xbin_edges= np.histogram(Xsample, bins=bins, range=None, normed=False, weights=None, density=True)

  weights = np.ones_like(Xsample)/len(Xsample) #  np.ones(len(Xsample))  #
   # ax2.hist(ret5d,50, normed=0,weights=weights,  facecolor='green')
  ax.hist(Xsample, bins=N, normed=0, weights=weights, fc='#AAAAFF')


  kde = sk.neighbors.KernelDensity(kernel=kernel, bandwidth=bandwith).fit(Xsample.reshape(-1, 1))
  log_dens = kde.score_samples(X_plot)
  log_dens -= np.log(XN)  #Normalize
  ax.plot(X_plot[:,0], np.exp(log_dens), '-',  label="kernel = '{0}'".format(kernel))

  ax.set_xlim(xmin, xmax)
  plt.show()
  return kde


def plot_Y(Yval, typeplot='.b', tsize=None, labels=None, title='', xlabel='', ylabel='', zcolor_label='', figsize=(8, 6), dpi=75, savefile='', color_dot='Blues', doreturn=0) :
    plt.figure(figsize=figsize)
    plt.title('Values ' + title)
    plt.plot(Yval, typeplot)
    plt.show()


def plot_XY(xx, yy, zcolor=None, tsize=None, labels=None, title='', xlabel='', ylabel='', zcolor_label='', figsize=(8, 6), dpi=75, savefile='', color_dot='Blues', doreturn=0) :
  '''
      labels= numpy array, ---> Generate HTML File with the labels interactives
      Color: Plasma
  '''


  #Color change
  if zcolor is None :  c= [[0, 0, 0]]
  elif isinstance(zcolor, int) : zcolor= zcolor
  else :
     aux= np.array(zcolor, dtype= np.float64)
     c= np.abs(aux)
  cmhot = plt.get_cmap(color_dot)

  #Marker size
  if tsize is None :  tsize=  50
  elif isinstance(tsize, int) : tsize= tsize
  else :
     aux= np.array(tsize, dtype= np.float64)
     tsize=  np.abs(aux)
     tsize=(tsize - np.min(tsize)) / (np.max(tsize) - np.min(tsize)) * 130 + 1

  #Plot
  fig, ax1 = plt.subplots(nrows=1, ncols=1)

  #Overall Plot
  fig.set_size_inches(figsize[0], figsize[1] )
  fig.set_dpi(dpi)
  fig.tight_layout()

  #Scatter
  scatter= ax1.scatter(xx, yy, c=c, cmap=cmhot, s=tsize, alpha=0.5)
  ax1.set_xlabel(xlabel, fontsize=9)
  ax1.set_ylabel(ylabel, fontsize=9)
  ax1.set_title(title)
  ax1.grid(True)
  # fig.autoscale(enable=True, axis='both')
  # fig.colorbar(ax1)

  c_min, c_max= np.min(c), np.max(c)
  scatter.set_clim([c_min,c_max])
  cb = fig.colorbar(scatter)
  cb.set_label(zcolor_label)

  # Add colorbar, make sure to specify tick locations to match desired ticklabels
  #cax = ax1.imshow(c, interpolation='nearest', cmap=color_dot)

  # cbar = fig.colorbar(ax1, ticks= xrange(c_min, c_max, 10))
  # cbar.ax.set_yticklabels([str(c_min), str(c_max)])  # vertically oriented colorbar
  #plt.clim(-0.5, 9.5)

  if labels is not None :  #Interactive HTML
     import mpld3
     labels = list(labels)
     tooltip = mpld3.plugins.PointLabelTooltip(scatter, labels=labels)
     mpld3.plugins.connect(fig, tooltip)
     mpld3.save_html(fig, savefile +'.html')

  plt.show()
  if savefile != '' :
     util.os_folder_create(os.path.split(savefile)[0])
     plt.savefig(savefile)

  if doreturn :   return fig, ax1


def plot_XY_plotly(xx, yy, towhere='url') :
  ''' Create Interactive Plotly   '''
  import plotly.plotly as py
  import plotly.graph_objs as go
  from plotly.graph_objs import Marker, ColorBar

  '''
  trace = go.Scatter(x= xx, y= yy, marker= Marker(
            size=16,
            cmax=39,
            cmin=0,
            color=[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39],
            colorbar=ColorBar(title='Colorbar' )),  colorscale='Viridis')
  '''
  trace = go.Scatter(x= xx, y= yy, mode='markers')

  data = [trace]
  if towhere=='ipython' :
     py.iplot(data, filename='basic-scatter')
  else :
     url= py.plot(data, filename='basic-scatter')


def plot_XY_seaborn(X, Y, Zcolor=None):
    import seaborn as sns
    sns.set_context('poster');  sns.set_color_codes()
    plot_kwds = {'alpha' : 0.35, 's' : 60, 'linewidths':0}
    palette = sns.color_palette('deep', np.unique(Zcolor).max() + 1)
    colors = [palette[x] if x >= 0 else (0.0, 0.0, 0.0) for x in Zcolor]
    plt.scatter(X, Y, c=colors, **plot_kwds)
    frame = plt.gca()
    frame.axes.get_xaxis().set_visible(False); frame.axes.get_yaxis().set_visible(False)
    plt.title('X:   , Y:   ,Z:', fontsize=18)



'''
def plot_cluster_embedding(Xmat, title=None):
   # Scale and visualize the embedding vectors
   x_min, x_max=np.min(Xmat, 0), np.max(Xmat, 0)
   Xmat=(Xmat - x_min) / (x_max - x_min)
   nX= Xmat.shape[0]

   plt.figure()
   ax=plt.subplot(111)
   colors= np.arange(0, nX, 5)
   for i in range(nX):
      plt.text(Xmat[i, 0], Xmat[i, 1], str(labels[i]), color=plt.cm.Set1(colors[i] / 10.), fontdict={'weight': 'bold', 'size': 9})

   if hasattr(offsetbox, 'AnnotationBbox'):
      # only print thumbnails with matplotlib > 1.0
      shown_images=np.array([[1., 1.]])  # just something big
      for i in range(digits.data.shape[0]):
         dist=np.sum((Xmat[i] - shown_images) ** 2, 1)
         if np.min(dist) < 4e-3: continue  # don't show points that are too close

         shown_images=np.r_[shown_images, [Xmat[i]]]
         imagebox=offsetbox.AnnotationBbox(offsetbox.OffsetImage(digits.images[i], cmap=plt.cm.gray_r), Xmat[i])
         ax.add_artist(imagebox)
   plt.xticks([]), plt.yticks([])
   if title is not None:  plt.title(title)
'''


######################### OPTIM   ###################################################
def optim_is_pareto_efficient(Xmat_cost, epsilon= 0.01, ret_boolean=1):
    """ Calculate Pareto Frontier of Multi-criteria Optimization program
    c1, c2  has to be minimized : -Sharpe, -Perf, +Drawdown
    :param Xmat_cost: An (n_points, k_costs) array
    :return: A (n_points, ) boolean array, indicating whether each point is Pareto efficient
    """
    pesp= 1.0 + epsilon   #Relax Pareto Constraints
    is_efficient = np.ones(Xmat_cost.shape[0], dtype = bool)
    for i, c in enumerate(Xmat_cost):
        if is_efficient[i]:
            is_efficient[is_efficient] = np.any(Xmat_cost[is_efficient] <= c * pesp, axis=1)  # Remove dominated points
    if ret_boolean : return is_efficient
    else :           return Xmat_cost[is_efficient]
    #return is_efficient





######################  Category Classifier Trees  #########################################################################
'''
Category Classifier
https://github.com/catboost/catboost/blob/master/catboost/tutorials/kaggle_paribas.ipynb

Very Efficient
D:\_devs\Python01\project27\linux_project27\mlearning\category_learning


https://tech.yandex.com/catboost/doc/dg/concepts/python-usages-examples-docpage/


clf = CatBoostClassifier(learning_rate=0.1, iterations=1000, random_seed=0)
clf.fit(train_df, labels, cat_features=cat_features_ids)


##### Base Approach
import pandas as pd
import numpy as np

from itertools import combinations
from catboost import CatBoostClassifier


labels = train_df.target
test_id = test_df.ID

train_df.drop(['ID', 'target'], axis=1, inplace=True)
test_df.drop(['ID'], axis=1, inplace=True)

train_df.fillna(-9999, inplace=True)
test_df.fillna(-9999, inplace=True)

# Keep list of all categorical features in dataset to specify this for CatBoost
cat_features_ids = np.where(train_df.apply(pd.Series.nunique) < 30000)[0].tolist()



########  Regularizer  
selected_features = [
    'v10', 'v12', 'v14', 'v21', 'v22', 'v24', 'v30', 'v31', 'v34', 'v38', 'v40', 'v47', 'v50',
    'v52', 'v56', 'v62', 'v66', 'v72', 'v75', 'v79', 'v91', 'v112', 'v113', 'v114', 'v129'
]

# drop some of the features that were not selected
train_df = train_df[selected_features]
test_df = test_df[selected_features]

# update the list of categorical features
cat_features_ids = np.where(train_df.apply(pd.Series.nunique) < 30000)[0].tolist()


char_features = list(train_df.columns[train_df.dtypes == np.object])
char_features_without_v22 = list(train_df.columns[(train_df.dtypes == np.object) & (train_df.columns != 'v22')])

cmbs = list(combinations(char_features, 2)) + map(lambda x: ("v22",) + x, combinations(char_features_without_v22, 2))


clf = CatBoostClassifier(learning_rate=0.1, iterations=1000, random_seed=0)
clf.fit(train_df, labels, cat_features=cat_features_ids)


'''



def sk_catboost_classifier(Xtrain, Ytrain, Xcolname=None,
                        pars= {"learning_rate":0.1, "iterations":1000, "random_seed":0, "loss_function": "MultiClass" }, isprint=0) :
  '''
  from catboost import Pool, CatBoostClassifier

TRAIN_FILE = '../data/cloudness_small/train_small'
TEST_FILE = '../data/cloudness_small/test_small'
CD_FILE = '../data/cloudness_small/train.cd'
# Load data from files to Pool
train_pool = Pool(TRAIN_FILE, column_description=CD_FILE)
test_pool = Pool(TEST_FILE, column_description=CD_FILE)
# Initialize CatBoostClassifier
model = CatBoostClassifier(iterations=2, learning_rate=1, depth=2, loss_function='MultiClass')
# Fit model
model.fit(train_pool)
# Get predicted classes
preds_class = model.predict(test_pool)
# Get predicted probabilities for each class
preds_proba = model.predict_proba(test_pool)
# Get predicted RawFormulaVal
  preds_raw = model.predict(test_pool, prediction_type='RawFormulaVal')  
  
  
  https://tech.yandex.com/catboost/doc/dg/concepts/python-usages-examples-docpage/
  
  '''
  import catboost
  pa= dict2(pars)

  if Xcolname is None :  Xcolname = [ str(i) for i  in range(0,  Xtrain.shape[1]) ]
  train_df= pd.DataFrame(Xtrain, Xcolname)
  cat_features_ids= Xcolname

  clf = catboost.CatBoostClassifier(learning_rate= pa.learning_rate, iterations= pa.iterations,
                                    random_seed= pa.random_seed,
                                    loss_function=pa.loss_function)
  clf.fit(Xtrain, Ytrain, cat_features=cat_features_ids)

  Y_pred = clf.predict(Xtrain)

  cm = sk.metrics.confusion_matrix(Ytrain, Y_pred); cm_norm = cm.astype('float') / cm.sum(axis=1)[:, np.newaxis]
  if isprint: print(( cm_norm[0,0] + cm_norm[1,1])); print(cm_norm); print(cm)
  return clf, cm, cm_norm



def sk_catboost_regressor():
   pass














######################  ALGO  #########################################################################
def sk_model_auto_tpot(Xmat, y,  outfolder='aaserialize/', model_type='regressor/classifier', train_size=0.5, generation=1, population_size=5, verbosity=2 ):
  ''' Automatic training of Xmat--->Y, Generate SKlearn code in outfile
      Very Slow Process, use lower number of Sample
  :param Xmat:
  :param y:
  :param outfolder:
  :param model_type:
  :param train_size:
  :param generation: 
  :param population_size:
  :param verbosity:
  :return:
  '''
  from tpot import TPOTClassifier,   TPOTRegressor
  X_train, X_test, y_train, y_test = train_test_split(Xmat, Y,     train_size=0.5,    test_size=0.5)

  if model_type == 'regressor' :
    tpot = TPOTRegressor(generations=generation, population_size=population_size, verbosity=verbosity)
  elif model_type=='classifier' :
    tpot = TPOTClassifier(generations=generation, population_size=population_size, verbosity=verbosity)

  tpot.fit(X_train, y_train)
  print((tpot.score(X_test, y_test)))
  file1= DIRCWD+ '/'+outfolder+'/tpot_regression_pipeline_'+ str(np.random.randint(1000, 9999)) +'.py'
  tpot.export(file1)
  return file1


def sk_params_search_best(Xmat, Ytarget, model1, param_grid={'alpha':  np.linspace(0,1, 5) }, method='gridsearch',
                          param_search= {'scoretype':'r2', 'cv':5, 'population_size':5, 'generations_number':3 }) :
  '''
   genetic: population_size=5, ngene_mutation_prob=0.10,,gene_crossover_prob=0.5, tournament_size=3,  generations_number=3

  :param Xmat:
  :param Ytarget:
  :param model1:
  :param param_grid:
  :param method:
  :param param_search:
  :return:
  '''
  p= param_search

  from sklearn.metrics import  make_scorer,  r2_score

  if param_search['scoretype'] =='r2' :  myscore = make_scorer(r2_score, sample_weight=None)

  if method=='gridsearch' :
    from sklearn.grid_search import GridSearchCV
    grid = GridSearchCV(model1,param_grid, cv=p["cv"], scoring=myscore) # 20-fold cross-validation
    grid.fit(Xmat, Ytarget)
    return grid.best_score_, grid.best_params_


  if method=="genetic" :
    from evolutionary_search import EvolutionaryAlgorithmSearchCV
    from sklearn.model_selection import StratifiedKFold
    # paramgrid = {"alpha":  np.linspace(0,1, 20) , "l1_ratio": np.linspace(0,1, 20) }
    cv = EvolutionaryAlgorithmSearchCV(estimator=model1,    params=param_grid,  scoring=p['scoretype'],
                                   cv=StratifiedKFold(y,    n_folds=p['cv']), verbose=True,  population_size=p['population_size'],
                                   gene_mutation_prob=0.10, gene_crossover_prob=0.5,
                                   tournament_size=3,       generations_number=p['generations_number'])

    cv.fit(Xmat, Ytarget)
    return cv.best_score_, cv.best_params_

  '''
   from sklearn.metrics import  make_scorer,  r2_score
from sklearn.grid_search import GridSearchCV

myscore = make_scorer(r2_score, sample_weight=None)

param_grid= {'alpha': np.linspace(0.01, 1.5, 10),
             'ww': [[0.05, 0.95]],
             'low_y_cut': [-10.0], 'high_y_cut': [9.0]   }

grid = GridSearchCV(model1(),param_grid, cv=10, scoring=myscore) # 20-fold cross-validation
grid.fit(Xtrain, Ytrain)
grid.best_params_

# Weight Search
wwl= np.linspace(0.01, 1.0, 5)
param_grid= {'alpha':  [0.01],
             'ww0': wwl,
             'low_y_cut': [-0.08609*1000], 'high_y_cut': [0.09347*1000]   }

grid = GridSearchCV(model1(),param_grid, cv=10, scoring=myscore) # 20-fold cross-validation
grid.fit(X*100.0, Ytarget*1000.0)
grid.best_params_

# {'alpha': 0.01, 'high_y_cut': 93.47, 'low_y_cut': -86.09, 'ww0': 0.01}

   '''


def sk_distribution_kernel_bestbandwidth(kde):
 '''Find best Bandwidht for a  given kernel
  :param kde:
  :return:
 '''
 from sklearn.grid_search import GridSearchCV
 grid = GridSearchCV(kde,{'bandwidth': np.linspace(0.1, 1.0, 30)}, cv=20) # 20-fold cross-validation
 grid.fit(x[:, None])
 return grid.best_params_


def sk_distribution_kernel_sample(kde=None, n=1):
 ''' 
  kde = sm.nonparametric.KDEUnivariate(np.array(Y[Y_cluster==0],dtype=np.float64))
  kde = sm.nonparametric.KDEMultivariate()  # ... you already did this
 '''

 from scipy.optimize import brentq
 samples= np.zeros(n)
 
 # 1-d root-finding  F-1(U) --> Sample
 def func(x):
    return kde.cdf([x]) - u

 for i in range(0,n):
   u = np.random.random()   # sample
   samples[i]=  brentq(func, -999, 999)  # read brentq-docs about these constants
 return samples

def sk_correl_rank(correl=[[1,0],[0,1]]) :
 """ Correl Ranking:  Col i, Col j, Correl_i_j, Abs_Correl_i_j    """
 m,n= np.shape(correl)
 correl_rank= np.zeros((n*(n-1)/2, 3), dtype=np.float32)
 k=0
 for i in range(0,n) :
    for j in range(i+1,n) :
      k+=1
      correl_rank[k, 0]= i
      correl_rank[k, 1]= j
      correl_rank[k, 2]= correl[i, j]
      correl_rank[k, 3]= abs(correl[i, j])
 correl_rank= util.sortcol(correl_rank, 3, asc=False)
 return correl_rank

def sk_error_r2(Ypred, y_true, sample_weight=None, multioutput=None):
    from sklearn.metrics import r2_score
    r2 = r2_score(y_true, Ypred, sample_weight=sample_weight, multioutput=multioutput)
    r =  np.sign(r2)*np.sqrt(np.abs(r2))
    if r <= -1:  return -1
    else:        return r

def sk_error_rmse(Ypred, Ytrue) :
 aux= np.sqrt(np.sum((Ypred-Ytrue)**2) ) / len(Ytrue)
 return "Error:", aux, "Error/Stdev:", aux / np.std(Ytrue)

def sk_cluster_distance_pair(Xmat, metric='jaccard'):
   '''
    'euclidean, 'minkowski', 'cityblock', 'seuclidean', 'sqeuclidean, 'cosine, 'correlation, 'hamming, 'jaccard, 'chebyshev, 'canberra, 'braycurtis, 'mahalanobis', VI=None) 'yule, 'matching, 'dice, 'kulsinski, 'rogerstanimoto, 'russellrao, 'sokalmichener, 'sokalsneath,

    'braycurtis': hdbscan.dist_metrics.BrayCurtisDistance,
 'canberra': hdbscan.dist_metrics.CanberraDistance,
 'chebyshev': hdbscan.dist_metrics.ChebyshevDistance,
 'cityblock': hdbscan.dist_metrics.ManhattanDistance,
 'dice': hdbscan.dist_metrics.DiceDistance,
 'euclidean': hdbscan.dist_metrics.EuclideanDistance,
 'hamming': hdbscan.dist_metrics.HammingDistance,
 'haversine': hdbscan.dist_metrics.HaversineDistance,
 'infinity': hdbscan.dist_metrics.ChebyshevDistance,
 'jaccard': hdbscan.dist_metrics.JaccardDistance,
 'kulsinski': hdbscan.dist_metrics.KulsinskiDistance,
 'l1': hdbscan.dist_metrics.ManhattanDistance,
 'l2': hdbscan.dist_metrics.EuclideanDistance,
 'mahalanobis': hdbscan.dist_metrics.MahalanobisDistance,
 'manhattan': hdbscan.dist_metrics.ManhattanDistance,
 'matching': hdbscan.dist_metrics.MatchingDistance,
 'minkowski': hdbscan.dist_metrics.MinkowskiDistance,
 'p': hdbscan.dist_metrics.MinkowskiDistance,
 'pyfunc': hdbscan.dist_metrics.PyFuncDistance,
 'rogerstanimoto': hdbscan.dist_metrics.RogersTanimotoDistance,
 'russellrao': hdbscan.dist_metrics.RussellRaoDistance,
 'seuclidean': hdbscan.dist_metrics.SEuclideanDistance,
 'sokalmichener': hdbscan.dist_metrics.SokalMichenerDistance,
 'sokalsneath': hdbscan.dist_metrics.SokalSneathDistance,
 'wminkowski': hdbscan.dist_metrics.WMinkowskiDistance}
   #Visualize discretization scheme

   Xtrain_dist= sci.spatial.distance.squareform(sci.spatial.distance.pdist(Xtrain_d,
             metric='cityblock', p=2, w=None, V=None, VI=None))

   Xtsne= da.plot_cluster_tsne(Xtrain_dist, metric='', perplexity=40, ncomponent=2, isprecompute=True)

   '''
   if metric=='jaccard':  return fast.distance_jaccard_X(Xmat)

   else  :  # if metric=='euclidian'
      return  sci.spatial.distance.squareform(sci.spatial.distance.pdist(Xmat, metric=metric, p=2, w=None, V=None, VI=None))

def sk_cluster(Xmat, method='kmode',  args=(), kwds={'metric':'euclidean', 'min_cluster_size':150, 'min_samples':3}, isprint=1, preprocess={'norm': False}) :
   '''
   'hdbscan',(), kwds={'metric':'euclidean', 'min_cluster_size':150, 'min_samples':3 }
   'kmodes',(), kwds={ n_clusters=2, n_init=5, init='Huang', verbose=1 }
   'kmeans',    kwds={ n_clusters= nbcluster }

   Xmat[ Xcluster== 5 ]
   # HDBSCAN Clustering
   Xcluster_hdbscan= da.sk_cluster_algo_custom(Xtrain_d, hdbscan.HDBSCAN, (),
                  {'metric':'euclidean', 'min_cluster_size':150, 'min_samples':3})

   print len(np.unique(Xcluster_hdbscan))

   Xcluster_use =  Xcluster_hdbscan

# Calculate Distribution for each cluster
kde= da.plot_distribution_density(Y[Xcluster_use== 2], kernel='gaussian', N=200, bandwith=1 / 500.)
kde.sample(5)

   '''
   if method=='kmode':
      # Kmode clustering data nbCategory,  NbSample, NbFeatures
      km=kmodes.kmodes.KModes(*args, **kwds)
      Xclus_class=km.fit_predict(Xmat)
      return Xclus_class, km, km.cluster_centroids_  # Class, km, centroid

   if method=='hdbscan' :
      import hdbscan
      Xcluster_id =  hdbscan.HDBSCAN(*args, **kwds).fit_predict(Xmat)
      print(('Nb Cluster', len(np.unique(Xcluster_id))))
      return Xcluster_id

   if method=='kmeans' :
    from sklearn.cluster import KMeans

    if preprocess['norm'] :
      stdev=  np.std(Xmat, axis=0)
      Xmat=   (Xmat - np.mean(Xmat, axis=0)) / stdev

    sh= Xmat.shape
    Xdim= 1 if len(sh) < 2 else sh[1]   #1Dim vector or 2dim-3dim vector
    print((len(Xmat.shape), Xdim))
    if Xdim==1 :  Xmat= Xmat.reshape((sh[0],1))

    kmeans = KMeans(**kwds)    #  KMeans(n_clusters= nbcluster)
    kmeans.fit(Xmat)
    centroids, labels= kmeans.cluster_centers_,  kmeans.labels_

    if isprint :
      import matplotlib.pyplot as plt
      colors = ["g.","r.","y.","b.", "k."]
      if Xdim==1 :
        for i in range(0, sh[0], 5):  plt.plot(Xmat[i], colors[labels[i]], markersize = 5)
        plt.show()
      elif Xdim==2 :
        for i in range(0, sh[0], 5):  plt.plot(Xmat[i,0], Xmat[i,1], colors[labels[i]], markersize = 2)
        plt.show()
      else :
        print('Cannot Show higher than 2dim')

    return labels, centroids


def sk_cluster_algo_custom(Xmat, algorithm, args, kwds, returnval=1):
    ''' Plot the cLuster using specific Algo
    distance_matrix = pairwise_distances(blobs)
    clusterer = hdbscan.HDBSCAN(metric='precomputed')
    clusterer.fit(distance_matrix)
    clusterer.labels_

    {'braycurtis': hdbscan.dist_metrics.BrayCurtisDistance,
 'canberra': hdbscan.dist_metrics.CanberraDistance,
 'chebyshev': hdbscan.dist_metrics.ChebyshevDistance,
 'cityblock': hdbscan.dist_metrics.ManhattanDistance,
 'dice': hdbscan.dist_metrics.DiceDistance,
 'euclidean': hdbscan.dist_metrics.EuclideanDistance,
 'hamming': hdbscan.dist_metrics.HammingDistance,
 'haversine': hdbscan.dist_metrics.HaversineDistance,
 'infinity': hdbscan.dist_metrics.ChebyshevDistance,
 'jaccard': hdbscan.dist_metrics.JaccardDistance,
 'kulsinski': hdbscan.dist_metrics.KulsinskiDistance,
 'l1': hdbscan.dist_metrics.ManhattanDistance,
 'l2': hdbscan.dist_metrics.EuclideanDistance,
 'mahalanobis': hdbscan.dist_metrics.MahalanobisDistance,
 'manhattan': hdbscan.dist_metrics.ManhattanDistance,
 'matching': hdbscan.dist_metrics.MatchingDistance,
 'minkowski': hdbscan.dist_metrics.MinkowskiDistance,
 'p': hdbscan.dist_metrics.MinkowskiDistance,
 'pyfunc': hdbscan.dist_metrics.PyFuncDistance,
 'rogerstanimoto': hdbscan.dist_metrics.RogersTanimotoDistance,
 'russellrao': hdbscan.dist_metrics.RussellRaoDistance,
 'seuclidean': hdbscan.dist_metrics.SEuclideanDistance,
 'sokalmichener': hdbscan.dist_metrics.SokalMichenerDistance,
 'sokalsneath': hdbscan.dist_metrics.SokalSneathDistance,
 'wminkowski': hdbscan.dist_metrics.WMinkowskiDistance}

    '''
    import  hdbscan
    import sklearn.cluster as cluster

    cluster_id = algorithm(*args, **kwds).fit_predict(Xmat)

    print(('Nb Cluster', len(np.unique(cluster_id))))
    if returnval : return cluster_id
    # plt.text(-0.5, 0.7, 'Clustering took {:.2f} s'.format(end_time - start_time), fontsize=14)

'''
def sk_cluster_kmeans(Xmat, nbcluster=5, isprint=False, isnorm=False) :
  from sklearn.cluster import k_means
  stdev=  np.std(Xmat, axis=0)
  if isnorm  : Xmat=   (Xmat - np.mean(Xmat, axis=0)) / stdev

  sh= Xmat.shape
  Xdim= 1 if len(sh) < 2 else sh[1]   #1Dim vector or 2dim-3dim vector
  print(len(Xmat.shape), Xdim)
  if Xdim==1 :  Xmat= Xmat.reshape((sh[0],1))

  kmeans = sk.cluster.KMeans(n_clusters= nbcluster)
  kmeans.fit(Xmat)
  centroids, labels= kmeans.cluster_centers_,  kmeans.labels_

  if isprint :
   import matplotlib.pyplot as plt
   colors = ["g.","r.","y.","b.", "k."]
   if Xdim==1 :
     for i in range(0, sh[0], 5):  plt.plot(Xmat[i], colors[labels[i]], markersize = 5)
     plt.show()
   elif Xdim==2 :
     for i in range(0, sh[0], 5):  plt.plot(Xmat[i,0], Xmat[i,1], colors[labels[i]], markersize = 2)
     plt.show()
   else :
      print('Cannot Show higher than 2dim')

  return labels, centroids, stdev
'''

def sk_optim_de(obj_fun, bounds, maxiter=1, name1='', solver1=None, isreset=1, popsize=15):
    ''' Optimization and Save Data into file'''
    import copy
    if isreset == 2:
        print('Traditionnal Optim, no saving')
        res = sci.optimize.differential_evolution(obj_fun, bounds=bounds, maxiter=maxiter)
        xbest, fbest, solver, i = res.x, res.fun, '', maxiter
    else:  # iterative solver
        print('Iterative Solver ')
        if name1 != '':  # wtih file
            print('/batch/' + name1)
            solver2 = load_obj('/batch/' + name1)
            imin = int(name1[-3:]) + 1
            solver = sci.optimize._differentialevolution.DifferentialEvolutionSolver(obj_fun, bounds=bounds,
                                                                                     popsize=popsize)
            solver.population = copy.deepcopy(solver2.population)
            solver.population_energies = copy.deepcopy(solver2.population_energies)
            del solver2

        elif solver1 is not None:  # Start from zero
            solver = copy.deepcopy(solver1)
            imin = 0
        else:
            solver = sci.optimize._differentialevolution.DifferentialEvolutionSolver(obj_fun, bounds=bounds,
                                                                                     popsize=popsize);
            imin = 0

        name1 = '/batch/solver_' + name1
        fbest0 = 1500000.0
        for i in range(imin, imin + maxiter):
            xbest, fbest = next(solver)
            print(0, i, fbest, xbest)
            res = (copy.deepcopy(solver), i, xbest, fbest)
            try:
                util.save_obj(solver, name1 + util.date_now() + '_' + util.np_int_tostr(i))
                print(( name1 + util.date_now() + '_' + util.np_int_tostr(i)))
            except:
                pass
            if np.mod(i + 1, 11) == 0:
                if np.abs(fbest - fbest0) < 0.001: break;
                fbest0 = fbest

    return fbest, xbest, solver



######## Valuation model template  ##########################################################
class sk_model_template1(sk.base.BaseEstimator) :
   def __init__(self, alpha=0.5, low_y_cut=-0.09, high_y_cut=0.09, ww0=0.95) :
      from sklearn.linear_model import Ridge
      self.alpha= alpha
      self.low_y_cut, self.high_y_cut,self.ww0  = 1000.0*low_y_cut, 1000.0*high_y_cut, ww0
      self.model= Ridge(alpha=self.alpha)

   def fit(self, X, Y=None) :
      X, Y= X*100.0, Y *1000.0

      y_is_above_cut = (Y >  self.high_y_cut)
      y_is_below_cut = (Y < self.low_y_cut)
      y_is_within_cut = (~y_is_above_cut & ~y_is_below_cut)
      if len(y_is_within_cut.shape) > 1 : y_is_within_cut= y_is_within_cut[:,0] 
  
      self.model.fit(X[y_is_within_cut,:], Y[y_is_within_cut])

      r2= self.model.score(X[y_is_within_cut,:], Y[y_is_within_cut])
      print(('R2:',  r2  ))      
      print(('Inter', self.model.intercept_ ))
      print(('Coef', self.model.coef_ ))

      self.ymedian = np.median(Y)
      return self, r2,  self.model.coef_ 

   def predict(self, X, y=None, ymedian=None) :
     X= X * 100.0

     if ymedian is None : ymedian= self.ymedian
     Y= self.model.predict(X)
     Y= Y.clip(self.low_y_cut, self.high_y_cut)
     Y=  (self.ww0 * Y  + (1-self.ww0 )* ymedian )

     Y= Y / 1000.0
     return Y

   def score(self, X, Ytrue=None, ymedian=None) :
     from sklearn.metrics import   r2_score
     X= X * 100.0

     if ymedian is None : ymedian= self.ymedian
     Y= self.model.predict(X)
     Y= Y.clip(self.low_y_cut, self.high_y_cut)
     Y=  (self.ww0 * Y  + (1-self.ww0 )* ymedian )
     Y= Y / 1000.0
     return r2_score(Ytrue, Y)
     


############################################################################
# ---------------------             ----------------
'''
 Reshape your data either using X.reshape(-1, 1) if your data has a single feature or
  X.reshape(1, -1) if it contains a single sample.

'''
def sk_feature_importance(clfrf, feature_name) :
 importances = clfrf.feature_importances_
 indices = np.argsort(importances)[::-1]
 for f in range(0, len(feature_name)):
    if importances[indices[f]] > 0.0001 :
      print(str(f + 1), str(indices[f]),  feature_name[indices[f]], str(importances[indices[f]]))


def sk_showconfusion(clfrf, X_train,Y_train, isprint=True):
  Y_pred = clfrf.predict(X_train) ;
  cm = sk.metrics.confusion_matrix(Y_train, Y_pred); cm_norm = cm.astype('float') / cm.sum(axis=1)[:, np.newaxis]
  if isprint: print(( cm_norm[0,0] + cm_norm[1,1])); print(cm_norm); print(cm)
  return cm, cm_norm, cm_norm[0,0] + cm_norm[1,1]



#-------- SK Learn TREE UTIL----------------------------------------------------------------
def sk_tree(Xtrain,Ytrain, nbtree, maxdepth, isprint1=1, njobs=1):
  # X_train, X_test, y_train, y_test = train_test_split(X, Y, test_size=.2)
  clfrf= sk.ensemble.RandomForestClassifier( n_estimators=nbtree, max_depth=maxdepth, max_features="sqrt",
                             criterion="entropy", n_jobs=njobs, min_samples_split=2, min_samples_leaf=2, class_weight= "balanced")
  clfrf.fit(Xtrain, Ytrain)
  Y_pred = clfrf.predict(Xtrain)

  cm = sk.metrics.confusion_matrix(Ytrain, Y_pred); cm_norm = cm.astype('float') / cm.sum(axis=1)[:, np.newaxis]
  if isprint1: print(( cm_norm[0,0] + cm_norm[1,1])); print(cm_norm); print(cm)
  return clfrf, cm, cm_norm


def sk_gen_ensemble_weight(vv, acclevel, maxlevel=0.88):
 imax= min(acclevel,len(vv))
 estlist= np.empty(imax, dtype= np.object) ; estww=[]
 for i in range(0, imax ) :
  #if vv[i,3]> acclevel:
   estlist[i]= vv[i,1] ;  estww.append( vv[i,3] )
   #print 5
  #Log Proba Weighted + Impact of recent False discovery
 estww= np.log( 1/(maxlevel - np.array(estww)/2.0) )
 # estww= estww/np.sum(estww)
# return np.array(estlist), np.array(estww)
 return estlist, np.array(estww)


def sk_votingpredict(estimators, voting, ww, X_test) :
  ww= ww/np.sum(ww)
  Yproba0= np.zeros((len(X_test),2))
  Y1= np.zeros((len(X_test)))

  for k,clf in enumerate(estimators) :
     Yproba= clf.predict_proba(X_test)
     Yproba0= Yproba0 + ww[k]*Yproba

  for k in range(0,  len(X_test)) :
     if  Yproba0[k,0] > Yproba0[k,1]:      Y1[k]=-1
     else : Y1[k]=1
  return  Y1, Yproba0


def sk_tree_get_ifthen(tree, feature_names, target_names, spacer_base=" "):
    """Produce psuedo-code for decision tree.
    tree -- scikit-leant DescisionTree.
    feature_names -- list of feature names.
    target_names -- list of target (output) names.
    spacer_base -- used for spacing code (default: "    ").
    """
    left      = tree.tree_.children_left
    right     = tree.tree_.children_right
    threshold = tree.tree_.threshold
    features  = [feature_names[i] for i in tree.tree_.feature]
    value = tree.tree_.value

    def recurse(left, right, threshold, features, node, depth):
        spacer = spacer_base * depth
        if (threshold[node] != -2):
            print((spacer + "if " + features[node] + " <= " + str(threshold[node]) + " :"))
#            print(spacer + "if ( " + features[node] + " <= " + str(threshold[node]) + " ) :")
            if left[node] != -1:
                    recurse(left, right, threshold, features, left[node], depth+1)
            print(( "" + spacer +"else :"))
            if right[node] != -1:
                    recurse(left, right, threshold, features, right[node], depth+1)
       #     print(spacer + "")
        else:
            target = value[node]
            for i, v in zip(np.nonzero(target)[1],  target[np.nonzero(target)]):
                target_name = target_names[i]
                target_count = int(v)
                print((spacer + "return " + str(target_name) +  " ( " + str(target_count) + ' examples )"'))

    recurse(left, right, threshold, features, 0, 0)



'''
class META_DB_CLASS(object):
   # Create Meta database to store infos on the tables : csv, zip, HFS, Postgres
ALL_DB['japancoupon']= {}
ALL_DB['japancoupon']['schema']=    df_schema
ALL_DB['japancoupon']['table_uri']= df_schema
ALL_DB['japancoupon']['table_columns']= df_schema


   def __init__(self, db_file='ALL_DB_META.pkl') :
     if db_file.find('.pkl') != -1 :
      self.filename= db_file
      self.db= util.load(db_file, isabsolutpath=1)

   def db_add(self, dbname ):
     self.db[dbname]= {}    # util.np_dictordered_create()

   def db_update_item(self, dbname, itemlistname='table_uri/schema/table_columns', itemlist=[]):
     self.db[dbname][itemlistname]=  itemlist

   def db_save(self, filename='') :
     if filename== '' :
        util.save(self.db, self.filename, isabsolutpath=1)
     else :
        self.filename= filename
        util.save(self.filename)

   def db_print_item(self):
       pass

meta_db= META_DB_CLASS( in1+'ALL_DB_META.pkl')

'''


#---------------------Execute rules on State Matrix --------------------------------
class sk_stateRule() :
  ''' Calculate Rule(True/False) based on State and Trigger
      Allow to add function to class externally                     '''
  def __init__(self, state, trigger, colname=[]):
    self.lrule= np.empty((3,20), dtype=np.object)  ;    self.nrule=20
    sh= np.shape(state);     self.tmax= sh[0] ;         self.nstate= sh[1]
    sh= np.shape(trigger);   self.ktrigger= sh[0] ;     self.ntrigger= sh[1]

    if len(colname) > 1 :     self.colname= colname
    else :                    self.colname= [ 'a'+str(i) for i in range(0, self.nstate ) ]

    self.state=    util.np_torecarray(state,  self.colname)
    self.trigger=  util.np_torecarray(trigger,  self.colname)


  def addrule(self,rulefun, name='', desc='') :
    kid= util.findnone(self.lrule[0,:])
    kid2= util.find(name, self.lrule[1,:])
    if kid2 != -1  and name != '' :     print('Name already exist !')
    else :
     if kid== -1 :
       lrule= util.np_addcolumn(self.lrule, 50)
       kid= self.nrule

     try :
       test= rulefun(self.state, self.trigger, 1)
       self.lrule[0,kid]= copy.deepcopy(rulefun)
       self.lrule[1,kid]= name
       self.lrule[2,kid]= desc
     except  ValueError as e:
      print(('Error with the function'+str(e)))


  def eval(self,idrule, t, ktrig=0) :
    if isinstance(idrule, str) :   #Evaluate by name
        kid= util.find(idrule, self.lrule[1,:])
        if kid !=-1 :         return self.lrule[0, kid](self.state, self.trigger, t)
        else:                 print(('cannot find '+idrule))
    else:
        return self.lrule[0, idrule](self.state, self.trigger, t)

  def help(self) :
    '''
s1= np.arange(5000).reshape((1000, 5))
trig1= np.ones((1,5))
state1= sk_stateRule(aa, trig1, ['drawdown','ma100d','ret10d','state_1','state_2'] )

def fun1(s, tr,t):
  return  s.drawdown[t] < tr.drawdown[0] and  s.drawdown[t] < tr.drawdown[0]

def fun2(s, tr,t):
 return  s.drawdown[t] > tr.drawdown[0] and  s.drawdown[t] < tr.drawdown[0]

state1.addrule(fun1, 'rule6')
state1.addrule(fun2, 'rule5')

state1.eval(idrule=0,t=5)

state1.eval(idrule=1,t=5)

state1.eval(idrule='rule5',t=6)

util.save_obj(state1, 'state1')

np.shape(aa2)

aa2= util.np_torecarray(aa,  ['drawdown','a2','a3','a4','a5'])

util.find(5.0, aa2[0])

recordarr = np.rec.array([(1,2.,7),(2,3.,5)],
                   dtype=[('col1', 'f8'),('col2', 'f8'), ('col3', 'f8')])
recordarr.col3[0]

state1= stateRule(np.ones((100,10)), np.ones((1,10)))

col= aa2.a2

'''



'''

def (X):
    return X[:, 1:]

def drop_first_component(X, y):
    "" Create a pipeline with PCA and the column selector and use it to transform the dataset. ""
    pipeline = make_pipeline( PCA(), FunctionTransformer(all_but_first_column))
    X_train, X_test, y_train, y_test = train_test_split(X, y)
    pipeline.fit(X_train, y_train)
    return pipeline.transform(X_test), y_test


'''


############################################################################
# ---------------------             --------------------
'''
Symbolic Regression:

http://gplearn.readthedocs.io/en/latest/examples.html#example-2-symbolic-tranformer


!pip install gplearn


x0 = np.arange(-1, 1, 1/10.)
x1 = np.arange(-1, 1, 1/10.)
x0, x1 = np.meshgrid(x0, x1)
y_truth = x0**2 - x1**2 + x1 - 1

ax = plt.figure().gca(projection='3d')
ax.set_xlim(-1, 1)
ax.set_ylim(-1, 1)
surf = ax.plot_surface(x0, x1, y_truth, rstride=1, cstride=1,
                       color='green', alpha=0.5)
plt.show()

import gplearn as gp

rng = gp.check_random_state(0)
boston = load_boston()
perm = rng.permutation(boston.target.size)
boston.data = boston.data[perm]
boston.target = boston.target[perm]

'''

# ------¨Pre-Processors --------------------------------------------------------------
'''
One-Hot: one column per category, with a 1 or 0 in each cell for if the row contained that column’s category
Binary: first the categories are encoded as ordinal, then those integers are converted into binary code,
then the digits from that binary string are split into separate columns.  This encodes the data in fewer dimensions that one-hot,
 but with some distortion of the distances.

http://www.kdnuggets.com/2015/12/beyond-one-hot-exploration-categorical-variables.html

import category_encoders as ce

encoder = ce.BackwardDifferenceEncoder(cols=[...])
encoder = ce.BinaryEncoder(cols=[...])
encoder = ce.HashingEncoder(cols=[...])
encoder = ce.HelmertEncoder(cols=[...])
encoder = ce.OneHotEncoder(cols=[...])
encoder = ce.OrdinalEncoder(cols=[...])
encoder = ce.SumEncoder(cols=[...])
encoder = ce.PolynomialEncoder(cols=[...])

Best is Binary Encoder

Splice
Coding	Dimensionality	Avg. Score	Elapsed Time
14	Ordinal	61	0.68	5.11
17	Sum Coding	3465	0.92	25.90
16	Binary Encoded	134	0.94	3.35
15	One-Hot Encoded	3465	0.95	2.56


Value ---> Hash  (limited in value)
      ---> Reduce Dimensionality of the Hash

def hash_fn(x):
tmp = [0for_inrange(N)]
for val in x.values:
tmp[hash(val)% N] += 1
return pd.Series(tmp, index=cols)

cols = ['col_%d'% d for d in range(N)]
X = X.apply(hash_fn, axis=1)


@profile(precision=4)
def onehot():
    X, _, _ = get_mushroom_data()
    print(X.info())
    enc = ce.OneHotEncoder()
    enc.fit(X, None)
    out = enc.transform(X)
    print(out.info())
    del enc, _, X, out

def binary(X):
    enc = ce.BinaryEncoder()
    enc.fit(X, None)
    out = enc.transform(X)
    print(out.info())
    del enc, _, X, out

enc = ce.OneHotEncoder()
X_bin = enc.fit_transform(X)

import matplotlib.pyplot as plt
import category_encoders as ce
from examples.source_data.loaders import get_mushroom_data, get_cars_data, get_splice_data


'''



############################################################################






###################################################################################################################
############################ UNIT TEST ############################################################################
if __name__ == '__main__' :
  import argparse;    ppa = argparse.ArgumentParser()       # Command Line input
  ppa.add_argument('--do', type=str, default= 'action',  help='test / test02')
  arg = ppa.parse_args()



if __name__ == '__main__' and arg.do == "test":
 print(__file__)
 try :
  import util; UNIQUE_ID= util.py_log_write( DIRCWD + '/aapackage/ztest_log_all.txt', "datanalysis")


  ##########################################################################################################
  import numpy as np, pandas as pd, scipy as sci
  import datanalysis as da; print(da)

  vv  =   np.random.rand(1,10)
  mm  =   np.random.rand(100,5)
  df= pd.DataFrame(mm, columns=[ 'a', 'b', 'c', 'd', 'e'])


  Xtrain= mm ; Ytrain= np.random.randint(0, 1, len(Xtrain))
  clfrf= da.sk_tree(Xtrain=Xtrain, Ytrain=Ytrain, nbtree=2, maxdepth=5, isprint1=0)
  print(clfrf)



  da.sk_cluster(Xmat=Xtrain, method='kmeans', kwds={'n_clusters': 5 } )






  ##########################################################################################################
  print("\n\n"+ UNIQUE_ID +" ###################### End:" + arrow.utcnow().to('Japan').format() + "###########################") ; sys.stdout.flush()
 except Exception as e : util.py_exception_print()




'''
  try :

  except Exception as e: print(e)

'''










