"""
Methods for reading the data from csv, xls, DB's etc, and writting to csv 
and does some preprocessing
"""

import copy
import itertools
import math
import os
import re
import sys
from calendar import isleap
from collections import OrderedDict
from datetime import datetime, timedelta

# import arrow
import numpy as np
import pandas as pd
import requests
import scipy as sci
from dateutil.parser import parse

import sklearn as sk
from sklearn.model_selection import train_test_split

# from attrdict import AttrDict as dict2
# from tabulate import tabulate


######## Read file and extract data pattern:  .csv,  .txt, .xls  ##################################
############## Excel processing #######################################################################


def xl_setstyle(file1):
    """
   http://openpyxl.readthedocs.io/en/default/styles.html#cell-styles-and-named-styles
  import openpyxl.styles.builtins  as bi
  import openpyxl.styles.builtins

  col = ws.column_dimensions['A']
  col.font = Font(bold=True)

  for cell in ws['A'] + ws[1]:
    cell.style = 'data01'

  bd = Side(style='thick', color="000000")
  highlight.border = BORDER_NONE
  from openpyxl.styles import
 """
    import openpyxl as xl

    try:
        from openpyxl.cell import get_column_letter as gcol
        from openpyxl.cell import column_index_from_string as gstr
    except:
        from openpyxl.utils import get_column_letter as gcol
        from openpyxl.utils import column_index_from_string as gstr
    from openpyxl import load_workbook

    try:
        wb = load_workbook(file1)
    except:
        print(("File protected: " + file1))

    from openpyxl.styles import named_styles, Font, Border, Side

    data01 = named_styles.NamedStyle(name="data01")
    data01.font = Font(name="Arial", bold=False, size=8, italic=False, color="000000")

    ws_names = wb.get_sheet_names()
    for wsname in ws_names:
        ws = wb.get_sheet_by_name(wsname)
        # cr= ws.get_squared_range('A1')
        for coli in range(1, 20):

            #  issues, does not work
            col = ws.column_dimensions[gcol(coli)]
            col.font = Font(name="Arial", bold=False, size=8, italic=False, color="000000")
    wb.save(file1)


def xl_val(ws, colj, rowi):
    from openpyxl.utils import get_column_letter as gcol

    try:
        return ws[gcol(colj) + str(rowi)].value
    except:
        return None


def xl_get_rowcol(ws, i0, j0, imax, jmax):
    rmat = []
    for j in range(j0, j0 + jmax):
        if isnull(xl_val(ws, j, i0)) and isnull(
            xl_val(ws, j + 1, i0)
        ):  # Stop conditions : column j+1 is empty
            return rmat
        rmatj = [xl_val(ws, j, i) for i in range(i0, i0 + imax)]  # add all the rows
        rmat.append(rmatj)
    return rmat


'''
def xl_getschema(dirxl="", filepattern="*.xlsx", dirlevel=1, outfile=".xlsx"):
    """Take All excel in a folder and provide Table, Column Schema, type into master file
 """

    def xl_is_data_block_start(ws, i, colj):
        # Check if block of data start:  i,j+1,j+2   i+1 i+2 should be NO empty
        if isnull(xl_val(ws, colj, i)):
            return False
        if (
            not isnull(xl_val(ws, colj, i))
            and not isnull(xl_val(ws, colj + 2, i))
            and not isnull(xl_val(ws, colj + 1, i + 1))
            and not isnull(xl_val(ws, colj + 2, i + 2))
        ):
            return True

    def xl_find_start_block(ws, colmin, rowmin, colmax, rowmax):
        # locate 1st non blank cell
        for j in range(colmin, colmax):
            for i in range(rowmin, rowmax):
                if xl_is_data_block_start(ws, i, j):
                    return i, j
        return rowmin, colmin

    def _xl_getschema(file1="*.xlsx", output_istack=1):
        """ 1 file Get dataframe Schema: Table, Colnam, val1, val2 from Excel xlsx
  """
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter as gcol

        # filepath1, filen1=  util.os_split_dir_file(file1)
        filepath1 = file1
        filen1 = util.os_file_getname(file1)
        filetype1 = filen1[-4:]
        try:
            wb = load_workbook(file1, read_only=True)
        except:
            print(("File protected: " + file1))
            return None

        print(file1)
        ws_names = wb.get_sheet_names()
        df_list = []
        for wsname in ws_names:
            ws = wb.get_sheet_by_name(wsname)
            if ws.max_column > 2 and ws.max_row > 3:  # Minimum table size

                print("    Reading: ", ws.title, ws.max_row, ws.max_column)
                i0, j0 = ws.min_row, ws.min_column
                if not isnull(xl_val(ws, i0 + 1, j0 + 1)):
                    i1, j1 = i0 + 2, ws.max_column
                    lefttop, rightdown = gcol(j0) + str(i0), gcol(j1) + str(i1)
                    xmat = np.array([[i.value for i in j] for j in ws[lefttop:rightdown]]).T

                else:
                    # locate 1st non blank cell
                    max_col, max_row = 50, 100
                    i, j = xl_find_start_block(ws, ws.min_column, ws.min_row, max_col, max_row)
                    xmat = np.array(xl_get_rowcol(ws, i, j, imax=3, jmax=1000))

                cols = ["uri", "filetype", "file", "table", "column", "val1", "val2"]
                nlen = xmat.shape[0]
                xfile = np.array([[filepath1, filetype1, filen1, str(wsname)]] * nlen)
                datai = np.column_stack((xfile, xmat))

                df_list.append(pd.DataFrame(datai, columns=cols))

        if output_istack:
            return pd_stack_dflist(df_list)
        else:
            return df_list

    filelist1 = util.os_file_listall(dirxl, filepattern, dirlevel=dirlevel)
    df0 = None
    for i, file1 in enumerate(filelist1[2]):
        dfi = _xl_getschema(file1=file1, output_istack=1)
        if df0 is None:
            df0 = dfi
        else:
            try:
                df0 = df0.append(dfi, ignore_index=True)
            except:
                print(("Error on: " + file1))

    if df0 is not None:
        util.pd_toexcel(df0, outfile=outfile, sheet_name="schema", returnfile=1)
'''

############## csv processing #######################################################################


def csv_dtypes_getdict(df=None, csvfile=None):
    if df is not None:
        return df.dtypes.to_dict()
    elif csvfile is not None:
        df = pd.read_csv(csvfile, nrows=1000)
        return df.dtypes.to_dict()


def csv_fast_processing():
    """
   http://word.bitly.com/post/74069870671/optimizing-text-processing

import sys
from collections import defaultdict
OUT_FILES = defaultdict(dict)

open_outfiles()  # open all files I could possibly need

for line in sys.stdin:
    # 1. parse line for account_id and metric_type
    key = line.split(',')
    account_id = key[ACCOUNT_ID_INDEX][1:] # strip leading quote

    # 2. write to appropriate file for account_id and metric_type
    OUT_FILES[account_id][key[METRIC_TYPE_INDEX]].write(line)

   close_outfiles()  # close all the files we opened

   """


def csv_col_schema_toexcel(
    dircsv="",
    filepattern="*.csv",
    outfile=".xlsx",
    returntable=1,
    maxrow=5000000,
    maxcol_pertable=90,
    maxstrlen="U80",
):
    """Take All csv in a folder and provide Table, Column Schema, type
 str(df[col].dtype)  USE str always, otherwise BIG Issue

METHOD FOR Unicode / ASCII issue
1. Decode early:  Decode to <type 'unicode'> ASAP
    df['PREF_NAME']=       df['PREF_NAME'].apply(to_unicode)
2. Unicode everywhere
3. Encode late :f = open('/tmp/ivan_out.txt','w')
                f.write(ivan_uni.encode('utf-8'))
 """

    import gc

    filelist1 = util.os_file_listall(dircsv, filepattern, dirlevel=0)
    csvlist = filelist1[0]  # name
    nsheet = len(csvlist)
    nrows_schema = maxcol_pertable * nsheet

    type_guess_all = {}  # Dict of all new types
    schema = np.array(util.str_empty_string_array(15, nrows_schema), dtype=maxstrlen)
    jj = 0
    for ii, tablei in enumerate(csvlist):
        print((ii, filelist1[2][ii]))
        df = pd.read_csv(dircsv + "/" + tablei + ".csv", nrows=maxrow)  # , encoding='UTF-8')

        df_type1 = df.dtypes.tolist()
        dfcol = df.columns.values
        # Transform to unicode EARLY
        for i, col in enumerate(dfcol):
            if str(df_type1[i]) == "object":
                df[col].fillna("", inplace=True)
                df[col] = df[col].apply(str_to_unicode)

        util.pd_toexcel(df[:100], outfile, sheet_name=tablei)
        ncol = len(df.columns.values)

        # Stats on the dataframe
        df_uniquecount = df.apply(pd.Series.nunique).values
        df_count = len(df.index) + np.zeros(ncol, dtype=np.int32)
        df_max = df.apply(pd.Series.max).values
        df_min = df.apply(pd.Series.min).values

        df_type2 = [str(type(df.iloc[0, i])) for i in range(0, ncol)]
        df_type_len = [
            df[col].map(len).max() if str(df[col].dtype) == "object" else 0
            for col in df.columns.values
        ]
        # df_quantile= df.quantile(0.9, axis=1)

        # Type guessing
        df_type_guess_dict = df.dtypes.to_dict()
        df_type_guess = [""] * ncol
        for i, x in enumerate(df_type1):
            df_type_guess[i] = df_type1[i]
            if str(x) == "object":
                ratio = float(df_uniquecount[i]) / df_count[i]
                if (
                    df_uniquecount[i] < 1000 and ratio < 0.2
                ) or ratio < 0.01:  # Less than 1% of all
                    df_type_guess[i] = "category"
            if str(x).find("int") > -1:
                if df_max[i] < 10000:
                    df_type_guess[i] = "int16"
                elif df_max[i] < 4294967296:
                    df_type_guess[i] = "int32"
            if str(x).find("float") > -1:
                if df_max[i] < 10000.0:
                    df_type_guess[i] = "float16"
                elif df_max[i] < 3.4028235e38:
                    df_type_guess[i] = "float32"

            df_type_guess_dict[dfcol[i]] = str(df_type_guess[i])  # .replace('object', 'unicode')

        type_guess_all[tablei] = df_type_guess_dict
        # Schema: Table, column name, val into NUMPY array
        schema[jj : jj + ncol, 0] = tablei  # Table name
        schema[jj : jj + ncol, 1] = df.columns.values  # Col name
        schema[jj : jj + ncol, 2] = df.iloc[0, :].values  # Col 1st value
        schema[jj : jj + ncol, 3] = df_type1  # type Pandas
        schema[jj : jj + ncol, 4] = df_type2  # type numpy
        schema[jj : jj + ncol, 5] = df_type_guess  # Guess type (reduction memory)
        schema[jj : jj + ncol, 6] = df_type_len  # String length
        schema[jj : jj + ncol, 7] = df_uniquecount
        schema[jj : jj + ncol, 8] = df_count
        schema[jj : jj + ncol, 9] = df_max
        schema[jj : jj + ncol, 10] = df_min
        # schema[jj:jj + ncol, 11]= df_quantile
        schema[jj : jj + ncol, 12] = filelist1[2][ii]
        jj = jj + ncol + 1
        gc.collect()

    schema = pd.DataFrame(
        schema,
        columns=[
            "table",
            "column",
            "val",
            "df_type",
            "np_type",
            "guess_type",
            "type_len",
            "nb_unique",
            "nb",
            "max",
            "min",
            "quantile_90",
            "uri",
            "col14",
            "col15",
        ],
    )
    util.pd_toexcel(schema, outfile, sheet_name="schema")
    util.save(type_guess_all, outfile.replace(".xlsx", "") + "_type_guess.pkl")
    util.save(schema, outfile.replace(".xlsx", "") + "_schema.pkl")
    print(outfile)
    print(("_type_guess.pkl", "_schema.pkl"))
    if returntable:
        return schema, type_guess_all


def csv_col_get_dict_categoryfreq(
    dircsv, filepattern="*.csv", category_cols=[], maxline=-1, fileencoding="utf-8"
):
    """ Find Category Freq in large CSV Transaction Column   """
    start = datetime.now()
    from collections import defaultdict  # Ordersed dict

    catdict = {}  # List of Category to fill when parsing
    for colid in category_cols:
        catdict[colid] = defaultdict(int)

    csvlist = util.os_file_listall(dircsv, filepattern)
    csvlist = csvlist[:, 2]
    kk = 0
    for jj, filecsv in enumerate(csvlist):
        maxline = 1000000000 if maxline == -1 else maxline
        for i, line in enumerate(open(filecsv, encoding=fileencoding)):
            if i > 0:
                kk += 1
                ll = line.split(",")
                for colid in category_cols:
                    catdict[colid][ll[colid]] += 1

                if kk % 10000000 == 0:
                    print(i, datetime.now() - start)  # progress
            if kk > maxline:
                break

    return catdict


def csv_row_reduce_line(fromfile, tofile, condfilter, catval_tokeep, header=True, maxline=-1):
    """ Reduce Data Row by filtering on some Category
    file_category=  in1+ "offers.csv"
    ncol= 8
    catval_tokeep=[ {} for i in xrange(0, ncol)]
    for i, line in enumerate(open(file_category)):
      ll=  line.split(",")
      catval_tokeep[3][  ll[1] ]  = 1  # Offer_file_col1 --> Transact_file_col_4
      catval_tokeep[4][  ll[3] ] =  1  # Offer_file_col3 --> Transact_file_col_4

  def condfilter(colk, catval_tokeep) :
    if colk[3] in catval_tokeep[3] or colk[4] in catval_tokeep[4]: return True
    else: return False
  """
    start = datetime.now()
    maxline = 1000000000 if maxline == -1 else maxline

    with open(tofile, "wb") as outfile:
        jj_new = 0
        i = 0
        with open(fromfile) as f:
            if header:  # 1st line
                line = next(f)
                outfile.write(ll)
            for i, line in enumerate(f):  # go through transactions file and reduce
                ll = line.split(",")
                if condfilter(ll, catval_tokeep):  # Condition  Filter : if category in offers dict
                    outfile.write(",".join(line))
                    jj_new += 1

                # progress
                if i % 5000000 == 0:
                    print(i, jj_new, datetime.now() - start)
                if i > maxline:
                    break
    print(i, jj_new, datetime.now() - start)
    """
  does not work, issue with character encoding....
      with open(fromfile, 'r') as f :
     with csv.reader(f,  delimiter=',' ) as reader :
      for ll in reader:
  """


def csv_analysis():
    """
   https://csvkit.readthedocs.io/en/540/tutorial/1_getting_started.html

   sudo pip install csvkit

   :return:
   """


def csv_row_reduce_line_manual(file_category, file_transact, file_reduced):
    """ Reduce Data by filtering on some Category """
    start = datetime.now()

    # Parse all categories and comps on offer in a dict
    offers_cat, offers_co = {}, {}
    for i, line in enumerate(open(file_category)):
        ll = line.split(",")
        offers_cat[ll[1]] = 1
        offers_co[ll[3]] = 1

    # open output file
    with open(file_reduced, "wb") as outfile:
        # go through transactions file and reduce
        jj_new = 0
        for i, line in enumerate(open(file_transact)):
            if i == 0:
                outfile.write(line)  # print header
            else:
                ll = line.split(",")
                if (
                    ll[3] in offers_cat or ll[4] in offers_co
                ):  # Condition  Filter : if category in offers dict
                    outfile.write(line)
                    jj_new += 1

            # progress
            if i % 5000000 == 0:
                print(i, jj_new, datetime.now() - start)
    print(i, jj_new, datetime.now() - start)


def csv_row_mapreduce(dircsv="", outfile="", type_mapreduce="sum", nrow=1000000, chunk=5000000):
    """Take All csv in a folder and provide Table, Column Schema"""
    filelist1 = util.os_file_listall(dircsv, "*.csv")
    csvlist = filelist1[:, 0]

    colname = ""
    kchunk = int(nrow / chunk)

    dfout = pd.DataFrame([], columns=colname)
    for ii, tablei in enumerate(csvlist):
        for kk in range(0, kchunk):
            df = pd.read_csv(dircsv + "/" + tablei + ".csv", nrows=chunk, encoding="UTF-8")

    return dfout


def csv_pivotable(
    dircsv="",
    filepattern="*.csv",
    fileh5=".h5",
    leftX="col0",
    topY="col2",
    centerZ="coli",
    mapreduce="sum",
    chunksize=500000,
    tablename="df",
):
    """ return df Pivot Table from series of csv file (transfer to d5 temporary)

Edit: you can groupby/sum from the store iteratively since this "map-reduces" over the chunks:

reduce(lambda x, y: x.add(y, fill_value=0),
       (df.groupby().sum() for df in store.select('df', chunksize=50000)))

 """

    if dircsv != "":
        util.pd_h5_fromcsv_tohdfs(
            dircsv,
            filepattern=filepattern,
            tofilehdfs=fileh5,
            tablename=tablename,
            chunksize=chunksize,
        )

    store = pd.HDFStore(fileh5)
    if mapreduce == "sum":
        pivot0 = None
        for df in store.select(tablename, chunksize=chunksize):
            if pivot0 is None:
                pivot0 = pd.DataFrame.pivot_table(
                    df, values=centerZ, index=[leftX], columns=[topY], aggfunc=np.sum, fill_value=0
                )
            else:
                pivot_i = pd.DataFrame.pivot_table(
                    df, values=centerZ, index=[leftX], columns=[topY], aggfunc=np.sum, fill_value=0
                )
                pivot0 = pd.concat([pivot0, pivot_i]).groupby(level=0).sum()

    if mapreduce == "count":
        pass

    return pivot0


def csv_bigcompute():
    pass


######################## DB related items #######################################################################
def db_getdata():
    pass


def db_sql():
    pass


def db_meta_add(
    metadb, dbname, new_table=("", []), schema=None, df_table_uri=None, df_table_columns=None
):
    """ Create Meta database to store infos on the tables : csv, zip, HFS, Postgres
ALL_DB['japancoupon']= {}
ALL_DB['japancoupon']['schema']=    df_schema
ALL_DB['japancoupon']['df_table_uri']= df_schema_dictionnary
ALL_DB['japancoupon']['df_table_columns']= df_schema_dict
        DBname, db_schema, db_table_uri, db_table_columns(dict_table->colum_list),
   """

    def pd_df_todict(df, colkey="table", firstelt=True):
        df1 = df.drop_duplicates(colkey).reset_index(level=0, drop=True)
        dict0 = {}
        for i in range(len(df)):
            id0 = df.iloc[i, 0]
            val0 = df.iloc[i, 1]
            if id0 != "":
                dict0.setdefault(id0, [])
                if firstelt:
                    dict0[id0] = val0
                else:
                    dict0[id0].append(val0)
        return dict0

    if schema is not None:
        metadb[dbname]["schema"] = schema
        metadb[dbname]["table_columns"] = pd_df_todict(schema[["table", "column"]])
        metadb[dbname]["table_uri"] = pd_df_todict(schema[["table", "uri"]])

    elif df_table_uri is not None:
        metadb[dbname]["table_uri"] = pd_df_todict(df_table_uri)

    elif df_table_columns is not None:
        metadb[dbname]["table_columns"] = pd_df_todict(
            df_table_columns, onlyfirstelt=False
        )  # table, colname, dtype

    if new_table[0] != "":
        metadb[dbname][new_table[0]] = new_table[1]

    return metadb


def db_meta_find(ALLDB, query="", filter_db=[], filter_table=[], filter_column=[]):
    """ Find string in all the meta table name, column
  db_meta_find(ALLDB, query='bottler', filter_db=['cokeon'],   filter_table=['table'], filter_column=['table'] )
  dbname: should be exact name
  fitler_table: partial match is ok
  fitler_column : partial name is ok
  return   (dbname, meta_table_name,  meta_table_filtered_by_row_containing query)
  """
    rs = []
    for iname, dbi in list(ALLDB.items()):  # Search in All DB
        if iname in filter_db or len(filter_db) == 0:
            for jname, tablej in list(dbi.items()):  # inside sub-meta table of DB
                if isinstance(tablej, pd.DataFrame):  # Only Dataframe
                    isok = util.str_match_fuzzy(jname, filter_table)
                    if isok or len(filter_table) == 0:
                        # print iname, jname
                        # Each column of table
                        colj = list(tablej.columns.values)
                        if len(filter_column) > 0:
                            for colf in filter_column:
                                aux = util.find_fuzzy(
                                    colf, colj
                                )  # if colf name matches partially colj, OK
                                colj = list(set(colj + aux))
                                # colj= [ colji  for colji in colj if colji in filter_column ]

                        df_new = util.pd_find(
                            tablej,
                            query,
                            col_restrict=colj,
                            isnumeric=False,
                            doreturnposition=False,
                        )
                        rs.append((iname, jname, df_new))  # DB name, meta_table, column_where_True

    return rs


######################  Pre Processing  ###############################################################


def str_to_unicode(x, encoding="utf-8"):
    if not isinstance(x, str):
        x = str(x, encoding)
        return x
    else:
        return x


def isnull(x):
    return x is None or x == ""
